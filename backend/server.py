from fastapi import FastAPI, APIRouter, HTTPException, status, Depends, File, UploadFile, Form, Request, WebSocket, WebSocketDisconnect
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import json
import calendar
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Set
import uuid
from datetime import datetime, timezone, date, timedelta
from enum import Enum
from jose import JWTError, jwt
from passlib.context import CryptContext
import secrets
import string
import random
import hashlib
import base64
import aiosmtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import io
import asyncio
import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse
import razorpay


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Email Configuration
SMTP_HOST = os.environ.get('SMTP_HOST')
SMTP_PORT = int(os.environ.get('SMTP_PORT', 587))
SMTP_USER = os.environ.get('SMTP_USER')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD')
SMTP_FROM = os.environ.get('SMTP_FROM')
SMTP_BCC = os.environ.get('SMTP_BCC')

# Razorpay Configuration
RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET')
RAZORPAY_WEBHOOK_SECRET = os.environ.get('RAZORPAY_WEBHOOK_SECRET')

# Initialize Razorpay client
razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

# Create the main app without a prefix
app = FastAPI(title="Elevate - Payroll Management System", version="1.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# WebSocket Connection Manager for Real-time Notifications
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}
    
    async def connect(self, websocket: WebSocket, user_id: str):
        await websocket.accept()
        self.active_connections[user_id] = websocket
        logging.info(f"WebSocket connected for user: {user_id}")
    
    def disconnect(self, user_id: str):
        if user_id in self.active_connections:
            del self.active_connections[user_id]
            logging.info(f"WebSocket disconnected for user: {user_id}")
    
    async def send_notification(self, user_id: str, message: dict):
        if user_id in self.active_connections:
            try:
                await self.active_connections[user_id].send_json(message)
            except Exception as e:
                logging.error(f"Error sending notification to {user_id}: {e}")
                self.disconnect(user_id)
    
    async def broadcast_to_role(self, role: str, message: dict):
        """Broadcast notification to all connected users with specific role"""
        disconnected = []
        for user_id, websocket in self.active_connections.items():
            if user_id.startswith(f"{role}:"):
                try:
                    await websocket.send_json(message)
                except Exception as e:
                    logging.error(f"Error broadcasting to {user_id}: {e}")
                    disconnected.append(user_id)
        
        # Clean up disconnected users
        for user_id in disconnected:
            self.disconnect(user_id)

manager = ConnectionManager()

# Helper function to send notification via WebSocket
async def send_realtime_notification(notification_dict: dict):
    """Send notification to connected users via WebSocket"""
    try:
        recipient_role = notification_dict.get('recipient_role')
        recipient_id = notification_dict.get('recipient_id')
        
        message = {
            'type': 'new_notification',
            'data': notification_dict
        }
        
        if recipient_id:
            # Send to specific user
            await manager.send_notification(f"user:{recipient_id}", message)
        elif recipient_role == "admin":
            # Broadcast to all admin connections
            await manager.broadcast_to_role("admin", message)
        elif recipient_role == "employee":
            # If no specific recipient_id, broadcast to all employees
            await manager.broadcast_to_role("employee", message)
    except Exception as e:
        logging.error(f"Error sending realtime notification: {e}")

# Enums
class EmployeeStatus(str, Enum):
    ACTIVE = "active"
    RESIGNED = "resigned"
    TERMINATED = "terminated"

class Gender(str, Enum):
    MALE = "male"
    FEMALE = "female"
    OTHER = "other"

class MaritalStatus(str, Enum):
    SINGLE = "single"
    MARRIED = "married"
    DIVORCED = "divorced"
    WIDOWED = "widowed"

class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    ADMIN = "admin"
    EMPLOYEE = "employee"

# JWT Configuration
SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    raise ValueError("SECRET_KEY environment variable is required but not set. Please configure a secure secret key.")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 120  # 2 hours session duration
REFRESH_TOKEN_EXPIRE_DAYS = 7

# Password hashing - using pbkdf2_sha256 to avoid bcrypt initialization issues
pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
security = HTTPBearer()

# Helper functions for datetime serialization
def prepare_for_mongo(data):
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, date) and not isinstance(value, datetime):
                data[key] = value.isoformat()
            elif isinstance(value, datetime):
                data[key] = value.isoformat()
            elif isinstance(value, dict):
                data[key] = prepare_for_mongo(value)
            elif isinstance(value, list):
                data[key] = [prepare_for_mongo(item) if isinstance(item, dict) else item for item in value]
    return data

def prepare_from_mongo(data):
    """Convert MongoDB document to JSON-serializable format"""
    if isinstance(data, dict):
        # Remove MongoDB _id field and convert ObjectId to string if needed
        if '_id' in data:
            del data['_id']
        
        # Convert any remaining ObjectId fields to strings
        for key, value in data.items():
            if hasattr(value, '__class__') and value.__class__.__name__ == 'ObjectId':
                data[key] = str(value)
            elif isinstance(value, dict):
                data[key] = prepare_from_mongo(value)
            elif isinstance(value, list):
                data[key] = [prepare_from_mongo(item) if isinstance(item, dict) else item for item in value]
    return data

# Authentication utility functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def create_refresh_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def generate_otp():
    return ''.join(secrets.choice(string.digits) for _ in range(6))

def generate_random_pin():
    """Generate a random 4-digit PIN"""
    return ''.join(random.choices(string.digits, k=4))

def generate_random_pin():
    """Generate a random 4-digit PIN"""
    return ''.join(random.choices(string.digits, k=4))

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        role: str = payload.get("role")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username, role=role)
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"username": token_data.username})
    if user is None:
        raise credentials_exception
    return User(**user)

def require_role(required_role: str):
    def role_checker(current_user = Depends(get_current_user)):
        if current_user.role != required_role:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions"
            )
        return current_user
    return role_checker


# Store OTPs temporarily (in production, use Redis or similar)
active_otps = {}

# Pydantic Models
class BankInfo(BaseModel):
    bank_name: Optional[str] = ""
    account_number: Optional[str] = ""
    ifsc_code: Optional[str] = ""
    branch: Optional[str] = ""

class SalaryStructure(BaseModel):
    # DEPRECATED: Legacy hardcoded fields (for backward compatibility)
    # These will be migrated to salary_components
    basic_salary: float = 0
    house_rent_allowance: float = 0
    medical_allowance: float = 0
    leave_travel_allowance: float = 0
    conveyance_allowance: float = 0
    performance_incentive: float = 0
    other_benefits: float = 0
    hra: Optional[float] = 0
    travel_allowance: Optional[float] = 0
    food_allowance: Optional[float] = 0
    internet_allowance: Optional[float] = 0
    special_allowance: Optional[float] = 0
    pf_employee: float = 0
    pf_employer: float = 0
    esi_employee: float = 0
    esi_employer: float = 0
    professional_tax: float = 0
    tds: float = 0
    loan_deductions: float = 0
    others: float = 0
    
    # NEW: Component-based salary structure
    # This is a list of component assignments with employee-specific values
    salary_components: Optional[List[dict]] = []  # List of {"component_id": "", "amount": 0, "component_name": ""}
    use_component_based_salary: bool = False  # Flag to indicate if using new system


# New model for component-based salary assignment
class EmployeeSalaryComponent(BaseModel):
    """Employee-specific salary component with amount"""
    component_id: str  # Reference to SalaryComponent
    component_name: str  # Cached for quick access
    component_type: str  # Cached: earnings, deductions, benefits, reimbursements
    category: str  # Cached: Basic, HRA, etc.
    name_in_payslip: str  # What appears on payslip
    amount: float  # Employee-specific amount
    is_active: bool = True

# Authentication Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str  # For admin: 'admin', for employee: employee_id, for super_admin: email
    email: Optional[str] = None
    role: UserRole
    company_id: Optional[str] = None  # None for super_admin, company UUID for admin/employee
    employee_id: Optional[str] = None  # Only for employee users
    hashed_password: Optional[str] = None  # For admin and super_admin
    pin: Optional[str] = None  # For employee
    is_active: bool = True
    last_login: Optional[datetime] = None
    last_login_ip: Optional[str] = None  # IP address of last login
    last_login_device: Optional[str] = None  # Device/User-Agent of last login
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LoginHistory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    username: str
    login_time: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    ip_address: Optional[str] = None
    device_name: Optional[str] = None
    pc_name: Optional[str] = None
    location: Optional[dict] = None  # {city, region, country, latitude, longitude}




# Subscription Access Check
async def check_subscription_access(current_user: User = Depends(get_current_user)):
    """
    Check if the user's company has an active subscription
    Allows super_admin to bypass this check
    """
    # Super admin always has access
    if current_user.role == "super_admin":
        return current_user
    
    # Get company subscription info
    company = await db.companies.find_one({"company_id": current_user.company_id}, {"_id": 0})
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found"
        )
    
    subscription_info = company.get("subscription_info", {})
    status_str = subscription_info.get("status", "trial")
    
    # Check if trial is still valid
    if status_str == "trial":
        trial_end_date = subscription_info.get("trial_end_date")
        if trial_end_date:
            trial_end = datetime.fromisoformat(trial_end_date)
            now = datetime.now(timezone.utc)
            
            if now >= trial_end:
                raise HTTPException(
                    status_code=status.HTTP_402_PAYMENT_REQUIRED,
                    detail="Trial period has expired. Please subscribe to continue using the service."
                )
    
    # Check subscription status
    if status_str in ["expired", "cancelled", "payment_failed"]:
        raise HTTPException(
            status_code=status.HTTP_402_PAYMENT_REQUIRED,
            detail=f"Subscription {status_str}. Please renew your subscription to continue using the service."
        )
    
    # Allow access if status is active or trial
    if status_str not in ["active", "trial", "created"]:
        raise HTTPException(
            status_code=status.HTTP_402_PAYMENT_REQUIRED,
            detail="Subscription inactive. Please contact support."
        )
    
    return current_user


# Company Filter Helper Functions (defined after User model)
def get_company_filter(current_user: User = Depends(get_current_user)) -> dict:
    """
    Get company filter for database queries based on user role.
    - Super admin: No filter (can see all companies)
    - Admin/Employee: Filter by their company_id
    """
    if current_user.role == "super_admin":
        return {}
    
    if not current_user.company_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User does not belong to any company"
        )
    
    return {"company_id": current_user.company_id}


def require_admin_or_super_admin(current_user: User = Depends(get_current_user)):
    """Dependency to ensure user is admin or super admin"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return current_user


# Multi-Tenancy Models
class CompanySettings(BaseModel):
    working_days_config: Optional[dict] = None
    leave_policies: Optional[dict] = None
    default_working_hours: Optional[dict] = None


class SubscriptionInfo(BaseModel):
    plan: str = "basic"  # basic, professional, enterprise
    plan_id: Optional[str] = None  # MongoDB SubscriptionPlan reference
    start_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    end_date: Optional[str] = None
    trial_end_date: Optional[str] = None
    status: str = "trial"  # trial, active, expired, cancelled, payment_failed
    razorpay_subscription_id: Optional[str] = None
    razorpay_plan_id: Optional[str] = None  # Razorpay Plan ID
    last_payment_date: Optional[str] = None
    next_billing_date: Optional[str] = None
    auto_renew: bool = True


class Company(BaseModel):
    company_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_name: str
    company_logo_url: Optional[str] = None
    contact_email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None
    admin_user_id: Optional[str] = None  # Reference to first admin user
    settings: Optional[CompanySettings] = None
    status: str = "active"  # active, inactive
    subscription_info: Optional[SubscriptionInfo] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class CompanyCreate(BaseModel):
    company_name: str
    company_logo_url: Optional[str] = None
    contact_email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None
    admin_name: str
    admin_email: EmailStr
    admin_phone: Optional[str] = None
    settings: Optional[CompanySettings] = None


class CompanyUpdate(BaseModel):
    company_name: Optional[str] = None
    company_logo_url: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    settings: Optional[CompanySettings] = None
    status: Optional[str] = None


class Invitation(BaseModel):
    invitation_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    admin_name: str
    role: str = "admin"  # Only admin invitations for now
    status: str = "pending"  # pending, accepted, expired
    expires_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc) + timedelta(days=7))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    accepted_at: Optional[datetime] = None


class InvitationAccept(BaseModel):
    token: str
    password: str


# Subscription Plan Models
class PlanFeatures(BaseModel):
    # Limits
    employee_limit: int  # -1 for unlimited
    admin_users_limit: int  # -1 for unlimited
    trial_days: int = 0  # 0 for no trial, 30 for free plan
    
    # Core Features
    employee_database: bool = True
    payroll_processing_manual: bool = True
    payroll_processing_automated: bool = False
    payslip_generation: bool = True
    
    # Attendance & Leave
    attendance_tracking_basic: bool = True
    attendance_tracking_advanced: bool = False
    leave_management_basic: bool = True
    leave_management_advanced: bool = False
    
    # Payroll Features
    salary_structure_management: bool = False
    bank_advice_generation: bool = False
    custom_salary_components: bool = False
    bulk_employee_import: bool = False
    
    # Compliance
    compliance_reports_basic: bool = False
    compliance_reports_full: bool = False
    
    # Advanced Features
    employee_portal: bool = False
    loans_advances: bool = False
    deductions_advanced: bool = False
    event_management: bool = False
    payroll_analytics: bool = False
    multi_bank_accounts: bool = False
    notifications: bool = False
    dark_mode: bool = False
    
    # Enterprise Features
    api_access: bool = False
    white_labeling: bool = False
    custom_integrations: bool = False
    sso_security: bool = False
    custom_reports: bool = False
    audit_logs: bool = False
    sla_guarantee: bool = False
    
    # Support
    support_level: str = "email"  # email, priority_email, priority_email_chat, phone


class SubscriptionPlan(BaseModel):
    plan_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    plan_name: str
    slug: str
    description: str
    price_per_user_monthly: float  # Changed from monthly_price
    price_per_user_annual: float   # Changed from annual_price (10 months pricing)
    currency: str = "INR"
    razorpay_plan_id_monthly: Optional[str] = None
    razorpay_plan_id_annual: Optional[str] = None
    features: PlanFeatures
    is_active: bool = True
    is_trial: bool = False  # True for Free plan
    display_order: int = 1
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PlanUpdate(BaseModel):
    plan_name: Optional[str] = None
    description: Optional[str] = None
    price_per_user_monthly: Optional[float] = None
    price_per_user_annual: Optional[float] = None
    features: Optional[PlanFeatures] = None
    is_active: Optional[bool] = None


class PublicSignup(BaseModel):
    # Company info
    company_name: str
    contact_email: EmailStr
    phone: Optional[str] = None
    industry: Optional[str] = None
    country: str = "India"
    
    # Admin info
    admin_name: str
    admin_email: EmailStr
    password: str
    
    # Plan selection
    plan_id: str
    billing_cycle: str = "monthly"  # monthly or annual
    
    # Terms
    terms_accepted: bool = True


# Payment and Subscription Models
class CreateSubscriptionRequest(BaseModel):
    company_id: str
    billing_cycle: str = "monthly"  # monthly or annual


class VerifyPaymentRequest(BaseModel):
    subscription_id: str
    payment_id: str
    signature: str


class SubscriptionStatusResponse(BaseModel):
    status: str
    subscription_id: Optional[str] = None
    current_period_end: Optional[str] = None
    next_billing_date: Optional[str] = None
    amount: Optional[float] = None
    auto_renew: bool = True


class LoginRequest(BaseModel):
    username: str  # Employee ID or 'admin'
    password: Optional[str] = None  # For admin
    pin: Optional[str] = None  # For employee
    otp: Optional[str] = None  # For employee

class Token(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    user: dict

class TokenData(BaseModel):
    username: Optional[str] = None
    role: Optional[str] = None

class OTPRequest(BaseModel):
    employee_id: str

class OTPResponse(BaseModel):
    message: str
    otp: Optional[str] = None  # Only in development, remove in production

class PinUpdateRequest(BaseModel):
    employee_id: str
    new_pin: Optional[str] = None  # If None, generate random PIN

class EmployeePinChangeRequest(BaseModel):
    current_pin: str
    new_pin: str

# Leave Management Models
class LeaveRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    leave_type: str  # Annual Leave, Sick Leave, Casual Leave, etc.
    start_date: date
    end_date: date
    days: float  # Can be 0.5 for half day
    reason: str
    half_day: bool = False
    status: str = "pending"  # pending, approved, rejected, cancelled
    applied_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    approved_by: Optional[str] = None
    approved_date: Optional[datetime] = None
    rejected_by: Optional[str] = None
    rejected_date: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    cancelled_by: Optional[str] = None
    cancelled_date: Optional[datetime] = None
    cancellation_reason: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeaveRequestCreate(BaseModel):
    leave_type: str
    start_date: date
    end_date: date
    reason: str
    half_day: bool = False

class LeaveApprovalRequest(BaseModel):
    status: str  # approved, rejected
    admin_comment: Optional[str] = None  # Admin comment for both approval and rejection

class LeaveCancellationRequest(BaseModel):
    cancellation_reason: str

# Leave Entitlement Models
class LeaveBalance(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    year: int
    casual_leave_accrued: float = 0.0  # Accrued at 1.5 days/month
    casual_leave_used: float = 0.0
    casual_leave_balance: float = 0.0
    sick_leave_total: float = 7.0  # 7 days per year
    sick_leave_used: float = 0.0
    sick_leave_balance: float = 7.0
    carried_forward_leaves: float = 0.0  # From previous year, max 5
    last_accrual_date: Optional[date] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeaveEntitlementResponse(BaseModel):
    employee_id: str
    employee_name: str
    joining_date: Optional[date]
    months_of_service: int
    casual_leave_accrued: float
    casual_leave_used: float
    casual_leave_balance: float
    sick_leave_total: float
    sick_leave_used: float
    sick_leave_balance: float
    annual_leave_total: float = 0.0
    annual_leave_used: float = 0.0
    annual_leave_balance: float = 0.0
    carried_forward_leaves: float
    total_available_leaves: float

# Loan Management Models
class LoanRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    loan_type: str  # Personal Loan, Emergency Loan, Advance Salary, etc.
    amount: float
    tenure_months: int
    interest_rate: float
    monthly_emi: float
    purpose: str
    monthly_income: Optional[float] = None
    existing_loans: Optional[float] = None
    status: str = "pending"  # pending, approved, rejected, disbursed
    applied_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    approved_by: Optional[str] = None
    approved_date: Optional[datetime] = None
    rejected_by: Optional[str] = None
    rejected_date: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    disbursed_amount: Optional[float] = None
    outstanding_amount: Optional[float] = None
    paid_emis: Optional[int] = 0
    remaining_emis: Optional[int] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LoanRequestCreate(BaseModel):
    employee_id: Optional[str] = None  # Optional: for admins creating on behalf of employees
    loan_type: str
    amount: float
    tenure_months: int
    interest_rate: Optional[float] = None  # Optional: allow custom interest rate
    monthly_emi: Optional[float] = None  # Optional: allow custom EMI
    purpose: str
    monthly_income: Optional[float] = None
    existing_loans: Optional[float] = None
    guarantor_name: Optional[str] = None
    guarantor_employee_id: Optional[str] = None

class LoanApprovalRequest(BaseModel):
    status: str  # approved, rejected
    rejection_reason: Optional[str] = None
    disbursed_amount: Optional[float] = None

# Attendance Management Models
class Attendance(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    date: date
    status: str = "present"  # present, leave, half-day, holiday, weekend
    working_hours: float = 8.0  # Default 8 hours (8:30 AM - 5:30 PM, minus 1 hour lunch)
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AttendanceCorrection(BaseModel):
    date: date
    status: str  # present, leave, half-day, absent
    working_hours: float
    notes: Optional[str] = None

class AttendanceMarkRequest(BaseModel):
    employee_id: str
    date: date
    status: str  # present, absent, half-day
    working_hours: float

class LateArrival(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    date: date
    expected_check_in: str = "08:30"  # Expected check-in time (HH:MM)
    actual_check_in: str  # Actual check-in time (HH:MM)
    late_minutes: int  # Minutes late
    reason: Optional[str] = None
    recorded_by: str  # Admin who recorded this
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LateArrivalCreate(BaseModel):
    employee_id: str
    date: date
    actual_check_in: str  # Format: HH:MM
    reason: Optional[str] = None

# OT Log Models
class OTLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    date: date
    from_time: str  # Format: HH:MM (e.g., "07:00" or "18:00")
    to_time: str    # Format: HH:MM (e.g., "08:30" or "20:00")
    ot_hours: float  # Calculated from time range
    project: str  # Mandatory project name
    status: str = "pending"  # pending, approved, rejected
    notes: Optional[str] = None
    applied_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    approved_by: Optional[str] = None
    approved_date: Optional[datetime] = None
    rejected_by: Optional[str] = None
    rejected_date: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OTLogCreate(BaseModel):
    date: date
    from_time: str  # Format: HH:MM
    to_time: str    # Format: HH:MM
    project: str  # Mandatory project name
    notes: Optional[str] = None

class AdminOTLogCreate(BaseModel):
    employee_id: str  # Admin specifies which employee
    date: date
    from_time: str  # Format: HH:MM
    to_time: str    # Format: HH:MM
    project: str  # Mandatory project name
    notes: Optional[str] = None

class OTApprovalRequest(BaseModel):
    status: str  # approved, rejected
    rejection_reason: Optional[str] = None

class Employee(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    company_id: Optional[str] = None  # CRITICAL: Company ID for multi-tenancy
    name: str
    email: EmailStr
    phone: Optional[str] = None
    gender: Optional[Gender] = None
    date_of_birth: Optional[date] = None
    aadhar_number: Optional[str] = None
    pan_number: Optional[str] = None
    marital_status: Optional[MaritalStatus] = None
    address: Optional[str] = None
    
    # Job Information
    department: str
    designation: str
    date_of_joining: Optional[date] = None
    work_location: Optional[str] = None
    status: EmployeeStatus = EmployeeStatus.ACTIVE
    
    # Status Management
    resignation_date: Optional[date] = None
    termination_date: Optional[date] = None
    status_reason: Optional[str] = None
    
    # Bank Information
    bank_info: Optional[BankInfo] = None
    
    # Salary Information
    salary_structure: Optional[SalaryStructure] = None
    
    # Leave Configuration
    is_on_probation: Optional[bool] = False
    probation_end_date: Optional[date] = None
    custom_casual_leave_per_month: Optional[float] = None  # Overrides default 1.5 if set
    custom_sick_leave_per_year: Optional[float] = None  # Overrides default 7 if set
    annual_leave_days: Optional[float] = None  # Additional annual leave entitlement
    
    # Emergency Contact & Medical Information
    emergency_contact_name: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    blood_group: Optional[str] = None
    photo_url: Optional[str] = None
    
    # Timestamps
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeCreate(BaseModel):
    employee_id: str
    name: str
    email: EmailStr
    phone: Optional[str] = None
    gender: Optional[Gender] = None
    date_of_birth: Optional[date] = None
    aadhar_number: Optional[str] = None
    pan_number: Optional[str] = None
    marital_status: Optional[MaritalStatus] = None
    address: Optional[str] = None
    department: str
    designation: str
    date_of_joining: Optional[date] = None
    work_location: Optional[str] = None
    bank_info: Optional[BankInfo] = None
    salary_structure: Optional[SalaryStructure] = None
    is_on_probation: Optional[bool] = False
    probation_end_date: Optional[date] = None
    custom_casual_leave_per_month: Optional[float] = None
    custom_sick_leave_per_year: Optional[float] = None
    annual_leave_days: Optional[float] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    blood_group: Optional[str] = None

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    gender: Optional[Gender] = None
    date_of_birth: Optional[date] = None
    aadhar_number: Optional[str] = None
    pan_number: Optional[str] = None
    marital_status: Optional[MaritalStatus] = None
    address: Optional[str] = None
    department: Optional[str] = None
    designation: Optional[str] = None
    date_of_joining: Optional[date] = None
    work_location: Optional[str] = None
    status: Optional[EmployeeStatus] = None
    resignation_date: Optional[date] = None
    termination_date: Optional[date] = None
    status_reason: Optional[str] = None
    bank_info: Optional[BankInfo] = None
    is_on_probation: Optional[bool] = None
    probation_end_date: Optional[date] = None
    custom_casual_leave_per_month: Optional[float] = None
    custom_sick_leave_per_year: Optional[float] = None
    annual_leave_days: Optional[float] = None
    salary_structure: Optional[SalaryStructure] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    blood_group: Optional[str] = None

class EmployeeStatusUpdate(BaseModel):
    status: EmployeeStatus
    resignation_date: Optional[date] = None
    termination_date: Optional[date] = None
    status_reason: Optional[str] = None

class BulkDeleteRequest(BaseModel):
    employee_ids: List[str]

class NotificationCreate(BaseModel):
    title: str
    message: str
    notification_type: str  # info, success, warning, error
    category: Optional[str] = None  # For routing notifications (ot, leave, loan, etc.)
    related_id: Optional[str] = None  # ID of related entity
    recipient_id: Optional[str] = None  # If None, it's for all admins
    recipient_role: Optional[str] = None  # admin or employee
    
class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    message: str
    notification_type: str
    category: Optional[str] = None
    related_id: Optional[str] = None  # ID of related entity (leave_id, ot_id, loan_id)
    recipient_id: Optional[str] = None
    recipient_role: Optional[str] = None
    is_read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    read_at: Optional[datetime] = None


class NotificationUpdateRequest(BaseModel):
    is_read: bool

class EventCreate(BaseModel):
    title: str
    description: Optional[str] = None
    date: str  # ISO format date string
    event_type: str = "company"  # company, team, milestone, other
    
class EventUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None
    event_type: Optional[str] = None

class Event(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    date: str  # ISO format date string
    event_type: str = "company"
    created_by: str  # admin employee_id
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Payslip(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    month: int
    year: int
    gross_salary: float
    total_deductions: float
    net_salary: float
    earnings: dict
    deductions: dict
    status: str = "generated"  # generated, deleted
    generated_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_date: Optional[datetime] = None

class PayslipGenerate(BaseModel):
    month: int
    year: int
    employee_ids: Optional[List[str]] = None  # If None, generate for all active employees

# Payroll Models
class PayrollEmployee(BaseModel):
    employee_id: str
    days_worked: int
    days_in_month: int = 30  # Total days in the payroll month
    overtime_hours: float = 0
    bonus: float = 0
    adjustments: float = 0  # Can be positive or negative
    loan_deductions: float = 0  # Dynamic loan deductions for this payroll run
    tds: float = 0  # Tax Deducted at Source

class PayrollRun(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    month: int
    year: int
    employees: List[PayrollEmployee]
    total_employees: int
    total_gross: float
    total_deductions: float
    total_net: float
    status: str = "completed"  # completed, draft
    processed_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    processed_by: str  # Admin username

class PayrollRunRequest(BaseModel):
    month: int
    year: int
    employees: List[PayrollEmployee]

class DashboardStats(BaseModel):
    total_employees: int
    active_employees: int
    this_month_payroll: float
    payslips_generated: int
    upcoming_deadlines: int


# ==================== Bank Advice System Models ====================

# Company Bank Accounts Models
class CompanyBankAccount(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    account_name: str  # e.g., "Primary Salary Account"
    bank_name: str
    account_number: str
    ifsc_code: str
    branch: str
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class CompanyBankAccountCreate(BaseModel):
    account_name: str
    bank_name: str
    account_number: str
    ifsc_code: str
    branch: str
    is_active: bool = True

class CompanyBankAccountUpdate(BaseModel):
    account_name: Optional[str] = None
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    ifsc_code: Optional[str] = None
    branch: Optional[str] = None
    is_active: Optional[bool] = None

# Employee Source Mapping Models
class EmployeeSourceMapping(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str  # employee_id (e.g., ET-MUM-00001)
    company_account_id: str  # References CompanyBankAccount.id
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class EmployeeSourceMappingCreate(BaseModel):
    employee_id: str
    company_account_id: str

class BulkEmployeeSourceMappingCreate(BaseModel):
    employee_ids: List[str]
    company_account_id: str

# Bank Template Models
class BankTemplate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bank_name: str
    template_data: str  # Base64 encoded Excel file
    file_name: str
    uploaded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    uploaded_by: str  # Admin username

class BankTemplateUpload(BaseModel):
    bank_name: str
    file_name: str
    template_data: str  # Base64 encoded

# Bank Advice Generation Models
class BankAdviceGenerate(BaseModel):
    month: int
    year: int
    company_account_id: Optional[str] = None  # If None, generate for all accounts
    template_id: Optional[str] = None  # Bank template to use

class BankAdvice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    month: int
    year: int
    reference_number: str
    company_account_id: str
    template_id: Optional[str] = None  # Template used for generation
    total_amount: float
    employee_count: int
    status: str = "generated"  # generated, sent, completed
    generated_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    sent_date: Optional[datetime] = None
    completed_date: Optional[datetime] = None
    file_path: Optional[str] = None  # Path to generated file


# ==================== Salary Components Management Models ====================

class SalaryComponent(BaseModel):
    component_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str  # Multi-tenancy
    category: str  # earnings, deductions, benefits, reimbursements
    component_type: str  # Basic, HRA, Bonus, etc. (custom types allowed)
    component_name: str = ""
    name_in_payslip: str
    
    # Calculation (not for variable types like Bonus, Commission)
    is_variable: bool = False  # True for Bonus, Commission, etc.
    calculation_type: Optional[str] = None  # flat_amount, percentage_of_ctc, percentage_of_basic
    # Deprecated: amount_value is no longer set at component definition time
    amount_value: Optional[float] = None
    
    # Status
    is_active: bool = True
    
    # Other Configurations
    part_of_salary_structure: bool = True
    is_taxable: bool = False
    calculate_on_pro_rata: bool = False
    is_scheduled_earning: bool = False  # For variable types
    
    # Tax preference (for variable taxable earnings)
    tax_deduction_preference: Optional[str] = None  # subsequent_payrolls, same_payroll
    
    # Flexible Benefit Plan
    include_in_fbp: bool = False
    
    # EPF Configuration
    consider_for_epf: bool = False
    epf_contribution_rule: Optional[str] = None  # always, conditional
    pf_wage_threshold: float = 15000  # Default threshold
    
    # ESI Configuration
    consider_for_esi: bool = False
    
    # Display
    show_in_payslip: bool = True
    
    # Deduction specific fields
    deduction_frequency: Optional[str] = None  # one_time, recurring
    
    # Benefits specific fields
    benefit_plan: Optional[str] = None  # e.g., "Section 80C - Tax Saving Investments"
    benefit_association: Optional[str] = None  # Specific benefit under the plan
    include_employer_contribution: bool = False
    is_superannuation_fund: bool = False
    
    # Reimbursements specific fields
    unclaimed_handling: Optional[str] = None  # carry_forward, do_not_carry_forward
    # Deprecated: amount_per_month is no longer set at component definition time
    amount_per_month: Optional[float] = None
    
    # Metadata
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str  # Admin username


class SalaryComponentCreate(BaseModel):
    category: str  # earnings, deductions, benefits, reimbursements
    component_type: str = ""
    component_name: str = ""
    name_in_payslip: str
    is_variable: bool = False
    calculation_type: Optional[str] = None
    # Deprecated: amount_value is optional
    amount_value: Optional[float] = None
    is_active: bool = True
    part_of_salary_structure: bool = True
    is_taxable: bool = False
    calculate_on_pro_rata: bool = False
    is_scheduled_earning: bool = False
    tax_deduction_preference: Optional[str] = None
    include_in_fbp: bool = False
    consider_for_epf: bool = False
    epf_contribution_rule: Optional[str] = None
    pf_wage_threshold: float = 15000
    consider_for_esi: bool = False
    show_in_payslip: bool = True
    # New fields
    deduction_frequency: Optional[str] = None
    benefit_plan: Optional[str] = None
    benefit_association: Optional[str] = None
    include_employer_contribution: bool = False
    is_superannuation_fund: bool = False
    unclaimed_handling: Optional[str] = None
    # Deprecated: amount_per_month is optional
    amount_per_month: Optional[float] = None


class SalaryComponentUpdate(BaseModel):
    component_type: Optional[str] = None
    component_name: Optional[str] = None
    name_in_payslip: Optional[str] = None
    is_variable: Optional[bool] = None
    calculation_type: Optional[str] = None
    amount_value: Optional[float] = None
    is_active: Optional[bool] = None
    part_of_salary_structure: Optional[bool] = None
    is_taxable: Optional[bool] = None
    calculate_on_pro_rata: Optional[bool] = None
    is_scheduled_earning: Optional[bool] = None
    tax_deduction_preference: Optional[str] = None
    include_in_fbp: Optional[bool] = None
    consider_for_epf: Optional[bool] = None
    epf_contribution_rule: Optional[str] = None
    pf_wage_threshold: Optional[float] = None
    consider_for_esi: Optional[bool] = None
    show_in_payslip: Optional[bool] = None
    # New fields
    deduction_frequency: Optional[str] = None
    benefit_plan: Optional[str] = None
    benefit_association: Optional[str] = None
    include_employer_contribution: Optional[bool] = None
    is_superannuation_fund: Optional[bool] = None
    unclaimed_handling: Optional[str] = None
    amount_per_month: Optional[float] = None


# ==================== Tax Configuration Models ====================

class TaxConfiguration(BaseModel):
    config_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str  # Multi-tenancy
    component_type: str  # epf, esi, tds, professional_tax, lwf, statutory_bonus
    is_enabled: bool = False
    
    # EPF Configuration
    epf_number: Optional[str] = None
    epf_deduction_cycle: str = "monthly"
    epf_employee_contribution_rate: float = 12.0  # Percentage
    epf_employer_contribution_rate: float = 12.0  # Percentage
    epf_include_employer_contribution: bool = True
    epf_include_edli: bool = False
    epf_include_admin_charges: bool = False
    epf_override_at_employee_level: bool = False
    epf_pro_rate_restricted_wage: bool = False
    epf_consider_components_after_lop: bool = True
    
    # ESI Configuration
    esi_number: Optional[str] = None
    esi_deduction_cycle: str = "monthly"
    esi_employee_contribution_rate: float = 0.75  # Percentage
    esi_employer_contribution_rate: float = 3.25  # Percentage
    esi_include_employer_contribution: bool = False
    esi_wage_ceiling: float = 21000  # Monthly salary threshold
    
    # TDS Configuration
    tds_enabled: bool = False
    # Add more TDS fields as needed
    
    # Professional Tax Configuration
    pt_enabled: bool = False
    pt_state: Optional[str] = None
    # Add more PT fields as needed
    
    # Labour Welfare Fund Configuration
    lwf_enabled: bool = False
    lwf_state: Optional[str] = None
    # Add more LWF fields as needed
    
    # Statutory Bonus Configuration
    bonus_enabled: bool = False
    bonus_min_percentage: float = 8.33
    bonus_max_percentage: float = 20.0
    # Add more bonus fields as needed
    
    # Metadata
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str  # Admin username


class TaxConfigurationCreate(BaseModel):
    component_type: str
    is_enabled: bool = True
    epf_number: Optional[str] = None
    epf_deduction_cycle: str = "monthly"
    epf_employee_contribution_rate: float = 12.0
    epf_employer_contribution_rate: float = 12.0
    epf_include_employer_contribution: bool = True
    epf_include_edli: bool = False
    epf_include_admin_charges: bool = False
    epf_override_at_employee_level: bool = False
    epf_pro_rate_restricted_wage: bool = False
    epf_consider_components_after_lop: bool = True
    esi_number: Optional[str] = None
    esi_deduction_cycle: str = "monthly"
    esi_employee_contribution_rate: float = 0.75
    esi_employer_contribution_rate: float = 3.25
    esi_include_employer_contribution: bool = False
    esi_wage_ceiling: float = 21000


class TaxConfigurationUpdate(BaseModel):
    is_enabled: Optional[bool] = None
    epf_number: Optional[str] = None
    epf_deduction_cycle: Optional[str] = None
    epf_employee_contribution_rate: Optional[float] = None
    epf_employer_contribution_rate: Optional[float] = None
    epf_include_employer_contribution: Optional[bool] = None
    epf_include_edli: Optional[bool] = None
    epf_include_admin_charges: Optional[bool] = None
    epf_override_at_employee_level: Optional[bool] = None
    epf_pro_rate_restricted_wage: Optional[bool] = None
    epf_consider_components_after_lop: Optional[bool] = None
    esi_number: Optional[str] = None
    esi_deduction_cycle: Optional[str] = None
    esi_employee_contribution_rate: Optional[float] = None
    esi_employer_contribution_rate: Optional[float] = None
    esi_include_employer_contribution: Optional[bool] = None
    esi_wage_ceiling: Optional[float] = None


# API Routes
@api_router.get("/")
async def root():
    return {"message": "Elevate - Payroll Management System API"}

# Authentication endpoints
@api_router.post("/auth/generate-otp")
async def generate_employee_otp(request: OTPRequest):
    # Check if employee exists
    employee = await db.employees.find_one({"employee_id": request.employee_id})
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )
    
    # Generate OTP
    otp = generate_otp()
    active_otps[request.employee_id] = {
        "otp": otp,
        "expires_at": datetime.utcnow() + timedelta(minutes=5),
        "employee_email": employee.get("email")
    }
    
    # In production, send OTP via SMS/Email
    # For now, return OTP in response (development only)
    return OTPResponse(
        message=f"OTP sent to employee {request.employee_id}",
        otp=otp  # Remove this in production
    )

def parse_device_info(user_agent: str) -> str:
    """Parse User-Agent to extract device information"""
    if not user_agent or user_agent == "Unknown":
        return "Unknown Device"
    
    # Simple device detection
    user_agent_lower = user_agent.lower()
    
    if "mobile" in user_agent_lower or "android" in user_agent_lower:
        if "android" in user_agent_lower:
            return "Android Mobile"
        elif "iphone" in user_agent_lower:
            return "iPhone"
        else:
            return "Mobile Device"
    elif "ipad" in user_agent_lower:
        return "iPad"
    elif "macintosh" in user_agent_lower or "mac os" in user_agent_lower:
        return "Mac Desktop"
    elif "windows" in user_agent_lower:
        return "Windows Desktop"
    elif "linux" in user_agent_lower:
        return "Linux Desktop"
    else:
        # Try to extract browser name
        if "chrome" in user_agent_lower:
            return "Desktop (Chrome)"
        elif "firefox" in user_agent_lower:
            return "Desktop (Firefox)"
        elif "safari" in user_agent_lower:
            return "Desktop (Safari)"
        else:
            return "Desktop Browser"



async def get_ip_geolocation(ip_address: str) -> Optional[dict]:
    """
    Get geolocation information for an IP address using ip-api.com (free service)
    Returns: dict with city, region, country, lat, lon or None if lookup fails
    """
    # Skip private/local IPs
    if ip_address in ["Unknown", "127.0.0.1", "localhost"] or ip_address.startswith("192.168.") or ip_address.startswith("10."):
        return None
    
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.get(f"http://ip-api.com/json/{ip_address}?fields=status,message,country,regionName,city,lat,lon")
            
            if response.status_code == 200:
                data = response.json()
                
                if data.get("status") == "success":
                    return {
                        "city": data.get("city", "Unknown"),
                        "region": data.get("regionName", "Unknown"),
                        "country": data.get("country", "Unknown"),
                        "latitude": data.get("lat"),
                        "longitude": data.get("lon")
                    }
            
            return None
    except Exception as e:
        logging.error(f"Failed to get geolocation for IP {ip_address}: {str(e)}")
        return None



@api_router.post("/auth/login")
async def login(login_request: LoginRequest, request: Request):
    # Get client IP address
    client_ip = request.client.host if request.client else "Unknown"
    # Check for forwarded IP (in case of proxy)
    forwarded_for = request.headers.get("X-Forwarded-For")
    if forwarded_for:
        client_ip = forwarded_for.split(",")[0].strip()
    
    # Get User-Agent for device info
    user_agent = request.headers.get("User-Agent", "Unknown")
    device_name = parse_device_info(user_agent)
    
    # Check if user is admin or employee by looking up in database
    # Support both username and email for login
    user = await db.users.find_one({
        "$or": [
            {"username": login_request.username},
            {"email": login_request.username}
        ]
    })
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials"
        )
    
    # Admin login (check for password)
    if user.get("role") == "admin" or user.get("role") == "super_admin":
        if not login_request.password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Password required for admin/super admin login"
            )
        
        # Verify password
        if not verify_password(login_request.password, user["hashed_password"]):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid credentials"
            )
        
        user_data = User(**user)
        
    elif user.get("role") == "employee":
        # Employee login (check for password)
        if not login_request.password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Password required for employee login"
            )
        
        # Verify password
        if not verify_password(login_request.password, user["hashed_password"]):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid employee credentials"
            )
        
        # Check if employee is active
        employee_data = await db.employees.find_one({"employee_id": user["employee_id"]})
        if not employee_data:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Employee record not found"
            )
        
        employee_status = employee_data.get("status", "active")
        if employee_status != "active":
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail=f"Access denied: Employee status is {employee_status}"
            )
        
        user_data = User(**user)
    else:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid user role"
        )
    
    # Get geolocation for IP address
    location_data = await get_ip_geolocation(client_ip)
    
    # Update last login with device info and IP address
    await db.users.update_one(
        {"username": login_request.username},
        {"$set": {
            "last_login": datetime.now(timezone.utc).isoformat(),
            "last_login_ip": client_ip,
            "last_login_device": device_name
        }}
    )
    
    # Save to login history collection
    login_history_entry = {
        "id": str(uuid.uuid4()),
        "employee_id": user_data.employee_id if user_data.employee_id else login_request.username,
        "username": login_request.username,
        "login_time": datetime.now(timezone.utc),
        "ip_address": client_ip,
        "device_name": device_name,
        "pc_name": "Desktop",  # Default value, can be enhanced later
        "location": location_data
    }
    await db.login_history.insert_one(login_history_entry)
    
    # Create tokens
    access_token = create_access_token(
        data={"sub": user_data.username, "role": user_data.role, "company_id": user_data.company_id}
    )
    refresh_token = create_refresh_token(
        data={"sub": user_data.username, "role": user_data.role, "company_id": user_data.company_id}
    )
    
    return Token(
        access_token=access_token,
        refresh_token=refresh_token,
        user={
            "id": user_data.id,
            "username": user_data.username,
            "role": user_data.role,
            "company_id": user_data.company_id,
            "employee_id": user_data.employee_id,
            "email": user_data.email
        }
    )

@api_router.post("/auth/refresh")
async def refresh_token(request: dict):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        refresh_token_value = request.get("refresh_token")
        if not refresh_token_value:
            raise credentials_exception
        payload = jwt.decode(refresh_token_value, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        role: str = payload.get("role")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    # Create new access token
    access_token = create_access_token(
        data={"sub": username, "role": role}
    )
    
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.get("/auth/verify")
async def verify_token(current_user: User = Depends(get_current_user)):
    """Verify if the current access token is valid"""
    return {
        "valid": True,
        "user": {
            "id": current_user.id,
            "username": current_user.username,
            "role": current_user.role
        }
    }

@api_router.post("/auth/logout")
async def logout(current_user: User = Depends(get_current_user)):
    # In production, add token to blacklist
    return {"message": "Successfully logged out"}

@api_router.post("/auth/change-password")
async def change_password(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    current_password = request.get("current_password")
    new_password = request.get("new_password")
    
    if not current_password or not new_password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password and new password are required"
        )
    
    # Get user from database
    user = await db.users.find_one({"username": current_user.username})
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Verify current password
    if not verify_password(current_password, user["password"]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Hash new password and update
    new_password_hash = get_password_hash(new_password)
    await db.users.update_one(
        {"username": current_user.username},
        {"$set": {"password": new_password_hash, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Password changed successfully"}


# ============================================================================
# SUPER ADMIN ENDPOINTS - Multi-Tenancy Management
# ============================================================================

def require_super_admin(current_user: User = Depends(get_current_user)):
    """Dependency to ensure only super admin can access"""
    if current_user.role != "super_admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Super admin access required"
        )
    return current_user


async def send_invitation_email(email: str, company_name: str, admin_name: str, invitation_link: str):
    """Send invitation email to company admin"""
    try:
        message = MIMEMultipart("alternative")
        message["From"] = SMTP_FROM
        message["To"] = email
        message["Subject"] = f"Invitation to join {company_name} on Elevate Payroll"
        
        # HTML email body
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="color: #10B981;">Welcome to Elevate Payroll!</h2>
                    
                    <p>Hello {admin_name},</p>
                    
                    <p>You have been invited to be the administrator for <strong>{company_name}</strong> on Elevate Payroll System.</p>
                    
                    <p>To accept this invitation and set up your account, please click the button below:</p>
                    
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{invitation_link}" 
                           style="background-color: #10B981; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">
                            Accept Invitation
                        </a>
                    </div>
                    
                    <p>Or copy and paste this link into your browser:</p>
                    <p style="background-color: #f5f5f5; padding: 10px; border-radius: 5px; word-break: break-all;">
                        {invitation_link}
                    </p>
                    
                    <p style="color: #999; font-size: 12px; margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd;">
                        This invitation will expire in 7 days.<br>
                        If you didn't expect this invitation, you can safely ignore this email.
                    </p>
                </div>
            </body>
        </html>
        """
        
        message.attach(MIMEText(html_body, "html"))
        
        # Send email
        await aiosmtplib.send(
            message,
            hostname=SMTP_HOST,
            port=SMTP_PORT,
            username=SMTP_USER,
            password=SMTP_PASSWORD,
            start_tls=True
        )
        
        return True
    except Exception as e:
        logging.error(f"Error sending invitation email: {e}")
        return False


@api_router.get("/super-admin/dashboard/stats")
async def get_super_admin_dashboard_stats(current_user: User = Depends(require_super_admin)):
    """Get super admin dashboard statistics"""
    total_companies = await db.companies.count_documents({"status": "active"})
    total_users = await db.users.count_documents({"role": {"$in": ["admin", "employee"]}})
    total_employees = await db.employees.count_documents({})
    total_payroll_runs = await db.payroll_runs.count_documents({})
    
    # Get recent companies
    recent_companies = await db.companies.find(
        {},
        {"_id": 0}
    ).sort("created_at", -1).limit(5).to_list(length=None)
    
    return {
        "total_companies": total_companies,
        "total_users": total_users,
        "total_employees": total_employees,
        "total_payroll_runs": total_payroll_runs,
        "recent_companies": recent_companies
    }


@api_router.get("/super-admin/companies")
async def get_companies(
    current_user: User = Depends(require_super_admin),
    search: Optional[str] = None,
    status_filter: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all companies with optional search and filter"""
    query = {}
    
    if search:
        query["$or"] = [
            {"company_name": {"$regex": search, "$options": "i"}},
            {"contact_email": {"$regex": search, "$options": "i"}}
        ]
    
    if status_filter:
        query["status"] = status_filter
    
    companies = await db.companies.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(length=None)
    total = await db.companies.count_documents(query)
    
    return {
        "companies": companies,
        "total": total,
        "skip": skip,
        "limit": limit
    }


@api_router.get("/super-admin/companies/{company_id}")
async def get_company(
    company_id: str,
    current_user: User = Depends(require_super_admin)
):
    """Get company details"""
    company = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found"
        )
    
    # Get company statistics
    employee_count = await db.employees.count_documents({"company_id": company_id})
    user_count = await db.users.count_documents({"company_id": company_id})
    payroll_count = await db.payroll_runs.count_documents({"company_id": company_id})
    
    return {
        **company,
        "statistics": {
            "employee_count": employee_count,
            "user_count": user_count,
            "payroll_runs": payroll_count
        }
    }


@api_router.post("/super-admin/companies")
async def create_company(
    company_data: CompanyCreate,
    current_user: User = Depends(require_super_admin)
):
    """Create a new company and send invitation to admin"""
    # Check if company with same email already exists
    existing_company = await db.companies.find_one({"contact_email": company_data.contact_email})
    if existing_company:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Company with this email already exists"
        )
    
    # Create company
    company_id = str(uuid.uuid4())
    company = {
        "company_id": company_id,
        "company_name": company_data.company_name,
        "company_logo_url": company_data.company_logo_url,
        "contact_email": company_data.contact_email,
        "phone": company_data.phone,
        "address": company_data.address,
        "admin_user_id": None,  # Will be set when admin accepts invitation
        "settings": company_data.settings.dict() if company_data.settings else {
            "working_days_config": {
                "sunday_off": True,
                "saturday_policy": "alternate",
                "week_start": "Monday"
            },
            "leave_policies": {
                "annual_leave": 15,
                "sick_leave": 10,
                "casual_leave": 8
            },
            "default_working_hours": {
                "start_time": "08:30",
                "end_time": "17:30",
                "total_hours": 9
            }
        },
        "status": "active",
        "subscription_info": {
            "plan": "basic",
            "start_date": datetime.now(timezone.utc).isoformat(),
            "status": "active"
        },
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.companies.insert_one(company)
    
    # Create invitation
    invitation_token = str(uuid.uuid4())
    invitation = {
        "invitation_id": str(uuid.uuid4()),
        "company_id": company_id,
        "token": invitation_token,
        "email": company_data.admin_email,
        "admin_name": company_data.admin_name,
        "role": "admin",
        "status": "pending",
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "accepted_at": None
    }
    
    await db.invitations.insert_one(invitation)
    
    # Create invitation link
    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    invitation_link = f"{frontend_url}/accept-invitation?token={invitation_token}"
    
    # Try to send email
    email_sent = await send_invitation_email(
        company_data.admin_email,
        company_data.company_name,
        company_data.admin_name,
        invitation_link
    )
    
    return {
        "message": "Company created successfully",
        "company_id": company_id,
        "invitation_link": invitation_link,
        "email_sent": email_sent
    }


@api_router.put("/super-admin/companies/{company_id}")
async def update_company(
    company_id: str,
    company_data: CompanyUpdate,
    current_user: User = Depends(require_super_admin)
):
    """Update company details"""
    company = await db.companies.find_one({"company_id": company_id})
    
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found"
        )
    
    # Prepare update data
    update_data = {k: v for k, v in company_data.dict(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data:
        await db.companies.update_one(
            {"company_id": company_id},
            {"$set": update_data}
        )
    
    return {"message": "Company updated successfully"}


@api_router.delete("/super-admin/companies/{company_id}")
async def deactivate_company(
    company_id: str,
    current_user: User = Depends(require_super_admin)
):
    """Deactivate a company (soft delete)"""
    company = await db.companies.find_one({"company_id": company_id})
    
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found"
        )
    
    await db.companies.update_one(
        {"company_id": company_id},
        {"$set": {"status": "inactive", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Company deactivated successfully"}


# Invitation endpoints
@api_router.get("/invitations/verify/{token}")
async def verify_invitation(token: str):
    """Verify invitation token and get invitation details"""
    invitation = await db.invitations.find_one({"token": token}, {"_id": 0})
    
    if not invitation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invitation not found"
        )
    
    # Check if invitation is expired
    expires_at = datetime.fromisoformat(invitation["expires_at"])
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invitation has expired"
        )
    
    # Check if invitation is already accepted
    if invitation["status"] == "accepted":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invitation has already been accepted"
        )
    
    # Get company details
    company = await db.companies.find_one({"company_id": invitation["company_id"]}, {"_id": 0})
    
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found"
        )
    
    return {
        "invitation": {
            "email": invitation["email"],
            "admin_name": invitation["admin_name"],
            "company_name": company["company_name"],
            "company_logo_url": company.get("company_logo_url"),
            "expires_at": invitation["expires_at"]
        }
    }


@api_router.post("/invitations/accept")
async def accept_invitation(invitation_data: InvitationAccept):
    """Accept invitation and create admin user"""
    invitation = await db.invitations.find_one({"token": invitation_data.token}, {"_id": 0})
    
    if not invitation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invitation not found"
        )
    
    # Check if invitation is expired
    expires_at = datetime.fromisoformat(invitation["expires_at"])
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invitation has expired"
        )
    
    # Check if invitation is already accepted
    if invitation["status"] == "accepted":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invitation has already been accepted"
        )
    
    # Check if user with this email already exists
    existing_user = await db.users.find_one({"email": invitation["email"]})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User with this email already exists"
        )
    
    # Create admin user
    user_id = str(uuid.uuid4())
    # Generate username from email (before @)
    username = invitation["email"].split("@")[0] + "_" + invitation["company_id"][:8]
    
    admin_user = {
        "id": user_id,
        "username": username,
        "email": invitation["email"],
        "role": "admin",
        "company_id": invitation["company_id"],
        "employee_id": None,
        "hashed_password": get_password_hash(invitation_data.password),
        "pin": None,
        "is_active": True,
        "last_login": None,
        "last_login_ip": None,
        "last_login_device": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(admin_user)
    
    # Update company with admin_user_id
    await db.companies.update_one(
        {"company_id": invitation["company_id"]},
        {"$set": {"admin_user_id": user_id}}
    )
    
    # Mark invitation as accepted
    await db.invitations.update_one(
        {"token": invitation_data.token},
        {"$set": {
            "status": "accepted",
            "accepted_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create tokens for auto-login
    access_token = create_access_token(
        data={"sub": username, "role": "admin", "company_id": invitation["company_id"]}
    )
    refresh_token = create_refresh_token(
        data={"sub": username, "role": "admin", "company_id": invitation["company_id"]}
    )
    
    return {
        "message": "Invitation accepted successfully",
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "user": {
            "id": user_id,
            "username": username,
            "email": invitation["email"],
            "role": "admin",
            "company_id": invitation["company_id"]
        }
    }


# ============================================================================
# SUBSCRIPTION PLAN ENDPOINTS
# ============================================================================

@api_router.get("/plans/public")
async def get_public_plans():
    """Get all active subscription plans (public endpoint - no auth required)"""
    plans = await db.subscription_plans.find(
        {"is_active": True},
        {"_id": 0}
    ).sort("display_order", 1).to_list(length=None)
    
    return {"plans": plans}


@api_router.get("/super-admin/plans")
async def get_all_plans(current_user = Depends(require_super_admin)):
    """Get all subscription plans for super admin"""
    plans = await db.subscription_plans.find({}, {"_id": 0}).sort("display_order", 1).to_list(length=None)
    
    # Get subscriber count for each plan
    for plan in plans:
        subscriber_count = await db.companies.count_documents({
            "subscription_info.plan_id": plan["plan_id"],
            "status": "active"
        })
        plan["subscriber_count"] = subscriber_count
    
    return {"plans": plans}


@api_router.get("/super-admin/plans/{plan_id}")
async def get_plan(
    plan_id: str,
    current_user = Depends(require_super_admin)
):
    """Get plan details"""
    plan = await db.subscription_plans.find_one({"plan_id": plan_id}, {"_id": 0})
    
    if not plan:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Plan not found"
        )
    
    return plan


@api_router.put("/super-admin/plans/{plan_id}")
async def update_plan(
    plan_id: str,
    plan_data: PlanUpdate,
    current_user = Depends(require_super_admin)
):
    """Update subscription plan"""
    # Log what we receive
    logging.info(f"Updating plan {plan_id}")
    logging.info(f"Received plan_data: {plan_data}")
    if plan_data.features:
        logging.info(f"Features dict: {plan_data.features.model_dump()}")
    
    plan = await db.subscription_plans.find_one({"plan_id": plan_id})
    
    if not plan:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Plan not found"
        )
    
    # Prepare update data
    update_data = {}
    if plan_data.plan_name is not None:
        update_data["plan_name"] = plan_data.plan_name
    if plan_data.description is not None:
        update_data["description"] = plan_data.description
    if plan_data.price_per_user_monthly is not None:
        update_data["price_per_user_monthly"] = plan_data.price_per_user_monthly
    if plan_data.price_per_user_annual is not None:
        update_data["price_per_user_annual"] = plan_data.price_per_user_annual
    if plan_data.features is not None:
        # Don't use .dict() - it adds default values for missing fields
        # Instead, convert to dict and preserve only the provided fields
        update_data["features"] = plan_data.features.model_dump(exclude_unset=False)
    if plan_data.is_active is not None:
        update_data["is_active"] = plan_data.is_active
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data:
        await db.subscription_plans.update_one(
            {"plan_id": plan_id},
            {"$set": update_data}
        )
    
    return {"message": "Plan updated successfully"}


# ============================================================================
# PUBLIC SIGNUP ENDPOINT
# ============================================================================

@api_router.post("/signup")
async def public_signup(signup_data: PublicSignup):
    """Public self-service signup endpoint"""
    try:
        # ----------------------------------------------------------------------
        # Step 1: Prevent duplicate company or user
        # ----------------------------------------------------------------------
        existing_company = await db.companies.find_one({"contact_email": signup_data.contact_email})
        if existing_company:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Company with this email already exists"
            )
 
        existing_user = await db.users.find_one({"email": signup_data.admin_email})
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User with this email already exists"
            )
 
        # ----------------------------------------------------------------------
        # Step 2: Validate selected plan
        # ----------------------------------------------------------------------
        plan = await db.subscription_plans.find_one({"plan_id": signup_data.plan_id}, {"_id": 0})
        if not plan or not plan.get("is_active"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Selected plan not found or inactive"
            )
 
        # ----------------------------------------------------------------------
        # Step 3: Compute plan and subscription parameters
        # ----------------------------------------------------------------------
        now = datetime.now(timezone.utc)
        company_id = str(uuid.uuid4())
 
        # Determine price based on billing cycle
        billing_cycle = signup_data.billing_cycle.lower()
        if billing_cycle not in ["monthly", "annual"]:
            billing_cycle = "monthly"
 
        amount = plan["price_per_user_annual"] if billing_cycle == "annual" else plan["price_per_user_monthly"]
 
        # Trial logic
        trial_days = plan.get("features", {}).get("trial_days", 0)
        trial_ends_at = now + timedelta(days=trial_days) if trial_days > 0 else None
 
        # Set billing dates
        if billing_cycle == "annual":
            next_billing_date = now + timedelta(days=365)
        else:
            next_billing_date = now + timedelta(days=30)
 
        # If trial exists, next billing should start after trial
        if trial_ends_at:
            next_billing_date = trial_ends_at
 
        # ----------------------------------------------------------------------
        # Step 4: Create company record
        # ----------------------------------------------------------------------
        company = {
            "company_id": company_id,
            "company_name": signup_data.company_name,
            "company_logo_url": None,
            "contact_email": signup_data.contact_email,
            "phone": signup_data.phone,
            "address": None,
            "industry": signup_data.industry,
            "country": signup_data.country,
            "admin_user_id": None,  # to be updated later
            "settings": {
                "working_days_config": {
                    "sunday_off": True,
                    "saturday_policy": "alternate",
                    "week_start": "Monday"
                },
                "leave_policies": {
                    "annual_leave": 15,
                    "sick_leave": 10,
                    "casual_leave": 8
                },
                "default_working_hours": {
                    "start_time": "08:30",
                    "end_time": "17:30",
                    "total_hours": 9
                }
            },
            "status": "active",
            "subscription_info": {
                "plan_id": plan["plan_id"],
                "plan_name": plan["plan_name"],
                "billing_cycle": billing_cycle,
                "currency": plan.get("currency", "INR"),
                "price_per_user_monthly": plan.get("price_per_user_monthly"),
                "price_per_user_annual": plan.get("price_per_user_annual"),
                "amount": amount,
                "auto_renew": True,
                "status": "trial" if trial_days > 0 else "active",
                "start_date": now.isoformat(),
                "trial_end_date": trial_ends_at.isoformat() if trial_ends_at else None,
                "next_billing_date": next_billing_date.isoformat(),
            },
            "created_at": now.isoformat(),
            "updated_at": now.isoformat()
        }
 
        await db.companies.insert_one(company)
 
        # ----------------------------------------------------------------------
        # Step 5: Create admin user
        # ----------------------------------------------------------------------
        user_id = str(uuid.uuid4())
        username = signup_data.admin_email.split("@")[0] + "_" + company_id[:8]
 
        admin_user = {
            "id": user_id,
            "username": username,
            "email": signup_data.admin_email,
            "role": "admin",
            "company_id": company_id,
            "employee_id": None,
            "hashed_password": get_password_hash(signup_data.password),
            "pin": None,
            "is_active": True,
            "last_login": None,
            "last_login_ip": None,
            "last_login_device": None,
            "created_at": now.isoformat(),
            "updated_at": now.isoformat()
        }
 
        await db.users.insert_one(admin_user)
 
        # Update company with admin reference
        await db.companies.update_one(
            {"company_id": company_id},
            {"$set": {"admin_user_id": user_id}}
        )
 
        # ----------------------------------------------------------------------
        # Step 6: Create tokens for immediate login
        # ----------------------------------------------------------------------
        access_token = create_access_token(
            data={"sub": username, "role": "admin", "company_id": company_id}
        )
        refresh_token = create_refresh_token(
            data={"sub": username, "role": "admin", "company_id": company_id}
        )
 
        # ----------------------------------------------------------------------
        # Step 7: Return response
        # ----------------------------------------------------------------------
        return {
            "message": "Account created successfully! Your free trial has started." if trial_days > 0 else "Account created successfully!",
            "access_token": access_token,
            "refresh_token": refresh_token,
            "token_type": "bearer",
            "user": {
                "id": user_id,
                "username": username,
                "email": signup_data.admin_email,
                "role": "admin",
                "company_id": company_id
            },
            "subscription_info": company["subscription_info"]
        }
 
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in public signup: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="An error occurred during signup. Please try again."
        )
 


# ============================================================================
# PAYMENT & SUBSCRIPTION ENDPOINTS
# ============================================================================

@api_router.post("/subscription/create")
async def create_subscription(
    request: CreateSubscriptionRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Create Razorpay subscription after trial period
    Should be called when company wants to activate paid subscription
    """
    try:
        # Fetch company
        company = await db.companies.find_one({"company_id": request.company_id}, {"_id": 0})
        if not company:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company not found"
            )
        
        # Verify user has access to this company
        if current_user.role != "super_admin" and current_user.company_id != request.company_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
        
        subscription_info = company.get("subscription_info", {})
        plan_id = subscription_info.get("plan_id")
        
        # Get the plan details
        plan = await db.subscription_plans.find_one({"plan_id": plan_id}, {"_id": 0})
        if not plan:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Subscription plan not found"
            )
        
        # Get the appropriate Razorpay plan ID
        if request.billing_cycle == "annual":
            razorpay_plan_id = plan.get("razorpay_plan_id_annual")
            amount = plan["annual_price"]
        else:
            razorpay_plan_id = plan.get("razorpay_plan_id_monthly")
            amount = plan["monthly_price"]
        
        if not razorpay_plan_id:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Razorpay plan not configured. Please contact support."
            )
        
        # Create Razorpay subscription
        subscription_data = {
            "plan_id": razorpay_plan_id,
            "customer_notify": 1,
            "quantity": 1,
            "total_count": 60 if request.billing_cycle == "monthly" else 5,  # 5 years for annual, 60 months for monthly
            "notes": {
                "company_id": request.company_id,
                "company_name": company["company_name"]
            }
        }
        
        razorpay_subscription = razorpay_client.subscription.create(subscription_data)
        
        # Update company with Razorpay subscription info
        now = datetime.now(timezone.utc)
        next_billing_date = now + (timedelta(days=365) if request.billing_cycle == "annual" else timedelta(days=30))
        
        await db.companies.update_one(
            {"company_id": request.company_id},
            {
                "$set": {
                    "subscription_info.razorpay_subscription_id": razorpay_subscription["id"],
                    "subscription_info.razorpay_plan_id": razorpay_plan_id,
                    "subscription_info.status": "created",  # Will be updated to active after payment
                    "subscription_info.billing_cycle": request.billing_cycle,
                    "subscription_info.amount": amount,
                    "subscription_info.next_billing_date": next_billing_date.isoformat(),
                    "updated_at": now.isoformat()
                }
            }
        )
        
        return {
            "subscription_id": razorpay_subscription["id"],
            "razorpay_key": RAZORPAY_KEY_ID,
            "amount": int(amount * 100),  # Convert to paise
            "currency": plan["currency"],
            "plan_name": plan["plan_name"],
            "billing_cycle": request.billing_cycle
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating subscription: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@api_router.post("/subscription/verify-payment")
async def verify_payment(
    request: VerifyPaymentRequest,
    current_user: User = Depends(get_current_user)
):
    """Verify Razorpay payment signature"""
    try:
        # Verify signature
        params_dict = {
            'razorpay_subscription_id': request.subscription_id,
            'razorpay_payment_id': request.payment_id,
            'razorpay_signature': request.signature
        }
        
        razorpay_client.utility.verify_payment_signature(params_dict)
        
        # Find company with this subscription
        company = await db.companies.find_one(
            {"subscription_info.razorpay_subscription_id": request.subscription_id},
            {"_id": 0}
        )
        
        if not company:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company not found for this subscription"
            )
        
        # Update subscription status to active
        now = datetime.now(timezone.utc)
        await db.companies.update_one(
            {"company_id": company["company_id"]},
            {
                "$set": {
                    "subscription_info.status": "active",
                    "subscription_info.last_payment_date": now.isoformat(),
                    "status": "active",  # Also update company status
                    "updated_at": now.isoformat()
                }
            }
        )
        
        return {
            "message": "Payment verified successfully",
            "subscription_status": "active"
        }
        
    except razorpay.errors.SignatureVerificationError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid payment signature"
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error verifying payment: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@api_router.get("/subscription/status")
async def get_subscription_status(current_user: User = Depends(get_current_user)):
    """Get current subscription status for the user's company"""
    try:
        # Get company
        company = await db.companies.find_one({"company_id": current_user.company_id}, {"_id": 0})
        if not company:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company not found"
            )
        
        subscription_info = company.get("subscription_info", {})
        
        # Check if trial
        trial_end_date = subscription_info.get("trial_end_date")
        if trial_end_date:
            trial_end = datetime.fromisoformat(trial_end_date)
            now = datetime.now(timezone.utc)
            
            if now < trial_end:
                days_left = (trial_end - now).days
                return {
                    "status": "trial",
                    "trial_days_left": days_left,
                    "trial_end_date": trial_end_date,
                    "plan": subscription_info.get("plan_name") or subscription_info.get("plan", "Free Trial"),
                    "requires_payment": False
                }
            else:
                # Trial expired, requires payment
                return {
                    "status": "trial_expired",
                    "trial_end_date": trial_end_date,
                    "plan": subscription_info.get("plan_name") or subscription_info.get("plan", "Free Trial"),
                    "requires_payment": True
                }
        
        # Get subscription details
        status_str = subscription_info.get("status", "unknown")
        
        return {
            "status": status_str,
            "subscription_id": subscription_info.get("razorpay_subscription_id"),
            "current_period_end": subscription_info.get("next_billing_date"),
            "next_billing_date": subscription_info.get("next_billing_date"),
            "amount": subscription_info.get("amount"),
            "billing_cycle": subscription_info.get("billing_cycle", "monthly"),
            "auto_renew": subscription_info.get("auto_renew", True),
            "plan": subscription_info.get("plan_name") or subscription_info.get("plan", "Unknown"),
            "requires_payment": status_str in ["trial_expired", "expired", "payment_failed"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting subscription status: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@api_router.get("/subscription/features")
async def get_subscription_features(current_user: User = Depends(get_current_user)):
    """Get subscription features for the current user's company"""
    try:
        # Get company
        company = await db.companies.find_one({"company_id": current_user.company_id}, {"_id": 0})
        if not company:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company not found"
            )
        
        subscription_info = company.get("subscription_info", {})
        plan_id = subscription_info.get("plan_id")
        
        # If no plan_id, check for legacy plan field
        if not plan_id:
            plan_name = subscription_info.get("plan", "free")
            # Get plan by slug
            plan = await db.subscription_plans.find_one({"slug": plan_name.lower()}, {"_id": 0})
        else:
            # Get plan by plan_id
            plan = await db.subscription_plans.find_one({"plan_id": plan_id}, {"_id": 0})
        
        # If still no plan found, return default free trial plan features
        if not plan:
            return {
                "plan_name": "Free Trial",
                "plan_slug": "free",
                "features": {
                    "employee_limit": 5,
                    "admin_users_limit": 1,
                    "employee_database": True,
                    "payroll_processing_manual": True,
                    "payroll_processing_automated": False,
                    "payslip_generation": True,
                    "attendance_tracking_basic": True,
                    "attendance_tracking_advanced": False,
                    "leave_management_basic": True,
                    "leave_management_advanced": False,
                    "salary_structure_management": False,
                    "bank_advice_generation": False,
                    "custom_salary_components": False,
                    "bulk_employee_import": False,
                    "compliance_reports_basic": False,
                    "compliance_reports_full": False,
                    "employee_portal": True,  # Free Trial HAS employee portal
                    "loans_advances": False,
                    "deductions_advanced": False,
                    "event_management": False,
                    "payroll_analytics": False,
                    "multi_bank_accounts": False,
                    "notifications": False,
                    "dark_mode": False,
                    "api_access": False,
                    "white_labeling": False,
                    "custom_integrations": False,
                    "sso_security": False,
                    "custom_reports": False,
                    "audit_logs": False,
                    "sla_guarantee": False,
                    "support_level": "email"
                }
            }
        
        return {
            "plan_name": plan.get("plan_name"),
            "plan_slug": plan.get("slug"),
            "features": plan.get("features", {})
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting subscription features: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@api_router.get("/subscription/upgrade-options")
async def get_upgrade_options(current_user: User = Depends(get_current_user)):
    """Get available upgrade options for current company"""
    try:
        # Get company
        company = await db.companies.find_one({"company_id": current_user.company_id}, {"_id": 0})
        if not company:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company not found"
            )
        
        subscription_info = company.get("subscription_info", {})
        current_plan_id = subscription_info.get("plan_id")
        
        if not current_plan_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No active subscription found"
            )
        
        # Get current plan
        current_plan = await db.subscription_plans.find_one({"plan_id": current_plan_id}, {"_id": 0})
        if not current_plan:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Current plan not found"
            )
        
        current_display_order = current_plan.get("display_order", 0)
        
        # Get all plans with higher display_order (upgrades only, no downgrades)
        upgrade_plans = await db.subscription_plans.find(
            {
                "display_order": {"$gt": current_display_order},
                "is_active": True
            },
            {"_id": 0}
        ).sort("display_order", 1).to_list(length=None)
        
        return {
            "current_plan": current_plan,
            "upgrade_options": upgrade_plans,
            "billing_cycle": subscription_info.get("billing_cycle", "monthly")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting upgrade options: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@api_router.post("/subscription/calculate-upgrade")
async def calculate_upgrade_cost(
    target_plan_id: str,
    new_billing_cycle: Optional[str] = None,   # <-- keep this EXACT name
    current_user: User = Depends(get_current_user),
):
    """
    Calculate cost for:
      1) Upgrading to a higher plan (same billing cycle), OR
      2) Switching billing cycle on the same plan (Monthly <-> Annual), OR
      3) Upgrading plan AND switching billing cycle in one go.

    Rules:
      - Credit the unused remainder of the *current* period (monthly=30d, annual=365d).
      - If switching to Annual *now*, charge: (annual_price - credit_remaining) * employees
        and set next_billing_date_new = today + 365 days (full fresh year).
      - If switching to Monthly from Annual, no immediate charge; the switch takes effect at renewal.
    """
    try:
        # --- Load company & subscription info
        company = await db.companies.find_one(
            {"company_id": current_user.company_id}, {"_id": 0}
        )
        if not company:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")

        subscription_info = company.get("subscription_info", {}) or {}
        current_plan_id = subscription_info.get("plan_id")
        current_cycle = subscription_info.get("billing_cycle", "monthly")
        next_billing_date_str = subscription_info.get("next_billing_date")

        if not current_plan_id or not next_billing_date_str:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid subscription information")

        # --- Plans
        current_plan = await db.subscription_plans.find_one({"plan_id": current_plan_id}, {"_id": 0})
        target_plan  = await db.subscription_plans.find_one({"plan_id": target_plan_id}, {"_id": 0})
        if not current_plan or not target_plan:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan not found")

        # --- Guard against downgrade by display_order
        if target_plan.get("display_order", 0) < current_plan.get("display_order", 0):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot downgrade plans")

        # --- Employees
        employee_count = await db.employees.count_documents({
            "company_id": current_user.company_id,
            "status": "active"
        })

        # --- Parse dates
        nxt_str = next_billing_date_str.replace("Z", "+00:00")
        try:
            next_billing_date = datetime.fromisoformat(nxt_str)
        except Exception:
            # try without tz then force UTC
            base = next_billing_date_str.split("+")[0].split("Z")[0]
            next_billing_date = datetime.fromisoformat(base).replace(tzinfo=timezone.utc)
        if next_billing_date.tzinfo is None:
            next_billing_date = next_billing_date.replace(tzinfo=timezone.utc)

        today = datetime.now(timezone.utc)
        days_remaining = max((next_billing_date - today).days, 0)

        # --- helpers
        def price(plan, cycle):
            return plan.get("price_per_user_annual", 0) if cycle == "annual" else plan.get("price_per_user_monthly", 0)

        def total_days_for(cycle):
            return 365 if cycle == "annual" else 30

        # Effective target cycle (what user picked on UI, else current)
        effective_cycle = (new_billing_cycle or current_cycle).lower()
        if effective_cycle not in ("monthly", "annual"):
            effective_cycle = current_cycle  # fallback safety

        # Scenario flags
        same_plan = (target_plan_id == current_plan_id)
        cycle_change = (effective_cycle != current_cycle)

        # --- CREDIT for unused portion of the *current* period
        current_period_days = total_days_for(current_cycle)
        current_price_per_user = price(current_plan, current_cycle)
        credit_per_user = 0.0
        if days_remaining > 0 and current_price_per_user > 0:
            credit_per_user = (current_price_per_user / current_period_days) * days_remaining

        # --- TARGET pricing
        target_price_per_user = price(target_plan if not same_plan else current_plan, effective_cycle)

        # --- Compute charge
        response_type = "plan_upgrade"
        message = None
        new_next_billing_date = None
        pro_rated_amount_per_user = 0.0
        total_upgrade_cost = 0.0
        price_difference_per_user = 0.0

        if same_plan and cycle_change:
            # (A) Switching billing cycle only
            if current_cycle == "monthly" and effective_cycle == "annual":
                # Charge annual minus remaining monthly credit; start a fresh 1-year term from today
                response_type = "billing_cycle_upgrade"
                price_difference_per_user = target_price_per_user - credit_per_user
                pro_rated_amount_per_user = max(price_difference_per_user, 0.0)  # never negative charge
                total_upgrade_cost = round(pro_rated_amount_per_user * employee_count, 2)
                new_next_billing_date = (today + timedelta(days=365)).isoformat()

                message = "Switching to annual now. Unused monthly time is credited; new annual cycle starts today."
            elif current_cycle == "annual" and effective_cycle == "monthly":
                # Make change at renewal; no immediate charge
                response_type = "billing_cycle_downgrade_scheduled"
                pro_rated_amount_per_user = 0.0
                total_upgrade_cost = 0.0
                price_difference_per_user = 0.0
                new_next_billing_date = next_billing_date_str
                message = "Switch to monthly will take effect at renewal. No immediate charge."
            else:
                # Unknown combination, fall back safe
                response_type = "billing_cycle_change"
                pro_rated_amount_per_user = 0.0
                total_upgrade_cost = 0.0
                new_next_billing_date = next_billing_date_str
                message = "Billing cycle change processed."

        else:
            # (B) Upgrading plan (with or without cycle change)
            if not cycle_change:
                # Same cycle plan upgrade: prorate difference for remaining days
                current_target_price = price(target_plan, current_cycle)
                price_difference_per_user = current_target_price - current_price_per_user
                per_day_diff = price_difference_per_user / total_days_for(current_cycle)
                pro_rated_amount_per_user = max(per_day_diff * days_remaining, 0.0)
                total_upgrade_cost = round(pro_rated_amount_per_user * employee_count, 2)
                new_next_billing_date = next_billing_date_str
                message = "Plan upgraded. Pro-rated difference charged for remaining days of current period."
            else:
                # Plan upgrade + cycle switch
                if current_cycle == "monthly" and effective_cycle == "annual":
                    # Credit remaining monthly; charge new annual for target plan
                    annual_target_per_user = price(target_plan, "annual")
                    price_difference_per_user = annual_target_per_user - credit_per_user
                    pro_rated_amount_per_user = max(price_difference_per_user, 0.0)
                    total_upgrade_cost = round(pro_rated_amount_per_user * employee_count, 2)
                    new_next_billing_date = (today + timedelta(days=365)).isoformat()
                    message = "Plan upgraded and switched to annual now. Credit applied; new annual cycle starts today."
                elif current_cycle == "annual" and effective_cycle == "monthly":
                    # Defer cycle change; charge only plan difference for remaining annual days (if any)
                    current_target_price = price(target_plan, "annual")
                    price_difference_per_user = current_target_price - current_price_per_user
                    per_day_diff = price_difference_per_user / total_days_for("annual")
                    pro_rated_amount_per_user = max(per_day_diff * days_remaining, 0.0)
                    total_upgrade_cost = round(pro_rated_amount_per_user * employee_count, 2)
                    new_next_billing_date = next_billing_date_str
                    response_type = "plan_upgrade_with_cycle_downgrade_scheduled"
                    message = "Plan upgraded now. Switch to monthly will take effect at renewal."

        return {
            "type": response_type,
            "current_plan": {
                "plan_id": current_plan.get("plan_id"),
                "plan_name": current_plan.get("plan_name"),
            },
            "target_plan": {
                "plan_id": target_plan.get("plan_id") if not same_plan else current_plan.get("plan_id"),
                "plan_name": target_plan.get("plan_name") if not same_plan else current_plan.get("plan_name"),
                "price_per_user": round(target_price_per_user, 2),
            },
            "employee_count": employee_count,
            "current_billing_cycle": current_cycle,
            "effective_billing_cycle": effective_cycle,
            "days_remaining": days_remaining,
            "credit_per_user": round(credit_per_user, 2),
            "price_difference_per_user": round(price_difference_per_user, 2),
            "pro_rated_amount_per_user": round(pro_rated_amount_per_user, 2),
            "total_upgrade_cost": total_upgrade_cost,
            "next_billing_date": next_billing_date_str,
            "next_billing_date_new": new_next_billing_date,  # may be None if unchanged
            "message": message,
        }

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error calculating upgrade cost: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@api_router.post("/subscription/upgrade")
async def upgrade_subscription(
    target_plan_id: str,
    payment_id: str,
    current_user: User = Depends(get_current_user)
):
    """Process subscription upgrade after payment"""
    try:
        # Get company
        company = await db.companies.find_one({"company_id": current_user.company_id}, {"_id": 0})
        if not company:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company not found"
            )
        
        subscription_info = company.get("subscription_info", {})
        current_plan_id = subscription_info.get("plan_id")
        
        # Get target plan
        target_plan = await db.subscription_plans.find_one({"plan_id": target_plan_id}, {"_id": 0})
        if not target_plan:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Target plan not found"
            )
        
        # Verify payment with Razorpay
        try:
            payment = razorpay_client.payment.fetch(payment_id)
            if payment['status'] != 'captured':
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Payment not captured"
                )
        except Exception as e:
            logging.error(f"Error verifying payment: {e}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Payment verification failed"
            )
        
        # Update subscription immediately (effective immediately)
        update_data = {
            "subscription_info.plan_id": target_plan_id,
            "subscription_info.plan_slug": target_plan.get("slug"),
            "subscription_info.plan_name": target_plan.get("plan_name"),
            "subscription_info.last_payment_date": datetime.now(timezone.utc).isoformat(),
            "subscription_info.last_upgrade_date": datetime.now(timezone.utc).isoformat(),
            "subscription_info.last_payment_id": payment_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.companies.update_one(
            {"company_id": current_user.company_id},
            {"$set": update_data}
        )
        
        logging.info(f"Upgraded company {current_user.company_id} from {current_plan_id} to {target_plan_id}")
        
        return {
            "message": "Subscription upgraded successfully",
            "new_plan": {
                "plan_id": target_plan_id,
                "plan_name": target_plan.get("plan_name"),
                "plan_slug": target_plan.get("slug")
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error upgrading subscription: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@api_router.post("/webhooks/razorpay")
async def razorpay_webhook(request: Request):
    """Handle Razorpay webhook events"""
    try:
        # Get the webhook payload and signature
        payload = await request.body()
        signature = request.headers.get("X-Razorpay-Signature", "")
        
        # Verify webhook signature
        razorpay_client.utility.verify_webhook_signature(
            payload.decode("utf-8"),
            signature,
            RAZORPAY_WEBHOOK_SECRET
        )
        
        # Parse the payload
        event_data = json.loads(payload)
        event = event_data.get("event")
        payload_data = event_data.get("payload", {})
        subscription_entity = payload_data.get("subscription", {}).get("entity", {})
        payment_entity = payload_data.get("payment", {}).get("entity", {})
        
        subscription_id = subscription_entity.get("id") or payment_entity.get("subscription_id")
        
        logging.info(f"Received Razorpay webhook event: {event} for subscription: {subscription_id}")
        
        # Find company with this subscription
        company = await db.companies.find_one(
            {"subscription_info.razorpay_subscription_id": subscription_id},
            {"_id": 0}
        )
        
        if not company:
            logging.warning(f"Company not found for subscription: {subscription_id}")
            return {"status": "ignored"}
        
        now = datetime.now(timezone.utc)
        
        # Handle different webhook events
        if event == "subscription.activated":
            # Subscription activated after payment
            await db.companies.update_one(
                {"company_id": company["company_id"]},
                {
                    "$set": {
                        "subscription_info.status": "active",
                        "subscription_info.last_payment_date": now.isoformat(),
                        "status": "active",
                        "updated_at": now.isoformat()
                    }
                }
            )
            logging.info(f"Subscription activated for company: {company['company_id']}")
            
        elif event == "subscription.charged":
            # Recurring payment successful
            await db.companies.update_one(
                {"company_id": company["company_id"]},
                {
                    "$set": {
                        "subscription_info.status": "active",
                        "subscription_info.last_payment_date": now.isoformat(),
                        "status": "active",
                        "updated_at": now.isoformat()
                    }
                }
            )
            
            # Send payment success notification to admin
            admin_user = await db.users.find_one(
                {"company_id": company["company_id"], "role": "admin"},
                {"_id": 0}
            )
            if admin_user:
                await create_notification_helper(
                    title="Payment Successful",
                    message=f"Your subscription payment of ₹{subscription_entity.get('plan_id', {}).get('amount', 0)/100} was successful.",
                    recipient_role="admin",
                    notification_type="success",
                    category="subscription",
                    recipient_id=admin_user.get("id")
                )
            
            logging.info(f"Subscription charged for company: {company['company_id']}")
            
        elif event == "payment.failed":
            # Payment failed
            await db.companies.update_one(
                {"company_id": company["company_id"]},
                {
                    "$set": {
                        "subscription_info.status": "payment_failed",
                        "updated_at": now.isoformat()
                    }
                }
            )
            
            # Send payment failed notification to admin
            admin_user = await db.users.find_one(
                {"company_id": company["company_id"], "role": "admin"},
                {"_id": 0}
            )
            if admin_user:
                await create_notification_helper(
                    title="Payment Failed",
                    message="Your subscription payment failed. Please update your payment method to continue using the service.",
                    recipient_role="admin",
                    notification_type="error",
                    category="subscription",
                    recipient_id=admin_user.get("id")
                )
            
            logging.warning(f"Payment failed for company: {company['company_id']}")
            
        elif event == "subscription.cancelled":
            # Subscription cancelled
            await db.companies.update_one(
                {"company_id": company["company_id"]},
                {
                    "$set": {
                        "subscription_info.status": "cancelled",
                        "status": "inactive",
                        "updated_at": now.isoformat()
                    }
                }
            )
            
            # Send cancellation notification to admin
            admin_user = await db.users.find_one(
                {"company_id": company["company_id"], "role": "admin"},
                {"_id": 0}
            )
            if admin_user:
                await create_notification_helper(
                    title="Subscription Cancelled",
                    message="Your subscription has been cancelled. Your access will be disabled at the end of the current billing cycle.",
                    recipient_role="admin",
                    notification_type="warning",
                    category="subscription",
                    recipient_id=admin_user.get("id")
                )
            
            logging.info(f"Subscription cancelled for company: {company['company_id']}")
            
        elif event == "subscription.completed":
            # Subscription period completed
            await db.companies.update_one(
                {"company_id": company["company_id"]},
                {
                    "$set": {
                        "subscription_info.status": "expired",
                        "status": "inactive",
                        "updated_at": now.isoformat()
                    }
                }
            )
            logging.info(f"Subscription completed for company: {company['company_id']}")
        
        return {"status": "processed"}
        
    except razorpay.errors.SignatureVerificationError:
        logging.error("Invalid webhook signature")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid webhook signature"
        )
    except Exception as e:
        logging.error(f"Error processing webhook: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )



@api_router.get("/notification-settings")
async def get_notification_settings(current_user: User = Depends(get_current_user)):
    """Get user notification settings"""
    settings = await db.notification_settings.find_one({"user_id": current_user.username})
    
    if settings:
        # Clean up MongoDB ObjectId and other non-serializable fields
        settings = prepare_from_mongo(settings)
    else:
        # Return default settings if none exist
        settings = {
            "user_id": current_user.username,
            "email_notifications": True,
            "sms_notifications": False,
            "payroll_reminders": True,
            "compliance_alerts": True,
            "birthday_reminders": True,
            "leave_notifications": True,
            "loan_reminders": True
        }
    
    return settings

@api_router.put("/notification-settings")
async def update_notification_settings(
    settings: dict,
    current_user: User = Depends(get_current_user)
):
    """Update user notification settings"""
    settings["user_id"] = current_user.username
    settings["updated_at"] = datetime.now(timezone.utc)
    
    await db.notification_settings.update_one(
        {"user_id": current_user.username},
        {"$set": settings},
        upsert=True
    )
    
    # Create a notification about settings change
    await create_notification_helper(
        title="Settings Updated",
        message="Your notification preferences have been updated successfully.",
        recipient_role="admin",
        notification_type="success",
        category="settings"
    )
    
    return {"message": "Notification settings updated successfully"}

@api_router.post("/test-notifications")
async def create_test_notifications(current_user: User = Depends(get_current_user)):
    """Create test notifications for different categories"""
    
    # Get user's notification settings
    user_settings = await db.notification_settings.find_one({"user_id": current_user.username}) or {}
    
    test_notifications = []
    
    if user_settings.get("payroll_reminders", True):
        await create_notification_helper(
            title="Payroll Reminder",
            message="Monthly payroll processing is due in 3 days. Please review and process employee salaries.",
            recipient_role="admin",
            notification_type="info",
            category="payroll"
        )
        test_notifications.append("Payroll Reminder")
    
    if user_settings.get("compliance_alerts", True):
        await create_notification_helper(
            title="Compliance Alert",
            message="PF filing deadline is approaching. Ensure all statutory compliance documents are ready.",
            recipient_role="admin",
            notification_type="warning",
            category="compliance"
        )
        test_notifications.append("Compliance Alert")
    
    if user_settings.get("birthday_reminders", True):
        await create_notification_helper(
            title="Birthday Reminder",
            message="Employee John Smith has a birthday today. Consider sending birthday wishes!",
            recipient_role="admin",
            notification_type="info",
            category="birthday"
        )
        test_notifications.append("Birthday Reminder")
    
    if user_settings.get("leave_notifications", True):
        await create_notification_helper(
            title="Leave Request",
            message="Test leave request has been submitted and requires your approval.",
            recipient_role="admin",
            notification_type="info",
            category="leave"
        )
        test_notifications.append("Leave Request")
    
    if user_settings.get("loan_reminders", True):
        await create_notification_helper(
            title="Loan Reminder",
            message="Test loan application is pending approval. Please review the application details.",
            recipient_role="admin",
            notification_type="info",
            category="loan"
        )
        test_notifications.append("Loan Reminder")
    
    return {
        "message": f"Created {len(test_notifications)} test notifications based on your settings",
        "notifications": test_notifications
    }

@api_router.get("/dashboard/pending-actions")
async def get_pending_actions():
    """Get counts of items requiring admin action"""
    try:
        pending_leaves = await db.leave_requests.count_documents({"status": "pending"})
        pending_loans = await db.loan_requests.count_documents({"status": "pending"})
        pending_ot = await db.ot_logs.count_documents({"status": "pending"})
        
        # Late arrivals from last 7 days (for review)
        seven_days_ago = (date.today() - timedelta(days=7)).isoformat()
        recent_late_arrivals = await db.late_arrivals.count_documents({
            "date": {"$gte": seven_days_ago}
        })
        
        return {
            "pending_leaves": pending_leaves,
            "pending_loans": pending_loans,
            "pending_ot": pending_ot,
            "recent_late_arrivals": recent_late_arrivals,
            "total_pending": pending_leaves + pending_loans + pending_ot
        }
    except Exception as e:
        logging.error(f"Error fetching pending actions: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch pending actions"
        )

@api_router.get("/dashboard/leave-statistics")
async def get_leave_statistics():
    """Get leave statistics for dashboard"""
    try:
        now = datetime.now()
        month_start = date(now.year, now.month, 1)
        
        # Pending leaves
        pending = await db.leave_requests.count_documents({"status": "pending"})
        
        # Approved leaves this month
        approved_this_month = await db.leave_requests.count_documents({
            "status": "approved",
            "start_date": {"$gte": month_start.isoformat()}
        })
        
        # Total leave balance utilization (average across all employees)
        employees = await db.employees.find({"status": "active"}).to_list(length=None)
        if employees:
            total_available = 0
            total_used = 0
            for emp in employees:
                total_available += emp.get("total_leave_balance", 12)
                leaves = await db.leave_requests.find({
                    "employee_id": emp["employee_id"],
                    "status": "approved"
                }).to_list(length=None)
                total_used += sum(l.get("days", 0) for l in leaves)
            
            utilization = (total_used / total_available * 100) if total_available > 0 else 0
        else:
            utilization = 0
        
        return {
            "pending_leaves": pending,
            "approved_this_month": approved_this_month,
            "utilization_percentage": round(utilization, 1)
        }
    except Exception as e:
        logging.error(f"Error fetching leave statistics: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch leave statistics"
        )

@api_router.get("/dashboard/loan-overview")
async def get_loan_overview():
    """Get loan overview for dashboard"""
    try:
        # Active loans (approved status - loans that have been approved and disbursed)
        active_loans = await db.loan_requests.find({
            "status": "approved"
        }).to_list(length=None)
        
        # Calculate totals from disbursed and outstanding amounts
        total_outstanding = 0
        monthly_deductions = 0
        
        for loan in active_loans:
            # Use outstanding_amount (set when approved) or fall back to disbursed_amount or original amount
            outstanding = loan.get("outstanding_amount") or loan.get("disbursed_amount") or loan.get("amount", 0)
            total_outstanding += outstanding
            
            # Use monthly_emi field
            emi = loan.get("monthly_emi", 0)
            monthly_deductions += emi
        
        return {
            "active_loans": len(active_loans),
            "total_outstanding": round(total_outstanding, 2),
            "monthly_deductions": round(monthly_deductions, 2)
        }
    except Exception as e:
        logging.error(f"Error fetching loan overview: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch loan overview"
        )

@api_router.get("/dashboard/employee-distribution")
async def get_employee_distribution():
    """Get employee distribution by department and designation"""
    try:
        employees = await db.employees.find({"status": "active"}).to_list(length=None)
        
        # Group by department
        dept_counts = {}
        for emp in employees:
            dept = emp.get("department", "Unassigned")
            dept_counts[dept] = dept_counts.get(dept, 0) + 1
        
        # Group by designation
        designation_counts = {}
        for emp in employees:
            desig = emp.get("designation", "Unassigned")
            designation_counts[desig] = designation_counts.get(desig, 0) + 1
        
        return {
            "by_department": [{"name": k, "count": v} for k, v in dept_counts.items()],
            "by_designation": [{"name": k, "count": v} for k, v in designation_counts.items()]
        }
    except Exception as e:
        logging.error(f"Error fetching employee distribution: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch employee distribution"
        )

@api_router.get("/dashboard/payroll-trends")
async def get_payroll_trends():
    """Get payroll trends for last 6 months"""
    try:
        now = datetime.now()
        trends = []
        
        for i in range(5, -1, -1):  # Last 6 months
            month = now.month - i
            year = now.year
            
            # Handle year rollback
            if month <= 0:
                month += 12
                year -= 1
            
            payroll_run = await db.payroll_runs.find_one({
                "month": month,
                "year": year
            })
            
            month_name = datetime(year, month, 1).strftime("%b %Y")
            trends.append({
                "month": month_name,
                "amount": payroll_run.get("total_net", 0) if payroll_run else 0
            })
        
        return {"trends": trends}
    except Exception as e:
        logging.error(f"Error fetching payroll trends: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch payroll trends"
        )

@api_router.get("/dashboard/attendance-overview")
async def get_attendance_overview():
    """Get today's attendance overview for dashboard"""
    try:
        today = date.today()
        
        # Get today's attendance records
        attendance_records = await db.attendance.find({
            "date": today.isoformat()
        }).to_list(length=None)
        
        # Count by status
        present = len([r for r in attendance_records if r.get("status") == "present"])
        leave = len([r for r in attendance_records if r.get("status") == "leave"])
        half_day = len([r for r in attendance_records if r.get("status") == "half-day"])
        
        # Get late arrivals for today
        late_arrivals = await db.late_arrivals.count_documents({
            "date": today.isoformat()
        })
        
        # Calculate total active employees
        total_active = await db.employees.count_documents({"status": "active"})
        
        # Calculate absent (active employees not in any attendance status)
        accounted_for = present + leave + half_day
        absent = max(0, total_active - accounted_for) if total_active > 0 else 0
        
        # Calculate rates
        present_rate = (present / total_active * 100) if total_active > 0 else 0
        on_time_rate = ((present - late_arrivals) / present * 100) if present > 0 else 0
        
        return {
            "present": present,
            "absent": absent,
            "late": late_arrivals,
            "leave": leave + half_day,
            "present_rate": round(present_rate, 1),
            "on_time_rate": round(on_time_rate, 1),
            "total_active": total_active
        }
    except Exception as e:
        logging.error(f"Error fetching attendance overview: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch attendance overview"
        )

# Dashboard endpoint
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    # Get total and active employees
    total_employees = await db.employees.count_documents({})
    active_employees = await db.employees.count_documents({"status": "active"})
    
    # Calculate this month's payroll
    now = datetime.now()
    month_start = date(now.year, now.month, 1)
    if now.month == 12:
        month_end = date(now.year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(now.year, now.month + 1, 1) - timedelta(days=1)
    
    # Get payroll run for current month
    payroll_run = await db.payroll_runs.find_one({
        "month": now.month,
        "year": now.year
    })
    
    this_month_payroll = 0
    if payroll_run:
        # Sum net salary from all employees in the payroll run
        this_month_payroll = sum(emp.get("net_salary", 0) for emp in payroll_run.get("employees", []))
    
    # Count payslips generated this month
    payslips_generated = await db.payslips.count_documents({
        "month": now.month,
        "year": now.year
    })
    
    # Calculate upcoming deadlines (pending leave requests + pending loan requests)
    upcoming_deadlines = 0
    pending_leaves = await db.leave_requests.count_documents({"status": "pending"})
    pending_loans = await db.loans.count_documents({"status": "pending"})
    upcoming_deadlines = pending_leaves + pending_loans
    
    return DashboardStats(
        total_employees=total_employees,
        active_employees=active_employees,
        this_month_payroll=this_month_payroll,
        payslips_generated=payslips_generated,
        upcoming_deadlines=upcoming_deadlines
    )

# Employee CRUD operations

# Helper function to check employee limit
async def check_employee_limit(company_id: str, additional_employees: int = 1):
    """
    Check if adding employees would exceed the plan limit
    Returns: (allowed: bool, message: str, current_count: int, limit: int)
    """
    # Get company
    company = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    if not company:
        return False, "Company not found", 0, 0
    
    # Get subscription info
    subscription_info = company.get("subscription_info", {})
    plan_id = subscription_info.get("plan_id")
    
    if not plan_id:
        # No plan, default to free trial limit
        employee_limit = 5
    else:
        # Get plan details
        plan = await db.subscription_plans.find_one({"plan_id": plan_id}, {"_id": 0})
        if not plan:
            employee_limit = 5  # Default to free trial
        else:
            employee_limit = plan.get("features", {}).get("employee_limit", 5)
    
    # Count current active employees
    current_employee_count = await db.employees.count_documents({
        "company_id": company_id,
        "status": "active"
    })
    
    # Check if unlimited
    if employee_limit == -1:
        return True, "Unlimited employees allowed", current_employee_count, -1
    
    # Check if adding would exceed limit
    new_total = current_employee_count + additional_employees
    if new_total > employee_limit:
        return False, f"Employee limit exceeded. Your plan allows {employee_limit} employees, you currently have {current_employee_count} active employees", current_employee_count, employee_limit
    
    return True, f"OK - {new_total}/{employee_limit} employees", current_employee_count, employee_limit


@api_router.get("/employees/limit-status")
async def get_employee_limit_status(current_user: User = Depends(get_current_user)):
    """Get current employee count and limit for the company"""
    allowed, message, current_count, limit = await check_employee_limit(current_user.company_id, 0)
    
    limit_str = "Unlimited" if limit == -1 else str(limit)
    
    return {
        "current_count": current_count,
        "limit": limit,
        "limit_str": limit_str,
        "can_add_more": allowed or limit == -1,
        "remaining": limit - current_count if limit != -1 else -1,
        "message": message
    }


@api_router.post("/employees", response_model=Employee)
async def create_employee(
    employee_data: EmployeeCreate,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    # Check employee limit before creating
    allowed, message, current_count, limit = await check_employee_limit(current_user.company_id, 1)
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=message
        )
    
    # Check if employee_id already exists in the company
    existing_employee = await db.employees.find_one({
        "employee_id": employee_data.employee_id,
        **company_filter
    })
    if existing_employee:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Employee ID already exists"
        )
    
    # Create employee with default empty objects for None values
    employee_data_dict = employee_data.dict()
    
    # Add company_id to employee (ALWAYS set it, even for super_admin if they have a company)
    if not employee_data_dict.get('company_id'):
        employee_data_dict['company_id'] = current_user.company_id
        logging.info(f"Setting company_id from current_user: {current_user.company_id} for employee {employee_data_dict.get('name')}")
    else:
        logging.info(f"Employee already has company_id: {employee_data_dict.get('company_id')}")
    
    if employee_data_dict.get('bank_info') is None:
        employee_data_dict['bank_info'] = {
            'bank_name': '',
            'account_number': '',
            'ifsc_code': '',
            'branch': ''
        }
    if employee_data_dict.get('salary_structure') is None:
        employee_data_dict['salary_structure'] = {
            'basic_salary': 0.0,
            'hra': 0.0,
            'medical_allowance': 0.0,
            'travel_allowance': 0.0,
            'food_allowance': 0.0,
            'internet_allowance': 0.0,
            'special_allowance': 0.0,
            'pf_employee': 0.0,
            'pf_employer': 0.0,
            'esi_employee': 0.0,
            'esi_employer': 0.0,
            'professional_tax': 0.0,
            'tds': 0.0
        }
    
    employee = Employee(**employee_data_dict)
    employee_dict = prepare_for_mongo(employee.dict())
    
    logging.info(f"Inserting employee: {employee_dict.get('name')} with company_id: {employee_dict.get('company_id')}")
    
    result = await db.employees.insert_one(employee_dict)
    logging.info(f"Insert result - inserted_id: {result.inserted_id}")
    
    if result.inserted_id:
        # Create corresponding user account for the employee
        # Set default password for employee
        default_password = "Test@1234"
        hashed_password = get_password_hash(default_password)
        employee_user = User(
            username=employee.employee_id,
            email=employee.email,
            role=UserRole.EMPLOYEE,
            company_id=employee.company_id,  # Use employee's company_id, not user's
            employee_id=employee.employee_id,
            hashed_password=hashed_password,
            is_active=True
        )
        await db.users.insert_one(prepare_for_mongo(employee_user.dict()))
        
        # Notify admins about new employee addition
        await notify_employee_added(
            employee_id=employee.employee_id,
            employee_name=employee.name,
            department=employee.department
        )
        
        return employee
    else:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create employee"
        )

@api_router.get("/employees", response_model=List[Employee])
async def get_employees(
    skip: int = 0, 
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    employees = await db.employees.find(company_filter).skip(skip).limit(limit).to_list(length=None)
    return [Employee(**employee) for employee in employees]

@api_router.get("/employees/{employee_id}", response_model=Employee)
async def get_employee(
    employee_id: str,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    query = {"employee_id": employee_id, **company_filter}
    employee = await db.employees.find_one(query)
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )
    return Employee(**employee)

@api_router.get("/employees/{employee_id}/pending-emi")
async def get_employee_pending_emi(
    employee_id: str,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    """Get pending EMI amount for an employee based on approved loans"""
    try:
        # Verify employee belongs to user's company
        query = {"employee_id": employee_id, **company_filter}
        employee = await db.employees.find_one(query)
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee not found"
            )
        
        # Find approved loans for the employee that are still active
        loan_query = {
            "employee_id": employee_id,
            "status": "approved",
            **company_filter
        }
        approved_loans = await db.loan_requests.find(loan_query).to_list(length=None)
        
        total_emi = 0
        loan_details = []
        
        for loan in approved_loans:
            # Calculate remaining months and EMI
            monthly_emi = loan.get("monthly_emi", 0)
            remaining_emis = loan.get("remaining_emis", 0)
            
            if monthly_emi > 0 and remaining_emis > 0:
                total_emi += monthly_emi
                loan_details.append({
                    "loan_id": loan.get("id"),
                    "loan_type": loan.get("loan_type", "Personal Loan"),
                    "amount": loan.get("amount", 0),
                    "emi_amount": monthly_emi,
                    "remaining_months": remaining_emis,
                    "total_months": loan.get("tenure_months", 0)
                })
        
        return {
            "employee_id": employee_id,
            "pending_emi": total_emi,
            "loan_details": loan_details[0] if loan_details else None,  # Return first loan details
            "total_loans": len(loan_details)
        }
        
    except Exception as e:
        logging.error(f"Error fetching pending EMI for {employee_id}: {str(e)}")
        return {
            "employee_id": employee_id,
            "pending_emi": 0,
            "loan_details": None,
            "total_loans": 0
        }


@api_router.get("/employees/{employee_id}/rating")
async def get_employee_rating(employee_id: str, month: Optional[int] = None, year: Optional[int] = None, period: Optional[str] = "current_month"):
    """
    Calculate employee performance rating based on CUMULATIVE Base 4.0 algorithm (Option B - Balanced System):
    - Starts at 4.0 in January
    - Carries forward month-to-month throughout the year
    - Resets to 4.0 in January of next year
    - Late arrivals: -0.02 per occurrence (balanced penalty)
    - Approved OT hours: +0.01 per hour (meaningful reward)
    - Punctuality bonus: +0.15 if no late arrivals in the month (fair incentive)
    - Maximum rating: 5.0
    
    Parameters:
    - period: "current_month" (default) or "ytd" (same as current_month for cumulative system)
    - month/year: Specific month/year
    """
    try:
        # Default to current month/year if not provided
        now = datetime.now(timezone.utc)
        target_month = month if month is not None else now.month
        target_year = year if year is not None else now.year
        
        # For cumulative rating, we need to calculate from January onwards
        async def calculate_month_rating(calc_month: int, calc_year: int) -> dict:
            """Calculate rating for a specific month, considering previous month's rating"""
            
            # Base rating: 4.0 for January, or previous month's rating for other months
            if calc_month == 1:
                starting_rating = 4.0
            else:
                # Get previous month's rating recursively
                prev_month = calc_month - 1 if calc_month > 1 else 12
                prev_year = calc_year if calc_month > 1 else calc_year - 1
                
                # Don't go back beyond January of target year
                if prev_year < target_year:
                    starting_rating = 4.0
                else:
                    prev_rating = await calculate_month_rating(prev_month, prev_year)
                    starting_rating = prev_rating["rating"]
            
            # Get late arrivals for this specific month
            start_date = datetime(calc_year, calc_month, 1, tzinfo=timezone.utc)
            if calc_month == 12:
                end_date = datetime(calc_year + 1, 1, 1, tzinfo=timezone.utc)
            else:
                end_date = datetime(calc_year, calc_month + 1, 1, tzinfo=timezone.utc)
            
            late_arrivals = await db.late_arrivals.count_documents({
                "employee_id": employee_id,
                "date": {
                    "$gte": start_date.isoformat(),
                    "$lt": end_date.isoformat()
                }
            })
            
            # Get approved OT hours for this specific month
            ot_logs = await db.ot_logs.find({
                "employee_id": employee_id,
                "status": "approved",
                "date": {
                    "$gte": start_date.isoformat(),
                    "$lt": end_date.isoformat()
                }
            }).to_list(length=None)
            
            total_ot_hours = sum(log.get("ot_hours", 0) for log in ot_logs)
            
            # Get attendance count for this month
            attendance_count = await db.attendance.count_documents({
                "employee_id": employee_id,
                "date": {
                    "$gte": start_date.isoformat(),
                    "$lt": end_date.isoformat()
                },
                "status": "present"
            })
            
            # Calculate rating changes for this month (Option B - Balanced System)
            rating = starting_rating
            
            # Apply late arrival penalty (-0.02 per occurrence)
            rating -= (late_arrivals * 0.02)
            
            # Apply OT bonus (+0.01 per hour)
            rating += (total_ot_hours * 0.01)
            
            # Apply punctuality bonus (+0.15 if no late arrivals THIS month)
            # IMPORTANT: Only credit punctuality bonus for COMPLETED months (past months)
            # For current ongoing month, bonus is pending until month ends
            is_current_month = (calc_month == now.month and calc_year == now.year)
            is_future_month = (calc_year > now.year) or (calc_year == now.year and calc_month > now.month)
            
            if is_future_month:
                # Future month - no bonus
                punctuality_bonus = 0.0
                bonus_status = "future"
            elif is_current_month:
                # Current ongoing month - bonus pending
                punctuality_bonus = 0.0
                bonus_status = "pending" if late_arrivals == 0 else "lost"
            else:
                # Past completed month - apply bonus if earned
                punctuality_bonus = 0.15 if late_arrivals == 0 else 0.0
                bonus_status = "earned" if late_arrivals == 0 else "lost"
            
            rating += punctuality_bonus
            
            # Cap at 5.0
            rating = min(rating, 5.0)
            
            return {
                "rating": round(rating, 2),
                "late_arrivals": late_arrivals,
                "ot_hours": total_ot_hours,
                "punctuality_bonus": punctuality_bonus,
                "punctuality_bonus_status": bonus_status,
                "attendance_days": attendance_count,
                "starting_rating": round(starting_rating, 2),
                "is_current_month": is_current_month
            }
        
        # Calculate rating for target month (which includes all previous months in the year)
        result = await calculate_month_rating(target_month, target_year)
        
        return {
            "employee_id": employee_id,
            "rating": result["rating"],
            "period": period,
            "month": target_month,
            "year": target_year,
            "details": {
                "starting_rating": result["starting_rating"],
                "base_rating": 4.0,
                "current_rating": result["rating"],
                "late_arrivals": result["late_arrivals"],
                "ot_hours": result["ot_hours"],
                "punctuality_bonus": result["punctuality_bonus"],
                "punctuality_bonus_status": result["punctuality_bonus_status"],
                "is_current_month": result["is_current_month"],
                "attendance_days": result["attendance_days"],
                "progress_from_baseline": round(result["rating"] - 4.0, 2)
            }
        }
        
    except Exception as e:
        logging.error(f"Error calculating rating for {employee_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to calculate rating: {str(e)}"
        )


@api_router.get("/employees/{employee_id}/earnings-overview")
async def get_employee_earnings_overview(
    employee_id: str,
    year: int = None,
    all_years: bool = False,
    current_user: User = Depends(get_current_user)
):
    """
    Get comprehensive earnings overview for an employee including:
    - YTD Earnings (January to current month of selected year, or total for all years)
    - Financial Year Earnings (April to March of selected FY, or total for all years)
    - Current Month Earnings (of selected year or current month)
    - Previous Month Earnings (of selected year or previous month)
    - Average Monthly Salary (last 6 months of selected year, or all time average)
    - Monthly trend (12 months of selected year, or all months for all years)
    """
    try:
        # Ensure user can only access their own data (unless admin)
        if current_user.role != UserRole.ADMIN and current_user.employee_id != employee_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
        
        now = datetime.now(timezone.utc)
        
        if all_years:
            # Get ALL payslips for the employee across all years
            payslips = await db.payslips.find({
                "employee_id": employee_id
            }).sort([("year", -1), ("month", -1)]).to_list(length=None)
        else:
            # Use provided year or current year
            selected_year = year if year else now.year
            
            # If selected year is current year, use current month, otherwise use December
            if selected_year == now.year:
                current_month = now.month
            else:
                current_month = 12  # For past years, show full year
            
            # Get all payslips for the employee in selected year
            payslips = await db.payslips.find({
                "employee_id": employee_id,
                "year": selected_year
            }).sort("month", -1).to_list(length=None)
        
        if not payslips:
            # Return zeros if no payslips found
            return {
                "ytd_earnings": 0,
                "fy_earnings": 0,
                "current_month_earnings": 0,
                "previous_month_earnings": 0,
                "average_monthly_salary": 0,
                "monthly_trend": [],
                "has_data": False
            }
        
        # Calculate earnings
        ytd_earnings = 0
        fy_earnings = 0
        current_month_earnings = 0
        previous_month_earnings = 0
        all_earnings = []
        monthly_trend = []
        
        if all_years:
            # For all years, calculate totals across all payslips
            for payslip in payslips:
                net_salary = payslip.get("net_salary", 0)
                gross_salary = payslip.get("gross_salary", 0)
                month = payslip.get("month")
                year_val = payslip.get("year")
                
                # Total earnings
                ytd_earnings += net_salary
                fy_earnings += net_salary
                all_earnings.append(net_salary)
                
                # Current month earnings (actual current month)
                if month == now.month and year_val == now.year:
                    current_month_earnings = net_salary
                
                # Previous month earnings
                prev_month = now.month - 1 if now.month > 1 else 12
                prev_year = now.year if now.month > 1 else now.year - 1
                if month == prev_month and year_val == prev_year:
                    previous_month_earnings = net_salary
                
                # Add to monthly trend for all years
                monthly_trend.append({
                    "month": month,
                    "year": year_val,
                    "month_name": f"{calendar.month_abbr[month]} {year_val}",
                    "gross_salary": float(gross_salary),
                    "net_salary": float(net_salary)
                })
            
            # Calculate average
            average_monthly_salary = sum(all_earnings) / len(all_earnings) if all_earnings else 0
            
            # Reverse trend to show oldest to newest
            monthly_trend.reverse()
            
        else:
            # Original single-year logic
            selected_year = year if year else now.year
            current_month = now.month if selected_year == now.year else 12
            
            last_6_months_earnings = []
            
            for payslip in payslips:
                month = payslip.get("month")
                year_val = payslip.get("year")  # Will always be selected_year due to query filter
                
                if not month or not year_val:
                    continue
                
                # Parse net salary
                net_salary = payslip.get("net_salary", 0)
                gross_salary = payslip.get("gross_salary", 0)
                
                # YTD Earnings (Jan to current_month of selected year)
                if month <= current_month:
                    ytd_earnings += net_salary
                
                # Financial Year Earnings (April to current_month of selected year)
                if current_month >= 4:
                    # Current FY: April to current_month of selected year
                    if month >= 4 and month <= current_month:
                        fy_earnings += net_salary
                else:
                    # Spanning FY: April-December of previous year + Jan-current_month of selected year
                    if month >= 4 or month <= current_month:
                        fy_earnings += net_salary
                
                # Current Month
                if month == current_month:
                    current_month_earnings = net_salary
                
                # Previous Month
                prev_month = current_month - 1 if current_month > 1 else 12
                if month == prev_month:
                    previous_month_earnings = net_salary
                
                # Last 6 months for average (within selected year)
                if month <= current_month and month >= max(1, current_month - 5):
                    last_6_months_earnings.append(net_salary)
                
                # All 12 months for trend
                monthly_trend.append({
                    "month": month,
                    "year": year_val,
                    "month_name": datetime(year_val, month, 1).strftime("%b %Y"),
                    "gross_salary": gross_salary,
                    "net_salary": net_salary
                })
            
            # Calculate average
            average_monthly_salary = sum(last_6_months_earnings) / len(last_6_months_earnings) if last_6_months_earnings else 0
            
            # Sort monthly trend by date (oldest first for chart)
            monthly_trend.sort(key=lambda x: (x["year"], x["month"]))
        
        return {
            "ytd_earnings": round(ytd_earnings, 2),
            "fy_earnings": round(fy_earnings, 2),
            "current_month_earnings": round(current_month_earnings, 2),
            "previous_month_earnings": round(previous_month_earnings, 2),
            "average_monthly_salary": round(average_monthly_salary, 2),
            "monthly_trend": monthly_trend,
            "has_data": True
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching earnings overview for {employee_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch earnings overview: {str(e)}"
        )


@api_router.put("/employees/{employee_id}", response_model=Employee)
async def update_employee(
    employee_id: str, 
    employee_data: EmployeeUpdate,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    # Check permissions: employees can only update their own records and can't update email
    if current_user.role == UserRole.EMPLOYEE:
        if current_user.employee_id != employee_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Employees can only update their own profile"
            )
        # Remove email from update data for employees
        employee_data.email = None
        
    # Include fields that are not None, but also include boolean fields even if False
    update_data = {k: v for k, v in employee_data.dict().items() if v is not None or k == 'is_on_probation'}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Log the update data for debugging
    logger.info(f"Updating employee {employee_id} with data: {update_data}")
    
    # Fetch existing employee data BEFORE update (for notification comparison)
    existing_employee = await db.employees.find_one({"employee_id": employee_id, **company_filter})
    if not existing_employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )
    
    update_data = prepare_for_mongo(update_data)
    
    result = await db.employees.update_one(
        {"employee_id": employee_id, **company_filter},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )
    
    updated_employee = await db.employees.find_one({"employee_id": employee_id, **company_filter})
    
    # Send notification to admins if employee updated their own profile
    if current_user.role == UserRole.EMPLOYEE:
        # Get the fields that were actually changed by comparing with existing data
        changed_fields = []
        for key, new_value in update_data.items():
            if key not in ['updated_at']:
                # Get the old value from existing employee
                old_value = existing_employee.get(key)
                
                # Check if value actually changed
                if old_value != new_value:
                    # Special handling for nested objects like bank_info
                    if key == 'bank_info' and isinstance(new_value, dict) and isinstance(old_value, dict):
                        # Check which bank fields changed
                        bank_changes = []
                        for bank_key, bank_new_value in new_value.items():
                            bank_old_value = old_value.get(bank_key)
                            if bank_old_value != bank_new_value:
                                bank_changes.append(bank_key.replace('_', ' ').title())
                        if bank_changes:
                            changed_fields.extend([f"Bank {field}" for field in bank_changes])
                    else:
                        # Regular field change
                        field_name = key.replace('_', ' ').title()
                        changed_fields.append(field_name)
        
        if changed_fields:
            # Create a readable list of changes
            if len(changed_fields) == 1:
                changes_text = changed_fields[0]
            elif len(changed_fields) == 2:
                changes_text = f"{changed_fields[0]} and {changed_fields[1]}"
            else:
                changes_text = f"{', '.join(changed_fields[:-1])}, and {changed_fields[-1]}"
            
            await notify_profile_change(
                employee_id=employee_id,
                employee_name=updated_employee.get('name', 'Unknown'),
                changes=changes_text
            )
    
    return Employee(**updated_employee)

@api_router.delete("/employees/bulk")
async def bulk_delete_employees(
    request: BulkDeleteRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    company_filter: dict = Depends(get_company_filter)
):
    """Delete multiple employees at once"""
    
    if not request.employee_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No employee IDs provided"
        )
    
    # Delete employees and associated user accounts
    deleted_count = 0
    errors = []
    
    for employee_id in request.employee_ids:
        try:
            # Delete employee - ensure within company scope
            result = await db.employees.delete_one({"employee_id": employee_id, **company_filter})
            if result.deleted_count > 0:
                deleted_count += 1
                # Also delete associated user account if exists (within company)
                await db.users.delete_one({"employee_id": employee_id, **company_filter})
            else:
                errors.append(f"Employee {employee_id} not found")
        except Exception as e:
            errors.append(f"Error deleting employee {employee_id}: {str(e)}")
    
    return {
        "message": f"Successfully deleted {deleted_count} employees",
        "deleted_count": deleted_count,
        "total_requested": len(request.employee_ids),
        "errors": errors
    }

@api_router.delete("/employees/{employee_id}")
async def delete_employee(
    employee_id: str,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    result = await db.employees.delete_one({"employee_id": employee_id, **company_filter})
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )
    return {"message": "Employee deleted successfully"}

@api_router.get("/employees/export/payroll")
async def export_payroll_data(
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    company_filter: dict = Depends(get_company_filter)
):
    """Export payroll data for all employees"""
    try:
        employees = await db.employees.find(company_filter).to_list(length=None)
        
        payroll_data = []
        for emp in employees:
            if emp.get('salary_structure'):
                salary = emp['salary_structure']
                
                # Calculate totals
                gross_salary = (
                    salary.get('basic_salary', 0) +
                    salary.get('hra', 0) +
                    salary.get('medical_allowance', 0) +
                    salary.get('travel_allowance', 0) +
                    salary.get('food_allowance', 0) +
                    salary.get('internet_allowance', 0) +
                    salary.get('special_allowance', 0)
                )
                
                total_deductions = (
                    salary.get('pf_employee', 0) +
                    salary.get('esi_employee', 0) +
                    salary.get('professional_tax', 0) +
                    salary.get('tds', 0)
                )
                
                net_salary = gross_salary - total_deductions
                
                payroll_data.append({
                    'Employee ID': emp.get('employee_id', ''),
                    'Name': emp.get('name', ''),
                    'Department': emp.get('department', ''),
                    'Designation': emp.get('designation', ''),
                    'Status': emp.get('status', ''),
                    'Basic Salary': salary.get('basic_salary', 0),
                    'HRA': salary.get('hra', 0),
                    'Medical Allowance': salary.get('medical_allowance', 0),
                    'Travel Allowance': salary.get('travel_allowance', 0),
                    'Food Allowance': salary.get('food_allowance', 0),
                    'Internet Allowance': salary.get('internet_allowance', 0),
                    'Special Allowance': salary.get('special_allowance', 0),
                    'Gross Salary': gross_salary,
                    'PF Employee': salary.get('pf_employee', 0),
                    'PF Employer': salary.get('pf_employer', 0),
                    'ESI Employee': salary.get('esi_employee', 0),
                    'ESI Employer': salary.get('esi_employer', 0),
                    'Professional Tax': salary.get('professional_tax', 0),
                    'TDS': salary.get('tds', 0),
                    'Total Deductions': total_deductions,
                    'Net Salary': net_salary,
                    'Bank Name': emp.get('bank_info', {}).get('bank_name', ''),
                    'Account Number': emp.get('bank_info', {}).get('account_number', ''),
                    'IFSC Code': emp.get('bank_info', {}).get('ifsc_code', ''),
                })
        
        return {"payroll_data": payroll_data}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export payroll data: {str(e)}"
        )

@api_router.put("/employees/{employee_id}/status", response_model=Employee)
async def update_employee_status(
    employee_id: str, 
    status_data: EmployeeStatusUpdate,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    company_filter: dict = Depends(get_company_filter)
):
    """Update employee status (active, resigned, terminated) with dates and reasons"""
    
    # Validate date requirements based on status
    if status_data.status == EmployeeStatus.RESIGNED and not status_data.resignation_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Resignation date is required for resigned status"
        )
    
    if status_data.status == EmployeeStatus.TERMINATED and not status_data.termination_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Termination date is required for terminated status"
        )
    
    # Prepare update data
    update_data = {
        "status": status_data.status,
        "updated_at": datetime.now(timezone.utc)
    }
    
    if status_data.status == EmployeeStatus.RESIGNED:
        update_data["resignation_date"] = status_data.resignation_date
        update_data["termination_date"] = None  # Clear termination date if exists
    elif status_data.status == EmployeeStatus.TERMINATED:
        update_data["termination_date"] = status_data.termination_date
        update_data["resignation_date"] = None  # Clear resignation date if exists
    else:
        # Active status - clear both dates
        update_data["resignation_date"] = None
        update_data["termination_date"] = None
    
    if status_data.status_reason:
        update_data["status_reason"] = status_data.status_reason
    
    update_data = prepare_for_mongo(update_data)
    
    result = await db.employees.update_one(
        {"employee_id": employee_id, **company_filter},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )
    
    updated_employee = await db.employees.find_one({"employee_id": employee_id, **company_filter})
    return Employee(**updated_employee)

# Settings Model
class CompanySettings(BaseModel):
    company_name: str
    address: str
    phone: str
    email: str
    website: str
    logo_url: Optional[str] = None
    pan: str
    tan: str
    gstin: str
    cin: str
    pf_registration: str
    esi_registration: str
    pt_registration: str

class PayrollSettings(BaseModel):
    financial_year_start: str
    salary_cycle: str
    pay_date: int
    working_days_per_week: int
    working_hours_per_day: int
    overtime_calculation: str
    auto_calculate_tax: bool
    include_weekends: bool
    leave_encashment: bool

class WorkingDaysConfig(BaseModel):
    saturday_policy: str = "alternate"  # "all_working", "all_off", "alternate", "custom"
    off_saturdays: List[int] = [1, 3]  # Which Saturdays are off (1st, 2nd, 3rd, 4th, 5th)
    sunday_off: bool = True

class Holiday(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # ISO format date string
    name: str
    description: Optional[str] = None
    is_optional: bool = False  # Optional holidays
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class HolidayImport(BaseModel):
    date: str
    name: str
    description: Optional[str] = None
    is_optional: bool = False

class SystemSettings(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_settings: CompanySettings
    payroll_settings: PayrollSettings
    working_days_config: Optional[WorkingDaysConfig] = Field(default_factory=lambda: WorkingDaysConfig())
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Settings endpoints
@api_router.get("/settings", response_model=SystemSettings)
async def get_settings():
    settings = await db.settings.find_one({})
    if not settings:
        # Return default settings if none exist
        default_settings = SystemSettings(
            company_settings=CompanySettings(
                company_name='Elevate',
                address='123 Business District, Mumbai, Maharashtra - 400001',
                phone='+91 22 1234 5678',
                email='hr@elevate.co.in',
                website='www.elevate.co.in',
                logo_url=None,
                pan='ABCDE1234F',
                tan='MUMB12345A',
                gstin='27ABCDE1234F1Z5',
                cin='U74999MH2020PTC123456',
                pf_registration='MAHA/MUM/1234567',
                esi_registration='ESI123456789',
                pt_registration='PT123456789'
            ),
            payroll_settings=PayrollSettings(
                financial_year_start='april',
                salary_cycle='monthly',
                pay_date=1,
                working_days_per_week=5,
                working_hours_per_day=8,
                overtime_calculation='enabled',
                auto_calculate_tax=True,
                include_weekends=False,
                leave_encashment=True
            ),
            working_days_config=WorkingDaysConfig(
                saturday_policy='alternate',
                off_saturdays=[1, 3],
                sunday_off=True
            )
        )
        # Save default settings to database
        settings_dict = prepare_for_mongo(default_settings.dict())
        await db.settings.insert_one(settings_dict)
        return default_settings
    else:
        return SystemSettings(**settings)

@api_router.put("/settings")
async def update_settings(settings_data: SystemSettings):
    try:
        settings_data.updated_at = datetime.now(timezone.utc)
        settings_dict = prepare_for_mongo(settings_data.dict())
        
        # Update or insert settings
        result = await db.settings.replace_one(
            {},  # Match any document (we only have one settings document)
            settings_dict,
            upsert=True  # Insert if document doesn't exist
        )
        
        return {"message": "Settings updated successfully", "success": True}
    except Exception as e:
        logging.error(f"Error updating settings: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update settings"
        )

# Holiday Management Endpoints
@api_router.get("/holidays")
async def get_holidays(year: Optional[int] = None):
    """Get all holidays, optionally filtered by year"""
    try:
        query = {}
        if year:
            # Filter holidays for specific year
            start_date = f"{year}-01-01"
            end_date = f"{year}-12-31"
            query = {
                "date": {
                    "$gte": start_date,
                    "$lte": end_date
                }
            }
        
        holidays = await db.holidays.find(query, {"_id": 0}).sort("date", 1).to_list(length=None)
        return holidays
    except Exception as e:
        logging.error(f"Error fetching holidays: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch holidays"
        )

@api_router.post("/holidays")
async def create_holiday(holiday: HolidayImport, current_user: User = Depends(get_current_user)):
    """Create a single holiday (admin only)"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    
    try:
        # Validate date format
        try:
            parsed_date = datetime.fromisoformat(holiday.date)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use YYYY-MM-DD"
            )
        
        # Check if holiday already exists for this date in the current year only
        year = parsed_date.year
        existing = await db.holidays.find_one({"date": holiday.date})
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Holiday already exists for {holiday.date}"
            )
        
        new_holiday = Holiday(
            date=holiday.date,
            name=holiday.name,
            description=holiday.description,
            is_optional=holiday.is_optional
        )
        
        holiday_dict = prepare_for_mongo(new_holiday.dict())
        await db.holidays.insert_one(holiday_dict)
        
        return {"message": "Holiday created successfully", "holiday": new_holiday}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating holiday: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create holiday"
        )

@api_router.post("/holidays/import")
async def import_holidays(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Import holidays from Excel file (admin only)"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    
    try:
        import openpyxl
        from io import BytesIO
        
        # Read Excel file
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        holidays_to_import = []
        errors = []
        skipped = []
        
        # Get current year for duplicate validation
        current_year = datetime.now(timezone.utc).year
        
        # Skip header row and process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            try:
                # Parse date
                date_value = row[0]
                parsed_date = None
                
                if isinstance(date_value, datetime):
                    parsed_date = date_value
                    date_str = date_value.strftime('%Y-%m-%d')
                elif isinstance(date_value, date):
                    parsed_date = datetime.combine(date_value, datetime.min.time())
                    date_str = date_value.isoformat()
                else:
                    date_str = str(date_value).strip()
                    # Try to parse common date formats - DD/MM/YYYY is the preferred format
                    try:
                        parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
                        date_str = parsed_date.strftime('%Y-%m-%d')
                    except:
                        try:
                            parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
                            date_str = parsed_date.strftime('%Y-%m-%d')
                        except:
                            errors.append(f"Row {row_idx}: Invalid date format '{date_str}'. Use DD/MM/YYYY")
                            continue
                
                name = str(row[1]).strip() if row[1] else ""
                description = str(row[2]).strip() if row[2] and len(row) > 2 else None
                is_optional = bool(row[3]) if row[3] and len(row) > 3 else False
                
                if not name:
                    errors.append(f"Row {row_idx}: Holiday name is required")
                    continue
                
                # Check if holiday already exists (only for current year)
                logger.info(f"Row {row_idx}: Checking date {date_str}, year={parsed_date.year}, current_year={current_year}")
                if parsed_date.year == current_year:
                    existing = await db.holidays.find_one({"date": date_str})
                    logger.info(f"Row {row_idx}: Found existing? {existing is not None}")
                    if existing:
                        skipped.append(f"Row {row_idx}: Holiday on {date_str} already exists for current year")
                        continue
                else:
                    logger.info(f"Row {row_idx}: Skipping because year {parsed_date.year} != {current_year}")
                    skipped.append(f"Row {row_idx}: Holiday on {date_str} is not for current year ({current_year})")
                    continue
                
                holiday = Holiday(
                    date=date_str,
                    name=name,
                    description=description,
                    is_optional=is_optional
                )
                holidays_to_import.append(prepare_for_mongo(holiday.dict()))
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        # Bulk insert holidays
        imported_count = 0
        if holidays_to_import:
            result = await db.holidays.insert_many(holidays_to_import)
            imported_count = len(result.inserted_ids)
        
        return {
            "message": "Import completed",
            "imported": imported_count,
            "skipped": len(skipped),
            "errors": len(errors),
            "details": {
                "skipped_items": skipped,
                "error_items": errors
            }
        }
        
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library not installed. Please install it to import Excel files."
        )
    except Exception as e:
        logging.error(f"Error importing holidays: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import holidays: {str(e)}"
        )

@api_router.get("/holidays/export")
async def export_holidays(template: bool = False):
    """Export holidays to Excel or download template"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Holidays"
        
        # Header row
        headers = ['Date', 'Holiday Name']
        ws.append(headers)
        
        # Add instruction row
        instruction_row = ['Format: DD/MM/YYYY (e.g., 26/01/2025)', 'Enter holiday name here']
        ws.append(instruction_row)
        
        # Style header
        header_fill = PatternFill(start_color="008B74", end_color="008B74", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style instruction row with light gray
        instruction_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        instruction_font = Font(italic=True, size=9)
        for cell in ws[2]:
            cell.fill = instruction_fill
            cell.font = instruction_font
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        if not template:
            # Export existing holidays
            holidays = await db.holidays.find({}).sort("date", 1).to_list(length=None)
            for holiday in holidays:
                # Convert YYYY-MM-DD to DD/MM/YYYY for export
                date_parts = holiday['date'].split('-')
                formatted_date = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]}"
                ws.append([
                    formatted_date,
                    holiday['name']
                ])
        else:
            # Add sample data for template in DD/MM/YYYY format
            current_year = datetime.now().year
            sample_holidays = [
                [f'26/01/{current_year}', 'Republic Day'],
                [f'29/03/{current_year}', 'Holi'],
                [f'15/08/{current_year}', 'Independence Day'],
                [f'02/10/{current_year}', 'Gandhi Jayanti'],
                [f'01/11/{current_year}', 'Diwali'],
            ]
            for row in sample_holidays:
                ws.append(row)
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = "holidays_template.xlsx" if template else f"holidays_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library not installed. Please install it to export Excel files."
        )
    except Exception as e:
        logging.error(f"Error exporting holidays: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export holidays: {str(e)}"
        )

@api_router.put("/holidays/{holiday_id}")
async def update_holiday(
    holiday_id: str, 
    holiday_update: HolidayImport,
    current_user: User = Depends(get_current_user)
):
    """Update a holiday (admin only)"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    
    try:
        # Validate date format
        try:
            datetime.fromisoformat(holiday_update.date)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use ISO format (YYYY-MM-DD)"
            )
        
        # Check if another holiday exists with same date (excluding current holiday)
        existing = await db.holidays.find_one({
            "date": holiday_update.date,
            "id": {"$ne": holiday_id}
        })
        
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"A holiday already exists on {holiday_update.date}: {existing['name']}"
            )
        
        # Prepare update data
        update_data = holiday_update.dict()
        
        # Update the holiday
        result = await db.holidays.update_one(
            {"id": holiday_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Holiday not found"
            )
        
        # Fetch and return updated holiday
        updated_holiday = await db.holidays.find_one({"id": holiday_id}, {"_id": 0})
        
        return {
            "message": "Holiday updated successfully",
            "holiday": updated_holiday
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating holiday: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update holiday"
        )

@api_router.delete("/holidays/{holiday_id}")
async def delete_holiday(holiday_id: str, current_user: User = Depends(get_current_user)):
    """Delete a holiday (admin only)"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    
    try:
        result = await db.holidays.delete_one({"id": holiday_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Holiday not found"
            )
        
        return {"message": "Holiday deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting holiday: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete holiday"
        )

# Photo upload endpoints
from fastapi import File, UploadFile
import cloudinary
import cloudinary.uploader
import os

# Configure Cloudinary (mock configuration - replace with actual credentials)
cloudinary.config(
    cloud_name=os.environ.get("CLOUDINARY_CLOUD_NAME", "demo"),
    api_key=os.environ.get("CLOUDINARY_API_KEY", "demo"),
    api_secret=os.environ.get("CLOUDINARY_API_SECRET", "demo")
)

@api_router.post("/employees/upload-photo")
async def upload_employee_photo(
    photo: UploadFile = File(...),
    employee_id: str = None,
    current_user: User = Depends(get_current_user)
):
    # Determine employee_id from current user or parameter
    target_employee_id = employee_id or current_user.employee_id
    
    # For employees, they can only upload their own photo
    if current_user.role == UserRole.EMPLOYEE and target_employee_id != current_user.employee_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You can only upload your own photo"
        )
    
    try:
        # Validate file type
        if not photo.content_type.startswith('image/'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be an image"
            )
        
        # Read file content
        file_content = await photo.read()
        
        # Save file locally (for development) - replace with cloud storage in production
        import base64
        
        # Convert image to base64 for storage in database
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        photo_url = f"data:{photo.content_type};base64,{file_base64}"
        
        # Update employee record with photo URL
        result = await db.employees.update_one(
            {"employee_id": target_employee_id},
            {"$set": {"photo_url": photo_url, "updated_at": datetime.now(timezone.utc)}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee not found"
            )
        
        return {
            "message": "Photo uploaded successfully",
            "photo_url": photo_url
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error uploading photo: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to upload photo"
        )

@api_router.post("/employees/upload-photo-base64")
async def upload_employee_photo_base64(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    """Upload employee photo using base64 encoded image (from cropper)"""
    employee_id = request.get("employee_id")
    photo_base64 = request.get("photo_base64")
    
    if not employee_id or not photo_base64:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="employee_id and photo_base64 are required"
        )
    
    # For employees, they can only upload their own photo
    if current_user.role == UserRole.EMPLOYEE and employee_id != current_user.employee_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You can only upload your own photo"
        )
    
    try:
        # Update employee record with photo URL (already in base64 format)
        result = await db.employees.update_one(
            {"employee_id": employee_id},
            {"$set": {"photo_url": photo_base64, "updated_at": datetime.now(timezone.utc)}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee not found"
            )
        
        return {
            "message": "Photo uploaded successfully",
            "photo_url": photo_base64[:100] + "..."  # Return truncated URL for response
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error uploading photo: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to upload photo"
        )

@api_router.delete("/employees/{employee_id}/photo")
async def remove_employee_photo(
    employee_id: str,
    current_user: User = Depends(get_current_user)
):
    # For employees, they can only remove their own photo
    if current_user.role == UserRole.EMPLOYEE and employee_id != current_user.employee_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You can only remove your own photo"
        )
    
    try:
        # Remove photo URL from employee record
        result = await db.employees.update_one(
            {"employee_id": employee_id},
            {"$unset": {"photo_url": ""}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee not found"
            )
        
        # In real implementation, also delete from Cloudinary:
        # cloudinary.uploader.destroy(f"employees/{employee_id}")
        
        return {"message": "Photo removed successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error removing photo: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to remove photo"
        )

# Enhanced notification system
async def create_notification_helper(
    title: str, 
    message: str, 
    recipient_id: Optional[str] = None, 
    recipient_role: Optional[str] = None,
    notification_type: str = "info",
    category: str = "general",
    related_id: Optional[str] = None
):
    """Enhanced helper function to create notifications"""
    try:
        notification = Notification(
            title=title,
            message=message,
            notification_type=notification_type,
            recipient_id=recipient_id,
            recipient_role=recipient_role,
            category=category,
            related_id=related_id
        )
        notification_dict = prepare_for_mongo(notification.dict())
        await db.notifications.insert_one(notification_dict)
        
        # Send real-time notification
        await send_realtime_notification(notification_dict)
        logger.info(f"Notification created: {title} for {recipient_role}:{recipient_id}")
    except Exception as e:
        logging.error(f"Failed to create notification: {str(e)}")

# Notification categories and helpers
async def notify_leave_application(employee_id: str, employee_name: str, leave_type: str, start_date: str, end_date: str, leave_id: str = None):
    """Notify admins about new leave application"""
    await create_notification_helper(
        title="New Leave Application",
        message=f"{employee_name} ({employee_id}) applied for {leave_type} leave from {start_date} to {end_date}",
        recipient_role="admin",
        notification_type="info",
        category="leave",
        related_id=leave_id
    )

async def notify_leave_approval(employee_id: str, leave_type: str, start_date: str, end_date: str, approved: bool, admin_comment: str = ""):
    """Notify employee about leave approval/rejection"""
    status = "approved" if approved else "rejected"
    notification_type = "success" if approved else "error"
    comment_text = f" Comment: {admin_comment}" if admin_comment else ""
    
    await create_notification_helper(
        title=f"Leave {status.title()}",
        message=f"Your {leave_type} leave from {start_date} to {end_date} has been {status}.{comment_text}",
        recipient_id=employee_id,
        recipient_role="employee",
        notification_type=notification_type,
        category="leave"
    )

async def notify_loan_application(employee_id: str, employee_name: str, loan_amount: float, loan_id: str = None):
    """Notify admins about new loan application"""
    await create_notification_helper(
        title="New Loan Application",
        message=f"{employee_name} ({employee_id}) applied for a loan of ₹{loan_amount:,.2f}",
        recipient_role="admin",
        notification_type="info",
        category="loan",
        related_id=loan_id
    )

async def notify_loan_approval(employee_id: str, loan_amount: float, approved: bool, admin_comment: str = ""):
    """Notify employee about loan approval/rejection"""
    status = "approved" if approved else "rejected"
    notification_type = "success" if approved else "error"
    comment_text = f" Comment: {admin_comment}" if admin_comment else ""
    
    await create_notification_helper(
        title=f"Loan {status.title()}",
        message=f"Your loan application for ₹{loan_amount:,.2f} has been {status}.{comment_text}",
        recipient_id=employee_id,
        recipient_role="employee",
        notification_type=notification_type,
        category="loan"
    )

async def notify_payslip_generated(employee_id: str, month: str, year: int):
    """Notify employee about payslip generation"""
    await create_notification_helper(
        title="Payslip Generated",
        message=f"Your payslip for {month} {year} has been generated and is now available for download",
        recipient_id=employee_id,
        recipient_role="employee",
        notification_type="success",
        category="payslip"
    )

async def notify_payslips_bulk_generated(employee_count: int, month: str, year: int):
    """Notify admins about bulk payslip generation"""
    await create_notification_helper(
        title="Payslips Generated",
        message=f"Successfully generated payslips for {employee_count} employees for {month} {year}",
        recipient_role="admin",
        notification_type="success",
        category="payslip"
    )

async def notify_document_upload(employee_id: str, employee_name: str, document_type: str):
    """Notify admins about document upload"""
    await create_notification_helper(
        title="Document Uploaded",
        message=f"{employee_name} ({employee_id}) uploaded a new {document_type}",
        recipient_role="admin",
        notification_type="info",
        category="document"
    )

async def notify_profile_change(employee_id: str, employee_name: str, changes: str):
    """Notify admins about profile changes"""
    await create_notification_helper(
        title="Profile Updated",
        message=f"{employee_name} ({employee_id}) updated their profile: {changes}",
        recipient_role="admin",
        notification_type="info",
        category="profile"
    )

async def notify_employee_added(employee_id: str, employee_name: str, department: str):
    """Notify admins about new employee addition"""
    await create_notification_helper(
        title="New Employee Added",
        message=f"New employee {employee_name} ({employee_id}) has been added to {department} department",
        recipient_role="admin",
        notification_type="success",
        category="employee"
    )

async def notify_birthday_today(employee_id: str, employee_name: str):
    """Notify admins about employee birthdays today"""
    await create_notification_helper(
        title="Employee Birthday",
        message=f"🎉 It's {employee_name}'s birthday today! Don't forget to wish them well.",
        recipient_role="admin",
        notification_type="info",
        category="birthday"
    )

# Function to check and create birthday notifications (can be called daily)
async def check_daily_birthdays():
    """Check for today's birthdays and create notifications"""
    try:
        today = datetime.now(timezone.utc).date()
        employees = await db.employees.find({"status": "active"}).to_list(length=None)
        
        for employee in employees:
            if employee.get("date_of_birth"):
                try:
                    # Parse date_of_birth if it's a string
                    if isinstance(employee["date_of_birth"], str):
                        birth_date = datetime.fromisoformat(employee["date_of_birth"]).date()
                    else:
                        birth_date = employee["date_of_birth"]
                    
                    # Check if birthday is today
                    if birth_date.month == today.month and birth_date.day == today.day:
                        # Check if we already sent birthday notification today
                        existing_notification = await db.notifications.find_one({
                            "category": "birthday",
                            "recipient_role": "admin",
                            "message": {"$regex": employee.get('name', 'Unknown')},
                            "created_at": {
                                "$gte": datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc)
                            }
                        })
                        
                        if not existing_notification:
                            await notify_birthday_today(
                                employee_id=employee["employee_id"],
                                employee_name=employee.get('name', 'Unknown')
                            )
                except (ValueError, TypeError):
                    continue
    except Exception as e:
        logging.error(f"Error checking daily birthdays: {str(e)}")

# Leave Management Endpoints
@api_router.post("/leaves/with-document")
async def create_leave_request_with_document(
    leave_type: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    reason: str = Form(...),
    half_day: bool = Form(False),
    medical_certificate: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Create leave request with medical certificate upload for sick leave"""
    try:
        # Validate file type
        allowed_types = ['application/pdf', 'image/jpeg', 'image/png', 'image/jpg']
        if medical_certificate.content_type not in allowed_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only PDF, JPG, and PNG files are allowed for medical certificates"
            )
        
        # Validate file size (5MB limit)
        max_size = 5 * 1024 * 1024  # 5MB
        file_content = await medical_certificate.read()
        if len(file_content) > max_size:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File size must be less than 5MB"
            )
        
        # Store file as base64 in database for simplicity
        import base64
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        medical_cert_data = {
            "filename": medical_certificate.filename,
            "content_type": medical_certificate.content_type,
            "file_data": f"data:{medical_certificate.content_type};base64,{file_base64}"
        }
        
        # Parse dates
        start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Calculate working days only (excluding weekends and holidays)
        if half_day:
            days = 0.5
        else:
            # Fetch working days config and holidays
            settings = await db.settings.find_one({"id": "system_settings"})
            working_days_config = settings.get("working_days_config", {
                "saturday_policy": "alternate",
                "off_saturdays": [1, 3],
                "sunday_off": True
            }) if settings else {
                "saturday_policy": "alternate",
                "off_saturdays": [1, 3],
                "sunday_off": True
            }
            
            holidays = await db.holidays.find({}).to_list(length=None)
            holiday_dates = [h["date"] for h in holidays]
            
            # Helper function to check if date is working day
            def is_working_day(check_date):
                day_of_week = check_date.weekday()
                
                # Check if it's a holiday
                if check_date.isoformat() in holiday_dates:
                    return False
                
                # Check Sunday
                if day_of_week == 6:  # Sunday
                    if working_days_config.get("sunday_off", True):
                        return False
                
                # Check Saturday
                if day_of_week == 5:  # Saturday
                    saturday_policy = working_days_config.get("saturday_policy", "alternate")
                    if saturday_policy == "all_off":
                        return False
                    elif saturday_policy == "alternate" or saturday_policy == "custom":
                        # Calculate which Saturday of the month
                        first_saturday = date(check_date.year, check_date.month, 1)
                        while first_saturday.weekday() != 5:
                            first_saturday = first_saturday + timedelta(days=1)
                        saturday_number = ((check_date - first_saturday).days // 7) + 1
                        off_saturdays = working_days_config.get("off_saturdays", [1, 3])
                        if saturday_number in off_saturdays:
                            return False
                
                return True
            
            # Helper function to get holiday name
            def get_holiday_name(check_date):
                date_str = check_date.isoformat()
                for holiday in holidays:
                    if holiday["date"] == date_str:
                        return holiday.get("name", "Holiday")
                return None
            
            # Validate start date is a working day
            if not is_working_day(start_date_parsed):
                holiday_name = get_holiday_name(start_date_parsed)
                if holiday_name:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Start date falls on a holiday: {holiday_name}. Please select a working day."
                    )
                else:
                    day_name = start_date_parsed.strftime('%A')
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Start date falls on {day_name} which is a non-working day. Please select a working day."
                    )
            
            # Validate end date is a working day
            if not is_working_day(end_date_parsed):
                holiday_name = get_holiday_name(end_date_parsed)
                if holiday_name:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"End date falls on a holiday: {holiday_name}. Please select a working day."
                    )
                else:
                    day_name = end_date_parsed.strftime('%A')
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"End date falls on {day_name} which is a non-working day. Please select a working day."
                    )
            
            # Count working days between start and end dates (inclusive)
            working_days_count = 0
            current_date = start_date_parsed
            while current_date <= end_date_parsed:
                if is_working_day(current_date):
                    working_days_count += 1
                current_date = current_date + timedelta(days=1)
            
            # Validate that there's at least one working day
            if working_days_count == 0:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The selected date range contains no working days. Please select dates that include at least one working day."
                )
            
            days = working_days_count
        
        leave_data = {
            "id": str(uuid.uuid4()),
            "employee_id": current_user.employee_id,
            "leave_type": leave_type,
            "start_date": start_date_parsed.isoformat(),
            "end_date": end_date_parsed.isoformat(),
            "days": days,
            "reason": reason,
            "half_day": half_day,
            "status": "pending",
            "applied_date": datetime.now(timezone.utc).isoformat(),
            "medical_certificate": medical_cert_data,  # Store medical certificate data
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Insert into database
        result = await db.leave_requests.insert_one(leave_data)
        
        # Create notification for admin
        await create_notification_helper(
            title="New Leave Request",
            message=f"Employee {current_user.employee_id} has applied for {leave_type} from {start_date} to {end_date}",
            recipient_id="admin",
            category="leave_request"
        )
        
        # Return the created leave request
        created_leave = await db.leave_requests.find_one({"_id": result.inserted_id})
        return LeaveRequest(**created_leave)
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating leave request with document: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create leave request: {str(e)}"
        )

@api_router.post("/leaves", response_model=LeaveRequest)
async def create_leave_request(
    leave_data: LeaveRequestCreate,
    current_user: User = Depends(get_current_user)
):
    try:
        # Calculate working days only (excluding weekends and holidays)
        start = leave_data.start_date
        end = leave_data.end_date
        
        if leave_data.half_day:
            days = 0.5
        else:
            # Fetch working days config and holidays
            settings = await db.settings.find_one({"id": "system_settings"})
            working_days_config = settings.get("working_days_config", {
                "saturday_policy": "alternate",
                "off_saturdays": [1, 3],
                "sunday_off": True
            }) if settings else {
                "saturday_policy": "alternate",
                "off_saturdays": [1, 3],
                "sunday_off": True
            }
            
            holidays = await db.holidays.find({}).to_list(length=None)
            holiday_dates = [h["date"] for h in holidays]
            
            # Helper function to check if date is working day
            def is_working_day(check_date):
                day_of_week = check_date.weekday()
                
                # Check if it's a holiday
                if check_date.isoformat() in holiday_dates:
                    return False
                
                # Check Sunday
                if day_of_week == 6:  # Sunday
                    if working_days_config.get("sunday_off", True):
                        return False
                
                # Check Saturday
                if day_of_week == 5:  # Saturday
                    saturday_policy = working_days_config.get("saturday_policy", "alternate")
                    if saturday_policy == "all_off":
                        return False
                    elif saturday_policy == "alternate" or saturday_policy == "custom":
                        # Calculate which Saturday of the month
                        first_saturday = date(check_date.year, check_date.month, 1)
                        while first_saturday.weekday() != 5:
                            first_saturday = first_saturday + timedelta(days=1)
                        saturday_number = ((check_date - first_saturday).days // 7) + 1
                        off_saturdays = working_days_config.get("off_saturdays", [1, 3])
                        if saturday_number in off_saturdays:
                            return False
                
                return True
            
            # Helper function to get holiday name
            def get_holiday_name(check_date):
                date_str = check_date.isoformat()
                for holiday in holidays:
                    if holiday["date"] == date_str:
                        return holiday.get("name", "Holiday")
                return None
            
            # Validate start date is a working day
            if not is_working_day(start):
                holiday_name = get_holiday_name(start)
                if holiday_name:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Start date falls on a holiday: {holiday_name}. Please select a working day."
                    )
                else:
                    day_name = start.strftime('%A')
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Start date falls on {day_name} which is a non-working day. Please select a working day."
                    )
            
            # Validate end date is a working day
            if not is_working_day(end):
                holiday_name = get_holiday_name(end)
                if holiday_name:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"End date falls on a holiday: {holiday_name}. Please select a working day."
                    )
                else:
                    day_name = end.strftime('%A')
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"End date falls on {day_name} which is a non-working day. Please select a working day."
                    )
            
            # Count working days between start and end dates (inclusive)
            working_days_count = 0
            current_date = start
            while current_date <= end:
                if is_working_day(current_date):
                    working_days_count += 1
                current_date = current_date + timedelta(days=1)
            
            # Validate that there's at least one working day
            if working_days_count == 0:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The selected date range contains no working days. Please select dates that include at least one working day."
                )
            
            days = working_days_count
        
        leave_request = LeaveRequest(
            employee_id=current_user.employee_id or current_user.username,
            leave_type=leave_data.leave_type,
            start_date=leave_data.start_date,
            end_date=leave_data.end_date,
            days=days,
            reason=leave_data.reason,
            half_day=leave_data.half_day
        )
        
        leave_dict = prepare_for_mongo(leave_request.dict())
        result = await db.leave_requests.insert_one(leave_dict)
        
        if result.inserted_id:
            # Get employee details for notification
            employee = await db.employees.find_one({"employee_id": current_user.employee_id or current_user.username})
            employee_name = employee.get('name', 'Unknown') if employee else current_user.username
            
            # Notify admins about new leave application
            await notify_leave_application(
                employee_id=current_user.employee_id or current_user.username,
                employee_name=employee_name,
                leave_type=leave_data.leave_type,
                start_date=leave_data.start_date.strftime('%Y-%m-%d'),
                end_date=leave_data.end_date.strftime('%Y-%m-%d'),
                leave_id=leave_request.id
            )
            
            return leave_request
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create leave request"
            )
    except Exception as e:
        logging.error(f"Error creating leave request: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create leave request"
        )

@api_router.get("/leaves")
async def get_leave_requests(
    current_user: User = Depends(get_current_user)
):
    try:
        if current_user.role == UserRole.ADMIN:
            # Admin can see all leave requests
            leaves = await db.leave_requests.find({}).to_list(length=None)
        else:
            # Employee can only see their own requests
            leaves = await db.leave_requests.find(
                {"employee_id": current_user.employee_id}
            ).to_list(length=None)
        
        # Convert MongoDB documents to JSON-serializable format
        serialized_leaves = [prepare_from_mongo(leave) for leave in leaves]
        
        return serialized_leaves
    except Exception as e:
        logging.error(f"Error fetching leave requests: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch leave requests"
        )

@api_router.get("/leaves/{leave_id}/medical-certificate")
async def get_medical_certificate(
    leave_id: str, 
    current_user: User = Depends(get_current_user)
):
    """Get medical certificate for a leave request"""
    try:
        # Find the leave request
        leave_request = await db.leave_requests.find_one({"id": leave_id})
        if not leave_request:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Leave request not found"
            )
        
        # Check permissions - admin or the employee who submitted the request
        if current_user.role != UserRole.ADMIN and current_user.employee_id != leave_request.get("employee_id"):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't have permission to view this medical certificate"
            )
        
        # Check if medical certificate exists
        medical_cert = leave_request.get("medical_certificate")
        if not medical_cert:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No medical certificate found for this leave request"
            )
        
        # Return the medical certificate data
        return {
            "filename": medical_cert.get("filename", "medical_certificate"),
            "content_type": medical_cert.get("content_type", "application/octet-stream"),
            "file_data": medical_cert.get("file_data", ""),
            "leave_id": leave_id,
            "employee_id": leave_request.get("employee_id"),
            "leave_type": leave_request.get("leave_type")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error retrieving medical certificate: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retrieve medical certificate: {str(e)}"
        )

@api_router.put("/leaves/{leave_id}/approve")
async def approve_reject_leave(
    leave_id: str,
    approval_data: LeaveApprovalRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    try:
        update_data = {
            "status": approval_data.status,
            "updated_at": datetime.now(timezone.utc)
        }
        
        if approval_data.status == "approved":
            update_data["approved_by"] = current_user.username
            update_data["approved_date"] = datetime.now(timezone.utc)
            if approval_data.admin_comment:
                update_data["admin_comment"] = approval_data.admin_comment
        elif approval_data.status == "rejected":
            update_data["rejected_by"] = current_user.username
            update_data["rejected_date"] = datetime.now(timezone.utc)
            if approval_data.admin_comment:
                update_data["admin_comment"] = approval_data.admin_comment
            # Keep backward compatibility with rejection_reason
            if approval_data.admin_comment:
                update_data["rejection_reason"] = approval_data.admin_comment
        
        # Get leave request data before updating for notification
        leave_request = await db.leave_requests.find_one({"id": leave_id})
        if not leave_request:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Leave request not found"
            )
        
        # Update leave request
        result = await db.leave_requests.update_one(
            {"id": leave_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Leave request not found"
            )
        
        # Send notification to employee using enhanced system
        await notify_leave_approval(
            employee_id=leave_request['employee_id'],
            leave_type=leave_request['leave_type'],
            start_date=leave_request['start_date'],
            end_date=leave_request['end_date'],
            approved=(approval_data.status == "approved"),
            admin_comment=approval_data.admin_comment or ""
        )
        
        # Mark the leave application notification as read using related_id
        await db.notifications.update_many(
            {
                "related_id": leave_id,
                "category": "leave",
                "is_read": False
            },
            {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": f"Leave request {approval_data.status} successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating leave request: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update leave request"
        )

@api_router.put("/leaves/{leave_id}/cancel")
async def cancel_leave_request(
    leave_id: str,
    cancellation_data: LeaveCancellationRequest,
    current_user: User = Depends(get_current_user)
):
    """Cancel a leave request (employee can cancel their own pending or approved future leaves)"""
    try:
        # Get leave request
        leave_request = await db.leave_requests.find_one({"id": leave_id})
        if not leave_request:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Leave request not found"
            )
        
        # Verify the leave belongs to the current user
        if leave_request['employee_id'] != current_user.employee_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only cancel your own leave requests"
            )
        
        # Check if leave can be cancelled
        current_status = leave_request.get('status')
        if current_status == 'cancelled':
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Leave request is already cancelled"
            )
        
        if current_status == 'rejected':
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot cancel a rejected leave request"
            )
        
        # Check if leave has started or is in the past
        start_date = leave_request['start_date']
        if isinstance(start_date, str):
            start_date = datetime.fromisoformat(start_date).date()
        
        today = date.today()
        if start_date < today:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot cancel a leave that has already started or is in the past"
            )
        
        # Update leave request to cancelled
        update_data = {
            "status": "cancelled",
            "cancelled_by": current_user.employee_id,
            "cancelled_date": datetime.now(timezone.utc),
            "cancellation_reason": cancellation_data.cancellation_reason,
            "updated_at": datetime.now(timezone.utc)
        }
        
        result = await db.leave_requests.update_one(
            {"id": leave_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Leave request not found"
            )
        
        # Get employee details for notification
        employee = await db.employees.find_one({"employee_id": current_user.employee_id})
        employee_name = employee.get('name', 'Unknown') if employee else current_user.username
        
        # Notify admin about cancellation (especially important for approved leaves)
        if current_status == 'approved':
            notification_title = "Approved Leave Cancelled"
            notification_message = f"Employee {employee_name} ({current_user.employee_id}) has cancelled their approved {leave_request['leave_type']} from {start_date} to {leave_request['end_date']}. Reason: {cancellation_data.cancellation_reason}"
        else:
            notification_title = "Leave Request Cancelled"
            notification_message = f"Employee {employee_name} ({current_user.employee_id}) has cancelled their pending {leave_request['leave_type']} from {start_date} to {leave_request['end_date']}. Reason: {cancellation_data.cancellation_reason}"
        
        # Create notification for admin
        await create_notification_helper(
            title=notification_title,
            message=notification_message,
            recipient_id="admin",
            recipient_role="admin",
            category="leave_cancellation"
        )
        
        return {
            "message": "Leave request cancelled successfully",
            "was_approved": current_status == 'approved'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error cancelling leave request: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to cancel leave request: {str(e)}"
        )

@api_router.delete("/leaves/clear-all")
async def clear_all_leave_requests(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Delete all leave requests (admin only)"""
    try:
        result = await db.leave_requests.delete_many({})
        return {
            "message": f"Successfully deleted all leave requests",
            "deleted_count": result.deleted_count
        }
    except Exception as e:
        logging.error(f"Error clearing leave requests: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to clear leave requests"
        )

# Leave Entitlement Endpoints
@api_router.get("/leaves/entitlement/{employee_id}", response_model=LeaveEntitlementResponse)
async def get_leave_entitlement(
    employee_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get current leave entitlement and balance for an employee"""
    try:
        # Fetch employee
        employee = await db.employees.find_one({"employee_id": employee_id})
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee not found"
            )
        
        # Check probation status - checkbox takes priority
        is_in_probation = employee.get('is_on_probation', False)
        
        # If not explicitly marked, check probation end date
        if not is_in_probation:
            probation_end_date = employee.get('probation_end_date')
            if probation_end_date:
                if isinstance(probation_end_date, str):
                    probation_end_date = datetime.fromisoformat(probation_end_date).date()
                is_in_probation = date.today() < probation_end_date
        
        # If in probation, no leaves
        if is_in_probation:
            months_of_service = 0
            casual_leave_accrued = 0.0
            sick_leave_total = 0.0
            annual_leave_total = 0.0
        else:
            # Calculate casual leave starting from PROBATION END DATE
            joining_date = employee.get('date_of_joining')
            probation_end_date = employee.get('probation_end_date')
            today = date.today()
            current_year_start = date(today.year, 1, 1)
            
            if not joining_date:
                months_of_service = 0
                casual_leave_accrued = 0.0
            else:
                if isinstance(joining_date, str):
                    joining_date = datetime.fromisoformat(joining_date).date()
                
                # Convert probation end date if exists
                if probation_end_date:
                    if isinstance(probation_end_date, str):
                        probation_end_date = datetime.fromisoformat(probation_end_date).date()
                
                # Leave accrual starts from probation end date (if set), otherwise from joining date
                leave_accrual_start = probation_end_date if probation_end_date else joining_date
                
                # Calculate months in current year only
                # If leave accrual started this year, count from that date, otherwise from Jan 1
                accrual_start_this_year = max(leave_accrual_start, current_year_start)
                
                # Only accrue if probation has ended
                if today >= leave_accrual_start:
                    # Months from accrual start to today in current year
                    months_this_year = (today.year - accrual_start_this_year.year) * 12 + (today.month - accrual_start_this_year.month)
                    
                    # Add partial month if mid-month
                    if today.day >= accrual_start_this_year.day:
                        months_this_year += 1
                    
                    # Use custom casual leave rate if set, otherwise default 1.5 days per month
                    casual_rate = employee.get('custom_casual_leave_per_month', 1.5)
                    # Accrual is only for current year
                    casual_leave_accrued = round(months_this_year * casual_rate, 1)
                else:
                    # Still in probation, no accrual
                    casual_leave_accrued = 0.0
                
                # Total months of service (for display purposes - from joining date)
                months_of_service = (today.year - joining_date.year) * 12 + (today.month - joining_date.month)
            
            # Use custom sick leave if set, otherwise default 7 days per year
            sick_leave_total = employee.get('custom_sick_leave_per_year', 7.0)
            
            # Additional annual leave days
            annual_leave_total = employee.get('annual_leave_days', 0.0)
        
        # Get current year
        current_year = date.today().year
        
        # Get or create leave balance for current year
        leave_balance = await db.leave_balances.find_one({
            "employee_id": employee_id,
            "year": current_year
        })
        
        if not leave_balance:
            # Calculate carried forward from previous year (only unused Annual Leave, max 5)
            previous_year_balance = await db.leave_balances.find_one({
                "employee_id": employee_id,
                "year": current_year - 1
            })
            
            carried_forward_from_prev = 0.0
            if previous_year_balance:
                # Get previous year's annual leave
                prev_annual_total = previous_year_balance.get('annual_leave_total', 0.0)
                prev_annual_used = previous_year_balance.get('annual_leave_used', 0.0)
                prev_annual_unused = max(0, prev_annual_total - prev_annual_used)
                # Carry forward only Annual Leave, capped at 5 days
                carried_forward_from_prev = min(prev_annual_unused, 5.0)
            
            # Create new leave balance record
            leave_balance = {
                "id": str(uuid.uuid4()),
                "employee_id": employee_id,
                "year": current_year,
                "casual_leave_accrued": casual_leave_accrued,
                "casual_leave_used": 0.0,
                "casual_leave_balance": casual_leave_accrued,
                "sick_leave_total": sick_leave_total,
                "sick_leave_used": 0.0,
                "sick_leave_balance": sick_leave_total,
                "carried_forward_leaves": carried_forward_from_prev,
                "annual_leave_total": annual_leave_total,
                "annual_leave_used": 0.0,
                "last_accrual_date": datetime.now(timezone.utc).date().isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.leave_balances.insert_one(leave_balance)
        
        # Calculate used leaves from approved leave requests
        current_year_start = date(current_year, 1, 1)
        current_year_end = date(current_year, 12, 31)
        
        approved_leaves = await db.leave_requests.find({
            "employee_id": employee_id,
            "status": "approved",
            "start_date": {
                "$gte": current_year_start.isoformat(),
                "$lte": current_year_end.isoformat()
            }
        }).to_list(length=None)
        
        casual_leave_used = 0.0
        sick_leave_used = 0.0
        annual_leave_used = 0.0
        
        for leave in approved_leaves:
            days = leave.get('days', 0.0)
            leave_type = leave.get('leave_type', '').lower()
            
            if 'casual' in leave_type:
                casual_leave_used += days
            elif 'sick' in leave_type:
                sick_leave_used += days
            elif 'annual' in leave_type:
                annual_leave_used += days
        
        # Update balances
        casual_leave_balance = max(0, casual_leave_accrued - casual_leave_used)
        sick_leave_balance = max(0, sick_leave_total - sick_leave_used)
        
        # Annual leave balance includes current year's allocation + carried forward - used
        carried_forward = min(leave_balance.get('carried_forward_leaves', 0.0), 5.0)
        annual_leave_balance = max(0, annual_leave_total + carried_forward - annual_leave_used)
        
        total_available = casual_leave_balance + sick_leave_balance + annual_leave_balance
        
        # Update leave balance in database
        await db.leave_balances.update_one(
            {"employee_id": employee_id, "year": current_year},
            {"$set": {
                "casual_leave_accrued": casual_leave_accrued,
                "casual_leave_used": casual_leave_used,
                "casual_leave_balance": casual_leave_balance,
                "sick_leave_total": sick_leave_total,
                "sick_leave_used": sick_leave_used,
                "sick_leave_balance": sick_leave_balance,
                "annual_leave_total": annual_leave_total,
                "annual_leave_used": annual_leave_used,
                "annual_leave_balance": annual_leave_balance,
                "carried_forward_leaves": carried_forward,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return LeaveEntitlementResponse(
            employee_id=employee_id,
            employee_name=employee.get('name', 'Unknown'),
            joining_date=joining_date if not is_in_probation else None,
            months_of_service=months_of_service,
            casual_leave_accrued=casual_leave_accrued,
            casual_leave_used=casual_leave_used,
            casual_leave_balance=casual_leave_balance,
            sick_leave_total=sick_leave_total,
            sick_leave_used=sick_leave_used,
            sick_leave_balance=sick_leave_balance,
            annual_leave_total=annual_leave_total,
            annual_leave_used=annual_leave_used,
            annual_leave_balance=annual_leave_balance,
            carried_forward_leaves=carried_forward,
            total_available_leaves=total_available
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching leave entitlement: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch leave entitlement"
        )

@api_router.get("/leaves/approved-by-month")
async def get_approved_leaves_by_month(
    month: int,
    year: int,
    current_user: User = Depends(get_current_user)
):
    """Get approved leaves for all employees for a specific month, returning only excess leaves beyond entitlement"""
    try:
        # Get start and end dates for the month
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Fetch all approved leave requests that overlap with this month
        approved_leaves = await db.leave_requests.find({
            "status": "approved",
            "$or": [
                {"start_date": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}},
                {"end_date": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}},
                {
                    "$and": [
                        {"start_date": {"$lte": month_start.isoformat()}},
                        {"end_date": {"$gte": month_end.isoformat()}}
                    ]
                }
            ]
        }).to_list(length=None)
        
        # Group by employee and calculate days in the specific month
        employee_leaves = {}
        
        for leave in approved_leaves:
            employee_id = leave['employee_id']
            leave_type = leave.get('leave_type', '').lower()
            
            # Parse dates
            start_date = leave['start_date']
            end_date = leave['end_date']
            if isinstance(start_date, str):
                start_date = datetime.fromisoformat(start_date).date()
            if isinstance(end_date, str):
                end_date = datetime.fromisoformat(end_date).date()
            
            # Calculate overlap with the target month
            overlap_start = max(start_date, month_start)
            overlap_end = min(end_date, month_end)
            
            if overlap_start <= overlap_end:
                # Calculate days in this month
                days_in_month = (overlap_end - overlap_start).days + 1
                
                # Adjust for half day
                if leave.get('half_day', False) and start_date == end_date:
                    days_in_month = 0.5
                
                if employee_id not in employee_leaves:
                    employee_leaves[employee_id] = {
                        'casual': 0.0,
                        'sick': 0.0,
                        'other': 0.0
                    }
                
                # Categorize leave type
                # Casual: casual, annual, earned, privilege
                # Sick: sick, medical
                if any(word in leave_type for word in ['casual', 'annual', 'earned', 'privilege']):
                    employee_leaves[employee_id]['casual'] += days_in_month
                elif any(word in leave_type for word in ['sick', 'medical']):
                    employee_leaves[employee_id]['sick'] += days_in_month
                else:
                    employee_leaves[employee_id]['other'] += days_in_month
        
        # Now calculate excess leaves for each employee
        result = {}
        
        for employee_id, leave_data in employee_leaves.items():
            # Get employee entitlement
            employee = await db.employees.find_one({"employee_id": employee_id})
            if not employee:
                continue
            
            # Check probation status - checkbox takes priority
            is_in_probation = employee.get('is_on_probation', False)
            
            # If not explicitly marked, check probation end date
            if not is_in_probation:
                probation_end_date = employee.get('probation_end_date')
                if probation_end_date:
                    if isinstance(probation_end_date, str):
                        probation_end_date = datetime.fromisoformat(probation_end_date).date()
                    is_in_probation = date.today() < probation_end_date
            
            # If in probation, all leaves are excess (unpaid)
            if is_in_probation:
                casual_entitled = 0.0
                sick_entitled = 0.0
            else:
                # Calculate entitlement for this employee FOR CURRENT YEAR ONLY
                joining_date = employee.get('date_of_joining')
                current_year_start = date(year, 1, 1)
                
                if not joining_date:
                    casual_entitled = 0.0
                else:
                    if isinstance(joining_date, str):
                        joining_date = datetime.fromisoformat(joining_date).date()
                    
                    # Calculate months in current year only
                    accrual_start = max(joining_date, current_year_start)
                    today = date.today()
                    
                    # Months from accrual start to today in current year
                    months_this_year = (today.year - accrual_start.year) * 12 + (today.month - accrual_start.month)
                    if today.day > accrual_start.day:
                        months_this_year += 1
                    
                    # Use custom casual leave rate if set, otherwise default 1.5 days per month
                    casual_rate = employee.get('custom_casual_leave_per_month', 1.5)
                    casual_entitled = round(months_this_year * casual_rate, 1)
                
                # Use custom sick leave if set, otherwise default 7 days per year
                sick_entitled = employee.get('custom_sick_leave_per_year', 7.0)
            
            # Get year-to-date usage up to (but not including) this month
            year_start = date(year, 1, 1)
            prev_month_end = month_start - timedelta(days=1)
            
            ytd_leaves = await db.leave_requests.find({
                "employee_id": employee_id,
                "status": "approved",
                "start_date": {"$gte": year_start.isoformat(), "$lte": prev_month_end.isoformat()}
            }).to_list(length=None)
            
            # Map leave types to categories
            # Casual: casual, annual, earned, privilege
            # Sick: sick, medical
            ytd_casual_used = sum(l.get('days', 0.0) for l in ytd_leaves 
                                 if any(word in l.get('leave_type', '').lower() 
                                       for word in ['casual', 'annual', 'earned', 'privilege']))
            ytd_sick_used = sum(l.get('days', 0.0) for l in ytd_leaves 
                               if any(word in l.get('leave_type', '').lower() 
                                     for word in ['sick', 'medical']))
            
            # Calculate remaining entitlement
            casual_remaining = max(0, casual_entitled - ytd_casual_used)
            sick_remaining = max(0, sick_entitled - ytd_sick_used)
            
            # Calculate excess leaves in this month
            casual_excess = max(0, leave_data['casual'] - casual_remaining)
            sick_excess = max(0, leave_data['sick'] - sick_remaining)
            other_excess = leave_data['other']  # Other leaves are always deducted
            
            total_excess = casual_excess + sick_excess + other_excess
            
            result[employee_id] = {
                'total_excess_days': round(total_excess, 1),
                'casual_excess': round(casual_excess, 1),
                'sick_excess': round(sick_excess, 1),
                'other_days': round(other_excess, 1),
                'casual_taken': round(leave_data['casual'], 1),
                'sick_taken': round(leave_data['sick'], 1),
                'casual_entitled': round(casual_entitled, 1),
                'sick_entitled': sick_entitled,
                'casual_remaining_before': round(casual_remaining, 1),
                'sick_remaining_before': round(sick_remaining, 1)
            }
        
        return result
        
    except Exception as e:
        logging.error(f"Error fetching approved leaves by month: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch approved leaves"
        )


# OT (Overtime) Management Endpoints
@api_router.post("/ot/log")
async def log_overtime(
    ot_data: OTLogCreate,
    current_user: User = Depends(get_current_user)
):
    """Employee logs overtime hours with time range validation"""
    try:
        # Validate that OT can only be logged for today
        today = date.today()
        if ot_data.date != today:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"OT can only be logged for today's date ({today.strftime('%Y-%m-%d')}). Cannot log OT for past or future dates."
            )
        
        # Parse times
        from_hour, from_min = map(int, ot_data.from_time.split(':'))
        to_hour, to_min = map(int, ot_data.to_time.split(':'))
        from_minutes = from_hour * 60 + from_min
        to_minutes = to_hour * 60 + to_min
        
        # Validate time range
        if from_minutes >= to_minutes:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="End time must be after start time"
            )
        
        # Check if date is a working day or off day (weekend/holiday)
        settings = await db.settings.find_one({"id": "system_settings"})
        working_days_config = settings.get("working_days_config", {
            "saturday_policy": "alternate",
            "off_saturdays": [1, 3],
            "sunday_off": True
        }) if settings else {
            "saturday_policy": "alternate",
            "off_saturdays": [1, 3],
            "sunday_off": True
        }
        
        holidays = await db.holidays.find({}).to_list(length=None)
        holiday_dates = [h["date"] for h in holidays]
        
        # Helper function to check if date is working day
        def is_working_day(check_date):
            day_of_week = check_date.weekday()
            
            # Check if it's a holiday
            if check_date.isoformat() in holiday_dates:
                return False
            
            # Check Sunday
            if day_of_week == 6:  # Sunday
                if working_days_config.get("sunday_off", True):
                    return False
            
            # Check Saturday
            if day_of_week == 5:  # Saturday
                saturday_policy = working_days_config.get("saturday_policy", "alternate")
                if saturday_policy == "all_off":
                    return False
                elif saturday_policy == "alternate" or saturday_policy == "custom":
                    # Calculate which Saturday of the month
                    first_saturday = date(check_date.year, check_date.month, 1)
                    while first_saturday.weekday() != 5:
                        first_saturday = first_saturday + timedelta(days=1)
                    saturday_number = ((check_date - first_saturday).days // 7) + 1
                    off_saturdays = working_days_config.get("off_saturdays", [1, 3])
                    if saturday_number in off_saturdays:
                        return False
            
            return True
        
        is_work_day = is_working_day(ot_data.date)
        
        # Only validate office hours for working days
        # On weekends/holidays, allow OT for any time
        if is_work_day:
            # Office hours: 8:30 AM (510 minutes) to 5:30 PM (1050 minutes)
            office_start = 8 * 60 + 30  # 510 minutes (8:30 AM)
            office_end = 17 * 60 + 30   # 1050 minutes (5:30 PM)
            
            # Validate OT is outside office hours on working days
            if from_minutes >= office_start and from_minutes < office_end:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="On working days, OT start time must be before 8:30 AM or after 5:30 PM"
                )
            
            if to_minutes > office_start and to_minutes <= office_end:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="On working days, OT end time must be before 8:30 AM or after 5:30 PM"
                )
        
        # Calculate OT hours
        ot_hours = (to_minutes - from_minutes) / 60.0
        
        # Check for duplicate/overlapping OT entries on the same date
        existing_ot = await db.ot_logs.find({
            "employee_id": current_user.username,
            "date": ot_data.date.isoformat()
        }).to_list(length=None)
        
        for existing in existing_ot:
            existing_from = existing.get("from_time", "")
            existing_to = existing.get("to_time", "")
            
            if existing_from and existing_to:
                ex_from_hour, ex_from_min = map(int, existing_from.split(':'))
                ex_to_hour, ex_to_min = map(int, existing_to.split(':'))
                ex_from_minutes = ex_from_hour * 60 + ex_from_min
                ex_to_minutes = ex_to_hour * 60 + ex_to_min
                
                # Check for overlap
                if not (to_minutes <= ex_from_minutes or from_minutes >= ex_to_minutes):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"OT time overlaps with existing entry ({existing_from} - {existing_to})"
                    )
        
        ot_log = OTLog(
            employee_id=current_user.username,
            date=ot_data.date,
            from_time=ot_data.from_time,
            to_time=ot_data.to_time,
            ot_hours=round(ot_hours, 2),
            project=ot_data.project,
            notes=ot_data.notes,
            status="pending"
        )
        
        ot_dict = prepare_for_mongo(ot_log.dict())
        await db.ot_logs.insert_one(ot_dict)
        
        # Fetch employee name for notification
        employee = await db.employees.find_one({"employee_id": current_user.username})
        employee_name = employee.get('name', current_user.username) if employee else current_user.username
        
        # Create notification for admin
        notification = Notification(
            recipient_role="admin",
            title="New OT Log Submitted",
            message=f"{employee_name} has logged {round(ot_hours, 2)} hours of overtime ({ot_data.from_time} - {ot_data.to_time}) for {ot_data.date} (Project: {ot_data.project}).",
            category="ot",
            related_id=ot_log.id,
            notification_type="info"
        )
        notification_dict = prepare_for_mongo(notification.dict())
        await db.notifications.insert_one(notification_dict)
        
        toast_message = f"OT logged successfully: {round(ot_hours, 2)} hours ({ot_data.from_time} - {ot_data.to_time})"
        return {"message": toast_message, "ot_log": prepare_from_mongo(ot_dict)}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error logging OT: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to log overtime: {str(e)}"
        )

@api_router.get("/ot/my-ot")
async def get_my_ot_logs(current_user: User = Depends(get_current_user)):
    """Employee view their OT logs"""
    try:
        ot_logs = await db.ot_logs.find({"employee_id": current_user.username}).to_list(length=None)
        return [prepare_from_mongo(log) for log in ot_logs]
    except Exception as e:
        logging.error(f"Error fetching OT logs: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch OT logs"
        )

@api_router.post("/ot/admin-log")
async def admin_log_overtime(
    ot_data: AdminOTLogCreate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Admin logs overtime for any employee on any date (no date restrictions)"""
    try:
        # Verify employee exists
        employee = await db.employees.find_one({"employee_id": ot_data.employee_id})
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee {ot_data.employee_id} not found"
            )
        
        # Parse times
        from_hour, from_min = map(int, ot_data.from_time.split(':'))
        to_hour, to_min = map(int, ot_data.to_time.split(':'))
        from_minutes = from_hour * 60 + from_min
        to_minutes = to_hour * 60 + to_min
        
        # Validate time range
        if from_minutes >= to_minutes:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="End time must be after start time"
            )
        
        # Calculate OT hours
        ot_hours = (to_minutes - from_minutes) / 60.0
        
        # Check for duplicate/overlapping OT entries on the same date for this employee
        existing_ot = await db.ot_logs.find({
            "employee_id": ot_data.employee_id,
            "date": ot_data.date.isoformat()
        }).to_list(length=None)
        
        for existing in existing_ot:
            existing_from = existing.get("from_time", "")
            existing_to = existing.get("to_time", "")
            
            if existing_from and existing_to:
                ex_from_h, ex_from_m = map(int, existing_from.split(':'))
                ex_to_h, ex_to_m = map(int, existing_to.split(':'))
                ex_from_minutes = ex_from_h * 60 + ex_from_m
                ex_to_minutes = ex_to_h * 60 + ex_to_m
                
                # Check for overlap
                if not (to_minutes <= ex_from_minutes or from_minutes >= ex_to_minutes):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"OT time overlaps with existing entry ({existing_from} - {existing_to})"
                    )
        
        # Create OT log
        ot_log = OTLog(
            employee_id=ot_data.employee_id,
            date=ot_data.date,
            from_time=ot_data.from_time,
            to_time=ot_data.to_time,
            ot_hours=ot_hours,
            project=ot_data.project,
            notes=ot_data.notes,
            status="approved",  # Admin-logged OT is auto-approved
            approved_by=current_user.username,
            approved_date=datetime.now(timezone.utc)
        )
        
        ot_dict = prepare_for_mongo(ot_log.dict())
        await db.ot_logs.insert_one(ot_dict)
        
        # Create notification for employee
        employee_name = employee.get('name', 'Unknown')
        notification = Notification(
            recipient_id=ot_data.employee_id,
            recipient_role="employee",
            title="OT Logged by Admin",
            message=f"Admin has logged {round(ot_hours, 2)} hours of overtime for you on {ot_data.date.strftime('%d %b %Y')} ({ot_data.from_time} - {ot_data.to_time}) for project: {ot_data.project}",
            category="ot_log",
            related_id=ot_log.id,
            notification_type="info"
        )
        notification_dict = prepare_for_mongo(notification.dict())
        await db.notifications.insert_one(notification_dict)
        
        # Send real-time notification
        await send_realtime_notification(notification_dict)
        
        return {
            "message": f"OT logged successfully for {employee_name}: {round(ot_hours, 2)} hours",
            "ot_log": prepare_from_mongo(ot_dict)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error logging OT by admin: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to log overtime: {str(e)}"
        )

@api_router.get("/ot/all")
async def get_all_ot_logs(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Admin view all OT logs"""
    try:
        ot_logs = await db.ot_logs.find({}).to_list(length=None)
        
        # Enrich with employee details
        for log in ot_logs:
            employee = await db.employees.find_one({"employee_id": log["employee_id"]})
            if employee:
                log["employee_name"] = employee.get("name", "Unknown")
                log["department"] = employee.get("department", "N/A")
        
        return [prepare_from_mongo(log) for log in ot_logs]
    except Exception as e:
        logging.error(f"Error fetching all OT logs: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch OT logs"
        )

@api_router.put("/ot/{ot_id}/approve")
async def approve_reject_ot(
    ot_id: str,
    approval_data: OTApprovalRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Admin approve or reject OT log"""
    try:
        ot_log = await db.ot_logs.find_one({"id": ot_id})
        if not ot_log:
            raise HTTPException(status_code=404, detail="OT log not found")
        
        update_data = {
            "status": approval_data.status,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        if approval_data.status == "approved":
            update_data["approved_by"] = current_user.username
            update_data["approved_date"] = datetime.now(timezone.utc).isoformat()
        else:
            update_data["rejected_by"] = current_user.username
            update_data["rejected_date"] = datetime.now(timezone.utc).isoformat()
            update_data["rejection_reason"] = approval_data.rejection_reason
        
        await db.ot_logs.update_one({"id": ot_id}, {"$set": update_data})
        
        # Create notification for employee
        employee_id = ot_log["employee_id"]
        status_text = "approved" if approval_data.status == "approved" else "rejected"
        notification = Notification(
            recipient_role="employee",
            recipient_id=employee_id,
            title=f"OT Log {status_text.capitalize()}",
            message=f"Your overtime log for {ot_log['date']} ({ot_log['ot_hours']} hours, Project: {ot_log.get('project', 'N/A')}) has been {status_text}.",
            category="ot",
            related_id=ot_id,
            notification_type="info" if approval_data.status == "approved" else "warning"
        )
        notification_dict = prepare_for_mongo(notification.dict())
        await db.notifications.insert_one(notification_dict)
        
        # Send real-time notification
        await send_realtime_notification(notification_dict)
        
        # Mark the OT submission notification as read using related_id
        await db.notifications.update_many(
            {
                "related_id": ot_id,
                "category": "ot",
                "recipient_role": "admin",
                "is_read": False
            },
            {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": f"OT log {status_text} successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error approving/rejecting OT: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to process OT approval"
        )

# Attendance Management Endpoints
@api_router.get("/attendance/my-attendance")
async def get_my_attendance(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Get employee's attendance records"""
    try:
        # Default to current month if not specified
        if not month or not year:
            now = datetime.now()
            month = now.month
            year = now.year
        
        # Get date range for the month
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Fetch attendance records
        attendance = await db.attendance.find({
            "employee_id": current_user.username,
            "date": {
                "$gte": month_start.isoformat(),
                "$lte": month_end.isoformat()
            }
        }).to_list(length=None)
        
        return [prepare_from_mongo(record) for record in attendance]
    except Exception as e:
        logging.error(f"Error fetching attendance: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch attendance"
        )

@api_router.get("/attendance/all")
async def get_all_attendance(
    month: Optional[int] = None,
    year: Optional[int] = None,
    employee_id: Optional[str] = None,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Admin view all attendance records"""
    try:
        # Default to current month if not specified
        if not month or not year:
            now = datetime.now()
            month = now.month
            year = now.year
        
        # Get date range for the month
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Build query
        query = {
            "date": {
                "$gte": month_start.isoformat(),
                "$lte": month_end.isoformat()
            }
        }
        
        if employee_id:
            query["employee_id"] = employee_id
        
        # Fetch attendance records
        attendance = await db.attendance.find(query).to_list(length=None)
        
        # Enrich with employee details
        for record in attendance:
            employee = await db.employees.find_one({"employee_id": record["employee_id"]})
            if employee:
                record["employee_name"] = employee.get("name", "Unknown")
                record["department"] = employee.get("department", "N/A")
        
        return [prepare_from_mongo(record) for record in attendance]
    except Exception as e:
        logging.error(f"Error fetching all attendance: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch attendance"
        )

@api_router.put("/attendance/correct")
async def correct_attendance(
    employee_id: str,
    correction_data: AttendanceCorrection,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Admin manually correct attendance"""
    try:
        # Check if attendance record exists
        existing = await db.attendance.find_one({
            "employee_id": employee_id,
            "date": correction_data.date.isoformat()
        })
        
        attendance_data = {
            "employee_id": employee_id,
            "date": correction_data.date.isoformat(),
            "status": correction_data.status,
            "working_hours": correction_data.working_hours,
            "notes": correction_data.notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        if existing:
            # Update existing record
            await db.attendance.update_one(
                {
                    "employee_id": employee_id,
                    "date": correction_data.date.isoformat()
                },
                {"$set": attendance_data}
            )
        else:
            # Create new record
            attendance = Attendance(
                employee_id=employee_id,
                date=correction_data.date,
                status=correction_data.status,
                working_hours=correction_data.working_hours,
                notes=correction_data.notes
            )
            attendance_dict = prepare_for_mongo(attendance.dict())
            await db.attendance.insert_one(attendance_dict)
        
        return {"message": "Attendance corrected successfully"}
    except Exception as e:
        logging.error(f"Error correcting attendance: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to correct attendance"
        )

@api_router.post("/attendance/mark")
async def mark_attendance(
    attendance_data: AttendanceMarkRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Admin marks or updates attendance for an employee"""
    try:
        # Validate status
        if attendance_data.status not in ['present', 'absent', 'half-day']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Status must be 'present', 'absent', or 'half-day'"
            )
        
        # Validate working hours
        if attendance_data.working_hours < 0 or attendance_data.working_hours > 24:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Working hours must be between 0 and 24"
            )
        
        # Verify employee exists
        employee = await db.employees.find_one({"employee_id": attendance_data.employee_id})
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee {attendance_data.employee_id} not found"
            )
        
        # Check if attendance record exists for this date
        date_str = attendance_data.date.isoformat()
        existing_record = await db.attendance.find_one({
            "employee_id": attendance_data.employee_id,
            "date": date_str
        })
        
        if existing_record:
            # Update existing record
            update_data = {
                "status": attendance_data.status,
                "working_hours": attendance_data.working_hours,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.attendance.update_one(
                {"employee_id": attendance_data.employee_id, "date": date_str},
                {"$set": update_data}
            )
            
            message = f"Attendance updated for {employee.get('name', attendance_data.employee_id)} on {date_str}"
        else:
            # Create new record
            attendance_record = {
                "id": str(uuid.uuid4()),
                "employee_id": attendance_data.employee_id,
                "date": date_str,
                "status": attendance_data.status,
                "check_in": "08:30" if attendance_data.status == 'present' else None,
                "check_out": "18:00" if attendance_data.status == 'present' else None,
                "working_hours": attendance_data.working_hours,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.attendance.insert_one(attendance_record)
            message = f"Attendance marked for {employee.get('name', attendance_data.employee_id)} on {date_str}"
        
        return {
            "message": message,
            "employee_id": attendance_data.employee_id,
            "date": date_str,
            "status": attendance_data.status,
            "working_hours": attendance_data.working_hours
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error marking attendance: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to mark attendance: {str(e)}"
        )

@api_router.post("/attendance/generate")
async def generate_attendance_for_month(
    month: int,
    year: int,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Admin generates attendance for all active employees for a given month"""
    try:
        # Get all active employees
        employees = await db.employees.find({"status": "active"}).to_list(length=None)
        
        # Get working days config and holidays
        settings = await db.settings.find_one({"id": "system_settings"})
        working_days_config = settings.get("working_days_config", {
            "saturday_policy": "alternate",
            "off_saturdays": [1, 3],
            "sunday_off": True
        }) if settings else {
            "saturday_policy": "alternate",
            "off_saturdays": [1, 3],
            "sunday_off": True
        }
        
        holidays = await db.holidays.find({}).to_list(length=None)
        holiday_dates = [h["date"] for h in holidays]
        
        # Get approved leaves for the month
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        approved_leaves = await db.leave_requests.find({
            "status": "approved",
            "$or": [
                {"start_date": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}},
                {"end_date": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}},
                {
                    "$and": [
                        {"start_date": {"$lte": month_start.isoformat()}},
                        {"end_date": {"$gte": month_end.isoformat()}}
                    ]
                }
            ]
        }).to_list(length=None)
        
        # Create a map of employee leaves
        employee_leaves = {}
        for leave in approved_leaves:
            employee_id = leave["employee_id"]
            if employee_id not in employee_leaves:
                employee_leaves[employee_id] = []
            employee_leaves[employee_id].append(leave)
        
        # Helper function to check if date is working day
        def is_working_day(check_date):
            day_of_week = check_date.weekday()
            
            # Check if it's a holiday
            if check_date.isoformat() in holiday_dates:
                return False, "holiday"
            
            # Check Sunday
            if day_of_week == 6:  # Sunday
                if working_days_config.get("sunday_off", True):
                    return False, "weekend"
            
            # Check Saturday
            if day_of_week == 5:  # Saturday
                saturday_policy = working_days_config.get("saturday_policy", "alternate")
                if saturday_policy == "all_off":
                    return False, "weekend"
                elif saturday_policy == "alternate" or saturday_policy == "custom":
                    # Calculate which Saturday of the month
                    first_saturday = date(check_date.year, check_date.month, 1)
                    while first_saturday.weekday() != 5:
                        first_saturday = first_saturday + timedelta(days=1)
                    saturday_number = ((check_date - first_saturday).days // 7) + 1
                    off_saturdays = working_days_config.get("off_saturdays", [1, 3])
                    if saturday_number in off_saturdays:
                        return False, "weekend"
            
            return True, "present"
        
        # Helper function to check if employee is on leave
        def is_on_leave(employee_id, check_date):
            if employee_id not in employee_leaves:
                return False, False
            
            for leave in employee_leaves[employee_id]:
                start_date = leave["start_date"]
                end_date = leave["end_date"]
                if isinstance(start_date, str):
                    start_date = datetime.fromisoformat(start_date).date()
                if isinstance(end_date, str):
                    end_date = datetime.fromisoformat(end_date).date()
                
                if start_date <= check_date <= end_date:
                    is_half_day = leave.get("half_day", False) and start_date == end_date == check_date
                    return True, is_half_day
            
            return False, False
        
        # Generate attendance records
        generated_count = 0
        days_in_month = (month_end - month_start).days + 1
        
        for employee in employees:
            employee_id = employee["employee_id"]
            
            for day_offset in range(days_in_month):
                current_date = month_start + timedelta(days=day_offset)
                
                # Check if record already exists
                existing = await db.attendance.find_one({
                    "employee_id": employee_id,
                    "date": current_date.isoformat()
                })
                
                if existing:
                    continue  # Skip if already exists
                
                # Determine attendance status
                is_working, default_status = is_working_day(current_date)
                on_leave, half_day_leave = is_on_leave(employee_id, current_date)
                
                if not is_working:
                    status = default_status
                    working_hours = 0.0
                elif on_leave:
                    status = "half-day" if half_day_leave else "leave"
                    working_hours = 4.0 if half_day_leave else 0.0
                else:
                    status = "present"
                    working_hours = 8.0
                
                # Create attendance record
                attendance = Attendance(
                    employee_id=employee_id,
                    date=current_date,
                    status=status,
                    working_hours=working_hours
                )
                attendance_dict = prepare_for_mongo(attendance.dict())
                await db.attendance.insert_one(attendance_dict)
                generated_count += 1
        
        return {
            "message": f"Successfully generated {generated_count} attendance records",
            "month": month,
            "year": year,
            "employees_processed": len(employees)
        }
    except Exception as e:
        logging.error(f"Error generating attendance: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate attendance: {str(e)}"
        )



@api_router.post("/attendance/generate-year")
async def generate_attendance_for_year(
    year: int = 2025,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """
    Admin generates attendance for all active employees for entire year (Jan 1 to today).
    Marks all working days (excluding weekends and holidays) as present with 8 hours.
    """
    try:
        # Get all active employees
        employees = await db.employees.find({"status": "active"}).to_list(length=None)
        
        if not employees:
            return {
                "message": "No active employees found",
                "generated_count": 0,
                "year": year
            }
        
        # Get working days config and holidays
        settings = await db.settings.find_one({"id": "system_settings"})
        working_days_config = settings.get("working_days_config", {
            "saturday_policy": "alternate",
            "off_saturdays": [1, 3],
            "sunday_off": True
        }) if settings else {
            "saturday_policy": "alternate",
            "off_saturdays": [1, 3],
            "sunday_off": True
        }
        
        holidays = await db.holidays.find({}).to_list(length=None)
        holiday_dates = [h["date"] for h in holidays]
        
        # Helper function to check if date is working day
        def is_working_day(check_date):
            # Check if it's a holiday
            if check_date.isoformat() in holiday_dates:
                return False
            
            # Check Sunday
            if check_date.weekday() == 6 and working_days_config.get("sunday_off", True):
                return False
            
            # Check Saturday
            if check_date.weekday() == 5:
                saturday_policy = working_days_config.get("saturday_policy", "alternate")
                
                if saturday_policy == "all_off":
                    return False
                elif saturday_policy == "all_working":
                    return True
                elif saturday_policy == "alternate" or saturday_policy == "custom":
                    # Find which Saturday of the month it is
                    day = check_date.day
                    first_saturday = 7 - date(check_date.year, check_date.month, 1).weekday() + 5
                    if first_saturday > 7:
                        first_saturday -= 7
                    
                    saturday_number = ((day - first_saturday) // 7) + 1
                    
                    off_saturdays = working_days_config.get("off_saturdays", [1, 3])
                    return saturday_number not in off_saturdays
            
            return True
        
        # Date range: Jan 1 to today
        start_date = date(year, 1, 1)
        end_date = date.today()
        
        # If end_date is before start_date (future year), return error
        if end_date < start_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Cannot generate attendance for future year {year}"
            )
        
        # Delete existing attendance records for this period
        delete_result = await db.attendance.delete_many({
            "date": {
                "$gte": start_date.isoformat(),
                "$lte": end_date.isoformat()
            }
        })
        
        logger.info(f"Deleted {delete_result.deleted_count} existing attendance records")
        
        generated_count = 0
        current_date = start_date
        
        # Generate attendance for each day
        while current_date <= end_date:
            # Check if it's a working day
            if is_working_day(current_date):
                # Create attendance for all active employees
                for employee in employees:
                    employee_id = employee.get("employee_id")
                    
                    # Create attendance record
                    attendance = Attendance(
                        employee_id=employee_id,
                        date=current_date,
                        status="present",
                        working_hours=8.0
                    )
                    attendance_dict = prepare_for_mongo(attendance.dict())
                    await db.attendance.insert_one(attendance_dict)
                    generated_count += 1
            
            current_date += timedelta(days=1)
        
        return {
            "message": f"Successfully generated {generated_count} attendance records for year {year}",
            "year": year,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "employees_processed": len(employees),
            "deleted_existing": delete_result.deleted_count
        }
        
    except Exception as e:
        logging.error(f"Error generating year attendance: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate year attendance: {str(e)}"
        )



@api_router.post("/late-arrivals")
async def record_late_arrival(
    late_data: LateArrivalCreate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Admin records a late arrival"""
    try:
        # Parse times
        actual_parts = late_data.actual_check_in.split(':')
        actual_minutes = int(actual_parts[0]) * 60 + int(actual_parts[1])
        
        # Expected check-in: 8:30 AM = 510 minutes
        expected_minutes = 8 * 60 + 30  # 510 minutes
        
        # Calculate late minutes
        late_minutes = actual_minutes - expected_minutes
        
        if late_minutes <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Actual check-in time is not late. Employee arrived on time or early."
            )
        
        late_arrival = LateArrival(
            employee_id=late_data.employee_id,
            date=late_data.date,
            actual_check_in=late_data.actual_check_in,
            late_minutes=late_minutes,
            reason=late_data.reason,
            recorded_by=current_user.username
        )
        
        late_dict = prepare_for_mongo(late_arrival.dict())
        await db.late_arrivals.insert_one(late_dict)
        
        # Get employee details for notification
        employee = await db.employees.find_one({"employee_id": late_data.employee_id})
        employee_name = employee.get("name", late_data.employee_id) if employee else late_data.employee_id
        
        # Create notification for employee
        notification = Notification(
            recipient_role="employee",
            recipient_id=late_data.employee_id,
            title="Late Arrival Recorded",
            message=f"Your late arrival on {late_data.date} at {late_data.actual_check_in} has been recorded ({late_minutes} minutes late).",
            notification_type="warning"
        )
        notification_dict = prepare_for_mongo(notification.dict())
        await db.notifications.insert_one(notification_dict)
        
        # Send real-time notification
        await send_realtime_notification(notification_dict)
        
        return {
            "message": f"Late arrival recorded for {employee_name}",
            "late_minutes": late_minutes,
            "late_arrival": prepare_from_mongo(late_dict)
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error recording late arrival: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to record late arrival: {str(e)}"
        )

@api_router.get("/late-arrivals")
async def get_late_arrivals(
    month: Optional[int] = None,
    year: Optional[int] = None,
    employee_id: Optional[str] = None,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get late arrival records"""
    try:
        # Default to current month if not specified
        if not month or not year:
            now = datetime.now()
            month = now.month
            year = now.year
        
        # Get date range for the month
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Build query
        query = {
            "date": {
                "$gte": month_start.isoformat(),
                "$lte": month_end.isoformat()
            }
        }
        
        if employee_id:
            query["employee_id"] = employee_id
        
        # Fetch late arrivals
        late_arrivals = await db.late_arrivals.find(query).to_list(length=None)
        
        # Enrich with employee details
        for record in late_arrivals:
            employee = await db.employees.find_one({"employee_id": record["employee_id"]})
            if employee:
                record["employee_name"] = employee.get("name", "Unknown")
                record["department"] = employee.get("department", "N/A")
        
        return [prepare_from_mongo(record) for record in late_arrivals]
    except Exception as e:
        logging.error(f"Error fetching late arrivals: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch late arrivals"
        )

@api_router.delete("/late-arrivals/{late_id}")
async def delete_late_arrival(
    late_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete a late arrival record"""
    try:
        result = await db.late_arrivals.delete_one({"id": late_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Late arrival record not found")
        
        return {"message": "Late arrival record deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting late arrival: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete late arrival record"
        )

@api_router.get("/late-arrivals/my-late-arrivals")
async def get_my_late_arrivals(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Employee view their own late arrival records"""
    try:
        # Default to current month if not specified
        if not month or not year:
            now = datetime.now()
            month = now.month
            year = now.year
        
        # Get date range for the month
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Build query for current employee
        query = {
            "employee_id": current_user.username,
            "date": {
                "$gte": month_start.isoformat(),
                "$lte": month_end.isoformat()
            }
        }
        
        # Fetch late arrivals
        late_arrivals = await db.late_arrivals.find(query).to_list(length=None)
        
        return [prepare_from_mongo(record) for record in late_arrivals]
    except Exception as e:
        logging.error(f"Error fetching my late arrivals: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch late arrivals"
        )
    except Exception as e:
        logging.error(f"Error deleting late arrival: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete late arrival record"
        )

@api_router.get("/attendance/monthly-summary")
async def get_monthly_attendance_summary(
    month: Optional[int] = None,
    year: Optional[int] = None,
    employee_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get monthly attendance summary including regular hours and OT"""
    try:
        # Default to current month if not specified
        if not month or not year:
            now = datetime.now()
            month = now.month
            year = now.year
        
        # Determine employee_id
        target_employee_id = employee_id if employee_id else current_user.username
        
        # Get date range for the month
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Fetch attendance records
        attendance = await db.attendance.find({
            "employee_id": target_employee_id,
            "date": {
                "$gte": month_start.isoformat(),
                "$lte": month_end.isoformat()
            }
        }).to_list(length=None)
        
        # Calculate regular working hours
        total_working_hours = 0
        present_days = 0
        leave_days = 0
        half_days = 0
        
        for record in attendance:
            status = record.get("status", "")
            if status == "present":
                present_days += 1
                total_working_hours += record.get("working_hours", 0)
            elif status == "leave":
                leave_days += 1
            elif status == "half-day":
                half_days += 1
                total_working_hours += record.get("working_hours", 0)
        
        # Fetch approved OT hours for the month
        month_str = f"{year}-{str(month).zfill(2)}"
        ot_logs = await db.ot_logs.find({
            "employee_id": target_employee_id,
            "status": "approved",
            "date": {"$regex": f"^{month_str}"}
        }).to_list(length=None)
        
        total_ot_hours = sum(log.get("ot_hours", 0) for log in ot_logs)
        
        # Calculate total hours (regular + OT)
        total_hours = total_working_hours + total_ot_hours
        
        return {
            "month": month,
            "year": year,
            "employee_id": target_employee_id,
            "present_days": present_days,
            "leave_days": leave_days,
            "half_days": half_days,
            "regular_working_hours": round(total_working_hours, 2),
            "ot_hours": round(total_ot_hours, 2),
            "total_hours": round(total_hours, 2)
        }
    except Exception as e:
        logging.error(f"Error fetching monthly summary: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch monthly summary"
        )


# Loan Management Endpoints
@api_router.post("/loans")
async def create_loan_request(
    loan_data: LoanRequestCreate,
    current_user: User = Depends(get_current_user)
):
    try:
        # Determine employee_id: use provided employee_id (admin creating for employee) or current user's ID
        target_employee_id = loan_data.employee_id if loan_data.employee_id else (current_user.employee_id or current_user.username)
        
        # Use provided interest rate or calculate based on loan type
        if loan_data.interest_rate is not None:
            interest_rate = loan_data.interest_rate
        else:
            interest_rates = {
                'Personal Loan': 12,
                'Emergency Loan': 10,
                'Advance Salary': 0,
                'Education Loan': 8
            }
            interest_rate = interest_rates.get(loan_data.loan_type, 12)
        
        # Calculate EMI if not provided
        if loan_data.monthly_emi is not None:
            monthly_emi = loan_data.monthly_emi
        else:
            P = loan_data.amount
            R = interest_rate / 100 / 12 if interest_rate > 0 else 0
            N = loan_data.tenure_months
            
            if R == 0:
                monthly_emi = P / N
            else:
                monthly_emi = P * R * (1 + R)**N / ((1 + R)**N - 1)
        
        loan_request = LoanRequest(
            employee_id=target_employee_id,
            loan_type=loan_data.loan_type,
            amount=loan_data.amount,
            tenure_months=loan_data.tenure_months,
            interest_rate=interest_rate,
            monthly_emi=round(monthly_emi),
            purpose=loan_data.purpose,
            monthly_income=loan_data.monthly_income,
            existing_loans=loan_data.existing_loans,
            remaining_emis=loan_data.tenure_months,
            guarantor_name=loan_data.guarantor_name,
            guarantor_employee_id=loan_data.guarantor_employee_id
        )
        
        loan_dict = prepare_for_mongo(loan_request.dict())
        result = await db.loan_requests.insert_one(loan_dict)
        
        if result.inserted_id:
            # Get employee details for notification
            employee = await db.employees.find_one({"employee_id": target_employee_id})
            employee_name = employee.get('name', 'Unknown') if employee else target_employee_id
            
            # Notify admins about new loan application
            await notify_loan_application(
                employee_id=target_employee_id,
                employee_name=employee_name,
                loan_amount=loan_data.amount,
                loan_id=loan_request.id
            )
            
            return loan_request
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create loan request"
            )
    except Exception as e:
        logging.error(f"Error creating loan request: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create loan request"
        )

@api_router.get("/loans")
async def get_loan_requests(
    current_user: User = Depends(get_current_user)
):
    try:
        if current_user.role == UserRole.ADMIN:
            # Admin can see all loan requests
            loans = await db.loan_requests.find({}).to_list(length=None)
        else:
            # Employee can only see their own requests
            loans = await db.loan_requests.find(
                {"employee_id": current_user.employee_id}
            ).to_list(length=None)
        
        # Convert MongoDB documents to JSON-serializable format
        serialized_loans = [prepare_from_mongo(loan) for loan in loans]
        
        return serialized_loans
    except Exception as e:
        logging.error(f"Error fetching loan requests: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch loan requests"
        )

@api_router.put("/loans/{loan_id}/approve")
async def approve_reject_loan(
    loan_id: str,
    approval_data: LoanApprovalRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    try:
        update_data = {
            "status": approval_data.status,
            "updated_at": datetime.now(timezone.utc)
        }
        
        if approval_data.status == "approved":
            update_data["approved_by"] = current_user.username
            update_data["approved_date"] = datetime.now(timezone.utc)
            if approval_data.disbursed_amount:
                update_data["disbursed_amount"] = approval_data.disbursed_amount
                update_data["outstanding_amount"] = approval_data.disbursed_amount
        elif approval_data.status == "rejected":
            update_data["rejected_by"] = current_user.username
            update_data["rejected_date"] = datetime.now(timezone.utc)
            update_data["rejection_reason"] = approval_data.rejection_reason
        
        # Get loan request data before updating for notification
        loan_request = await db.loan_requests.find_one({"id": loan_id})
        if not loan_request:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Loan request not found"
            )
        
        # Update loan request
        result = await db.loan_requests.update_one(
            {"id": loan_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Loan request not found"
            )
        
        # Send notification to employee using enhanced system
        await notify_loan_approval(
            employee_id=loan_request['employee_id'],
            loan_amount=loan_request['amount'],
            approved=(approval_data.status == "approved"),
            admin_comment=approval_data.rejection_reason or ""
        )
        
        # Mark the loan application notification as read using related_id
        await db.notifications.update_many(
            {
                "related_id": loan_id,
                "category": "loan",
                "recipient_role": "admin",
                "is_read": False
            },
            {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"message": f"Loan request {approval_data.status} successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating loan request: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update loan request"
        )

@api_router.delete("/loans/{loan_id}")
async def delete_loan_request(
    loan_id: str, 
    current_user: User = Depends(get_current_user)
):
    """Delete a loan request (Admin can delete any, Employee can delete their own pending requests)"""
    try:
        # Get the loan request to check permissions
        loan_request = await db.loan_requests.find_one({"id": loan_id})
        
        if not loan_request:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Loan request not found"
            )
        
        # Check permissions
        if current_user.role == UserRole.ADMIN:
            # Admin can delete any loan request
            pass
        elif current_user.role == UserRole.EMPLOYEE:
            # Employee can only delete their own pending loan requests
            if loan_request['employee_id'] != current_user.username:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You can only delete your own loan requests"
                )
            if loan_request.get('status', 'pending') != 'pending':
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="You can only delete pending loan requests"
                )
        else:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions"
            )
        
        # Delete the loan request
        result = await db.loan_requests.delete_one({"id": loan_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Loan request not found"
            )
        
        # Create notification about deletion
        if current_user.role == UserRole.ADMIN:
            await create_notification_helper(
                title="Loan Request Deleted",
                message=f"Loan request for ₹{loan_request.get('amount', 0):,.0f} has been deleted by admin.",
                recipient_role="employee",
                notification_type="info", 
                category="loan"
            )
        
        return {"message": "Loan request deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting loan request: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete loan request"
        )

# Recent Activities endpoint for admin dashboard
@api_router.get("/admin/recent-activities")
async def get_recent_activities(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Get recent activities for admin dashboard"""
    try:
        activities = []
        
        # Get recent notifications (last 10) - simplified version
        try:
            recent_notifications = await db.notifications.find(
                {"recipient_role": "admin"}
            ).sort("_id", -1).limit(10).to_list(length=None)
            
            for notif in recent_notifications:
                activities.append({
                    "id": str(notif.get("_id", "")),
                    "type": "notification", 
                    "title": notif.get("title", "Notification"),
                    "message": notif.get("message", ""),
                    "timestamp": notif.get("created_at", datetime.now(timezone.utc)),
                    "category": notif.get("category", "general")
                })
                
        except Exception as e:
            logging.error(f"Error fetching notifications: {str(e)}")
        
        # Add some static activities if we don't have notifications
        if len(activities) == 0:
            activities.append({
                "id": "welcome",
                "type": "info",
                "title": "Welcome to Elevate",
                "message": "Your comprehensive notification system is ready to track all activities",
                "timestamp": datetime.now(timezone.utc),
                "category": "system"
            })
        
        return {
            "activities": activities[:15],
            "total": len(activities)
        }
        
    except Exception as e:
        logging.error(f"Error fetching recent activities: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch recent activities: {str(e)}"
        )

# Manual birthday notification trigger endpoint (for testing)
@api_router.post("/notifications/trigger-birthday-check")
async def trigger_birthday_check(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Manually trigger birthday notifications check (for testing purposes)"""
    try:
        await check_daily_birthdays()
        return {"message": "Birthday notifications check completed successfully"}
    except Exception as e:
        logging.error(f"Error in manual birthday check: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to trigger birthday check: {str(e)}"
        )

# Notification endpoints
@api_router.post("/notifications")
async def create_notification(
    notification_data: NotificationCreate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Create a new notification"""
    try:
        notification = Notification(**notification_data.dict())
        notification_dict = prepare_for_mongo(notification.dict())
        
        # Send real-time notification
        await send_realtime_notification(notification_dict)
        
        result = await db.notifications.insert_one(notification_dict)
        if result.inserted_id:
            return {"message": "Notification created successfully", "id": notification.id}
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create notification"
            )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create notification: {str(e)}"
        )

@api_router.get("/notifications", response_model=List[Notification])
async def get_notifications(
    current_user: User = Depends(get_current_user)
):
    """Get notifications for current user"""
    try:
        # Build query based on user role
        if current_user.role == UserRole.ADMIN:
            query = {"$or": [
                {"recipient_role": "admin"},
                {"recipient_id": current_user.employee_id},
                {"recipient_role": None, "recipient_id": None}  # Global notifications
            ]}
        else:
            # Employee should only see:
            # 1. Notifications specifically for them (recipient_id matches)
            # 2. Broadcast notifications for all employees (recipient_role="employee" AND recipient_id is None)
            query = {"$or": [
                {"recipient_id": current_user.employee_id},
                {"recipient_role": "employee", "recipient_id": None}
            ]}
        
        notifications = await db.notifications.find(query).sort("created_at", -1).to_list(length=50)
        return [Notification(**notification) for notification in notifications]
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch notifications: {str(e)}"
        )

@api_router.delete("/notifications/test")
async def remove_test_notifications(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Remove test notifications from all employees"""
    try:
        # Remove notifications that contain "test" in title or message (case insensitive)
        result = await db.notifications.delete_many({
            "$or": [
                {"title": {"$regex": "test", "$options": "i"}},
                {"message": {"$regex": "test", "$options": "i"}},
                {"title": {"$regex": "welcome", "$options": "i"}},
                {"message": {"$regex": "welcome", "$options": "i"}},
                {"title": {"$regex": "demo", "$options": "i"}},
                {"message": {"$regex": "demo", "$options": "i"}}
            ]
        })
        
        return {
            "message": f"Removed {result.deleted_count} test notifications",
            "deleted_count": result.deleted_count
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to remove test notifications: {str(e)}"
        )

# PIN Email Models
class SendPinEmailsRequest(BaseModel):
    employee_ids: List[str]

async def send_pin_email(employee_email: str, employee_name: str, username: str, pin: str, app_url: str):
    """Send PIN email to employee with app link and credentials"""
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"Your Employee Portal Login Credentials - {employee_name}"
        msg['From'] = SMTP_FROM
        msg['To'] = employee_email
        msg['Bcc'] = SMTP_BCC
        
        # Create HTML email body
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ 
                    font-family: 'Aptos Display', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                    line-height: 1.6; 
                    color: #006B5B; 
                    margin: 0; 
                    padding: 20px; 
                    background-color: #f9f9f9;
                }}
                .container {{ 
                    max-width: 600px; 
                    margin: 0 auto; 
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                .header {{
                    background: linear-gradient(135deg, #006B5B, #008B74);
                    color: white;
                    padding: 30px;
                    text-align: center;
                }}
                .content {{
                    padding: 30px;
                }}
                .credentials-box {{
                    background: #f0f9f7;
                    border: 2px solid #006B5B;
                    border-radius: 8px;
                    padding: 20px;
                    margin: 20px 0;
                    text-align: center;
                }}
                .credential-item {{
                    margin: 10px 0;
                    font-size: 18px;
                }}
                .credential-label {{
                    font-weight: 600;
                    color: #004d43;
                }}
                .credential-value {{
                    font-family: 'Courier New', monospace;
                    background: white;
                    padding: 8px 12px;
                    border-radius: 4px;
                    border: 1px solid #ccc;
                    display: inline-block;
                    margin-left: 10px;
                    font-size: 16px;
                    color: #333;
                }}
                .app-link {{
                    display: inline-block;
                    background: #006B5B;
                    color: white;
                    padding: 12px 24px;
                    text-decoration: none;
                    border-radius: 6px;
                    font-weight: 600;
                    margin: 20px 0;
                    transition: background-color 0.3s;
                }}
                .app-link:hover {{
                    background: #005248;
                    color: white;
                    text-decoration: none;
                }}
                .footer {{
                    background: #f8f9fa;
                    padding: 20px;
                    text-align: center;
                    color: #666;
                    font-size: 14px;
                    border-top: 1px solid #e9ecef;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🔑 Employee Portal Access</h1>
                    <p>Your login credentials are ready</p>
                </div>
                
                <div class="content">
                    <p>Dear <strong>{employee_name}</strong>,</p>
                    
                    <p>Your employee portal account has been set up successfully. Below are your login credentials:</p>
                    
                    <div class="credentials-box">
                        <div class="credential-item">
                            <span class="credential-label">Username:</span>
                            <span class="credential-value">{username}</span>
                        </div>
                        <div class="credential-item">
                            <span class="credential-label">PIN:</span>
                            <span class="credential-value">{pin}</span>
                        </div>
                    </div>
                    
                    <p>Click the button below to access the employee portal:</p>
                    
                    <div style="text-align: center;">
                        <a href="{app_url}" class="app-link">
                            🚀 Access Employee Portal
                        </a>
                    </div>
                    
                    <div style="margin-top: 30px; padding: 15px; background: #fff3cd; border-radius: 6px; border-left: 4px solid #ffc107;">
                        <p style="margin: 0; font-size: 14px; color: #856404;">
                            <strong>Security Note:</strong> Please keep your credentials secure and do not share them with anyone. 
                            If you have any issues accessing your account, please contact the HR department.
                        </p>
                    </div>
                </div>
                
                <div class="footer">
                    <p><strong>Important:</strong> This is an automated email. Please do not reply to this message.</p>
                    <p>For support, contact your HR department or system administrator.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        html_part = MIMEText(html_body, 'html')
        msg.attach(html_part)
        
        # Send email
        await aiosmtplib.send(
            msg,
            hostname=SMTP_HOST,
            port=SMTP_PORT,
            username=SMTP_USER,
            password=SMTP_PASSWORD,
            start_tls=True
        )
        
        return True, None
        
    except Exception as e:
        logging.error(f"Error sending PIN email to {employee_email}: {str(e)}")
        return False, str(e)

@api_router.post("/admin/send-pin-emails")
async def send_pin_emails(
    request: SendPinEmailsRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Send PIN emails to selected employees"""
    try:
        if not request.employee_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No employee IDs provided"
            )
        
        # Get app URL from environment or use default
        app_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
        
        successful = 0
        failed = 0
        results = []
        
        for employee_id in request.employee_ids:
            # Get employee details
            employee = await db.employees.find_one({"employee_id": employee_id})
            if not employee:
                results.append({
                    "employee_id": employee_id,
                    "status": "failed",
                    "error": "Employee not found"
                })
                failed += 1
                continue
            
            if not employee.get('email'):
                results.append({
                    "employee_id": employee_id,
                    "name": employee.get('name', 'Unknown'),
                    "status": "failed",
                    "error": "No email address"
                })
                failed += 1
                continue
                
            # Get user PIN
            user = await db.users.find_one({"employee_id": employee_id})
            if not user or not user.get('pin'):
                results.append({
                    "employee_id": employee_id,
                    "name": employee.get('name', 'Unknown'),
                    "status": "failed",
                    "error": "No PIN found for employee"
                })
                failed += 1
                continue
            
            # Send email
            success, error = await send_pin_email(
                employee.get('email'),
                employee.get('name', 'Employee'),
                user.get('username'),
                user.get('pin'),
                app_url
            )
            
            results.append({
                "employee_id": employee_id,
                "name": employee.get('name', 'Unknown'),
                "email": employee.get('email'),
                "status": "sent" if success else "failed",
                "error": error
            })
            
            if success:
                successful += 1
            else:
                failed += 1
            
            # Small delay to avoid spam limits
            await asyncio.sleep(0.1)
        
        return {
            "message": f"Email sending completed: {successful} sent, {failed} failed",
            "total": len(request.employee_ids),
            "successful": successful,
            "failed": failed,
            "results": results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error sending PIN emails: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to send PIN emails: {str(e)}"
        )

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str,
    request: Optional[NotificationUpdateRequest] = None,
    current_user: User = Depends(get_current_user)
):
    """Mark notification as read or unread"""
    try:
        # If request body provided, use its is_read value, otherwise default to True (mark as read)
        is_read = request.is_read if request else True
        
        update_data = {
            "$set": {
                "is_read": is_read,
            }
        }
        
        # Set read_at timestamp only when marking as read
        if is_read:
            update_data["$set"]["read_at"] = datetime.now(timezone.utc)
        else:
            # When marking as unread, remove the read_at timestamp
            update_data["$unset"] = {"read_at": ""}
        
        result = await db.notifications.update_one(
            {"id": notification_id},
            update_data
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Notification not found"
            )
        
        return {"message": f"Notification marked as {'read' if is_read else 'unread'}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update notification: {str(e)}"
        )

@api_router.put("/notifications/mark-all-read")
async def mark_all_notifications_read(
    current_user: User = Depends(get_current_user)
):
    """Mark all notifications as read for current user"""
    try:
        # Build query based on user role
        if current_user.role == UserRole.ADMIN:
            query = {"$or": [
                {"recipient_role": "admin"},
                {"recipient_id": current_user.employee_id},
                {"recipient_role": None, "recipient_id": None}
            ]}
        else:
            # Employee should only see:
            # 1. Notifications specifically for them (recipient_id matches)
            # 2. Broadcast notifications for all employees (recipient_role="employee" AND recipient_id is None)
            query = {"$or": [
                {"recipient_id": current_user.employee_id},
                {"recipient_role": "employee", "recipient_id": None}
            ]}
        
        result = await db.notifications.update_many(
            query,
            {
                "$set": {
                    "is_read": True,
                    "read_at": datetime.now(timezone.utc)
                }
            }
        )
        
        return {"message": f"Marked {result.modified_count} notifications as read"}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to mark notifications as read: {str(e)}"
        )

@api_router.delete("/notifications/clear-all")
async def clear_all_notifications(
    current_user: User = Depends(get_current_user)
):
    """Clear all notifications for current user"""
    try:
        # Determine user filter based on role
        user_filter = {}
        if current_user.role == UserRole.ADMIN:
            user_filter = {"user_type": "admin"}
        elif current_user.role == UserRole.EMPLOYEE:
            user_filter = {"user_type": "employee"}
        else:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Invalid user role"
            )
        
        # Delete all notifications for the user
        result = await db.notifications.delete_many(user_filter)
        
        return {"message": f"Cleared {result.deleted_count} notifications"}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error clearing notifications: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to clear notifications: {str(e)}"
        )


@api_router.delete("/notifications/clear-read")
async def clear_read_notifications(
    current_user: User = Depends(get_current_user)
):
    """Clear only read notifications for current user"""
    try:
        # Build query based on user role to get only read notifications
        if current_user.role == UserRole.ADMIN:
            query = {
                "is_read": True,
                "$or": [
                    {"recipient_role": "admin"},
                    {"recipient_id": current_user.employee_id},
                    {"recipient_role": None, "recipient_id": None}
                ]
            }
        else:
            query = {
                "is_read": True,
                "$or": [
                    {"recipient_id": current_user.employee_id},
                    {"recipient_role": "employee", "recipient_id": None}
                ]
            }
        
        # Delete only read notifications for the user
        result = await db.notifications.delete_many(query)
        
        return {"message": f"Cleared {result.deleted_count} read notifications"}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error clearing read notifications: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to clear read notifications: {str(e)}"
        )

# Payslip endpoints
@api_router.post("/payslips/generate")
async def generate_payslips(
    request: PayslipGenerate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Generate payslips for active employees"""
    try:
        # Get active employees
        if request.employee_ids:
            employees = await db.employees.find(
                {"employee_id": {"$in": request.employee_ids}, "status": "active"}
            ).to_list(length=None)
        else:
            employees = await db.employees.find({"status": "active"}).to_list(length=None)
        
        generated_count = 0
        updated_count = 0
        
        for employee in employees:
            payslip_id = f"{employee['employee_id']}-{request.year}-{request.month:02d}"
            
            # Check if payslip already exists
            existing_payslip = await db.payslips.find_one({"id": payslip_id})
            
            if employee.get('salary_structure'):
                salary = employee['salary_structure']
                
                # Calculate earnings
                earnings = {
                    "basic_salary": salary.get('basic_salary', 0),
                    "house_rent_allowance": salary.get('house_rent_allowance', 0) or salary.get('hra', 0),
                    "medical_allowance": salary.get('medical_allowance', 0),
                    "leave_travel_allowance": salary.get('leave_travel_allowance', 0) or salary.get('travel_allowance', 0),
                    "conveyance_allowance": salary.get('conveyance_allowance', 0) or salary.get('food_allowance', 0),
                    "performance_incentive": salary.get('performance_incentive', 0) or salary.get('internet_allowance', 0),
                    "other_benefits": salary.get('other_benefits', 0) or salary.get('special_allowance', 0)
                }
                
                # Calculate deductions
                deductions = {
                    "pf_employee": salary.get('pf_employee', 0),
                    "esi_employee": salary.get('esi_employee', 0),
                    "professional_tax": salary.get('professional_tax', 0),
                    "tds": salary.get('tds', 0),
                    "loan_deductions": salary.get('loan_deductions', 0),
                    "others": salary.get('others', 0)
                }
                
                gross_salary = sum(earnings.values())
                total_deductions = sum(deductions.values())
                net_salary = gross_salary - total_deductions
                
                payslip_data = {
                    "id": payslip_id,
                    "employee_id": employee['employee_id'],
                    "month": request.month,
                    "year": request.year,
                    "gross_salary": gross_salary,
                    "total_deductions": total_deductions,
                    "net_salary": net_salary,
                    "earnings": earnings,
                    "deductions": deductions,
                    "status": "generated"
                }
                
                if existing_payslip:
                    # Update existing payslip
                    payslip_data["updated_date"] = datetime.now(timezone.utc)
                    prepared_data = prepare_for_mongo(payslip_data)
                    await db.payslips.update_one(
                        {"id": payslip_id},
                        {"$set": prepared_data}
                    )
                    updated_count += 1
                else:
                    # Create new payslip
                    payslip_data["generated_date"] = datetime.now(timezone.utc)
                    prepared_data = prepare_for_mongo(payslip_data)
                    await db.payslips.insert_one(prepared_data)
                    generated_count += 1
                
                # Notify employee about payslip generation (for both new and updated)
                month_names = ["", "January", "February", "March", "April", "May", "June",
                              "July", "August", "September", "October", "November", "December"]
                month_name = month_names[request.month]
                
                await notify_payslip_generated(
                    employee_id=employee['employee_id'],
                    month=month_name,
                    year=request.year
                )
        
        # Notify admins about bulk payslip generation
        if generated_count > 0 or updated_count > 0:
            month_names = ["", "January", "February", "March", "April", "May", "June",
                          "July", "August", "September", "October", "November", "December"]
            month_name = month_names[request.month]
            
            await notify_payslips_bulk_generated(
                employee_count=generated_count + updated_count,
                month=month_name,
                year=request.year
            )
        
        return {
            "message": f"Payslips processed: {generated_count} generated, {updated_count} updated",
            "generated_count": generated_count,
            "updated_count": updated_count,
            "total_employees": len(employees)
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate payslips: {str(e)}"
        )

@api_router.get("/payslips")
async def get_payslips(
    month: Optional[int] = None,
    year: Optional[int] = None,
    employee_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get payslips for the specified period or employee"""
    try:
        query = {"status": "generated"}
        if month:
            query["month"] = month
        if year:
            query["year"] = year
        if employee_id:
            query["employee_id"] = employee_id
            
        payslips = await db.payslips.find(query).to_list(length=None)
        
        # Get employee details for each payslip and clean MongoDB data
        for payslip in payslips:
            employee = await db.employees.find_one({"employee_id": payslip["employee_id"]})
            if employee:
                payslip["employee"] = prepare_from_mongo(employee)
        
        # Clean payslips data for JSON serialization
        cleaned_payslips = [prepare_from_mongo(payslip) for payslip in payslips]
        
        return cleaned_payslips
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch payslips: {str(e)}"
        )

@api_router.delete("/payslips/{payslip_id}")
async def delete_payslip(
    payslip_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete a specific payslip"""
    try:
        result = await db.payslips.delete_one({"id": payslip_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Payslip not found"
            )
        
        return {"message": "Payslip deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete payslip: {str(e)}"
        )

@api_router.delete("/payslips")
async def clear_all_payslips(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Clear all payslips"""
    try:
        result = await db.payslips.delete_many({})
        
        return {
            "message": f"Deleted {result.deleted_count} payslips",
            "deleted_count": result.deleted_count
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to clear payslips: {str(e)}"
        )

@api_router.delete("/payslips/month/{month}/{year}")
async def delete_payslips_by_month(
    month: int,
    year: int,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete all payslips for a specific month and year"""
    try:
        result = await db.payslips.delete_many({
            "month": month,
            "year": year
        })
        
        return {
            "message": f"Deleted {result.deleted_count} payslips for {month}/{year}",
            "deleted_count": result.deleted_count,
            "month": month,
            "year": year
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete payslips: {str(e)}"
        )

@api_router.put("/payslips/{payslip_id}/regenerate")
async def regenerate_payslip(
    payslip_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Regenerate a specific payslip"""
    try:
        # Get existing payslip
        payslip = await db.payslips.find_one({"id": payslip_id})
        if not payslip:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Payslip not found"
            )
        
        # Get employee data
        employee = await db.employees.find_one({"employee_id": payslip["employee_id"]})
        if not employee or employee.get("status") != "active":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Employee not found or not active"
            )
        
        # Regenerate payslip with current employee data
        request_data = PayslipGenerate(
            month=payslip["month"],
            year=payslip["year"],
            employee_ids=[payslip["employee_id"]]
        )
        
        # This will update the existing payslip
        result = await generate_payslips(request_data, current_user)
        
        return {"message": "Payslip regenerated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to regenerate payslip: {str(e)}"
        )

# Payroll Processing Endpoints
@api_router.get("/payroll/check-exists")
async def check_payroll_exists(
    month: int,
    year: int,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Check if payroll run already exists for the given month/year"""
    try:
        existing_run = await db.payroll_runs.find_one({
            "month": month,
            "year": year
        })
        
        if existing_run:
            return {
                "exists": True,
                "payroll_run_id": existing_run["id"],
                "total_employees": existing_run.get("total_employees", 0),
                "processed_date": existing_run.get("processed_date")
            }
        else:
            return {"exists": False}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to check payroll existence: {str(e)}"
        )

@api_router.delete("/payroll/run/{payroll_run_id}")
async def delete_payroll_run(
    payroll_run_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete a payroll run and its associated payslips"""
    try:
        # Get payroll run details first
        payroll_run = await db.payroll_runs.find_one({"id": payroll_run_id})
        if not payroll_run:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Payroll run not found"
            )
        
        # Delete associated payslips
        month = payroll_run["month"]
        year = payroll_run["year"]
        await db.payslips.delete_many({
            "month": month,
            "year": year
        })
        
        # Delete payroll run
        await db.payroll_runs.delete_one({"id": payroll_run_id})
        
        return {
            "message": "Payroll run and associated payslips deleted successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete payroll run: {str(e)}"
        )

@api_router.post("/payroll/run")
@api_router.post("/payroll-runs")
async def run_payroll(
    payroll_request: PayrollRunRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Process and save payroll run"""
    try:
        # Check for existing payroll run and delete if exists
        existing_run = await db.payroll_runs.find_one({
            "month": payroll_request.month,
            "year": payroll_request.year
        })
        
        if existing_run:
            # Delete old payroll run and its payslips
            await db.payslips.delete_many({
                "month": payroll_request.month,
                "year": payroll_request.year
            })
            await db.payroll_runs.delete_one({"id": existing_run["id"]})
        
        # Calculate totals and validate
        total_gross = 0
        total_deductions = 0
        total_net = 0
        
        processed_employees = []
        
        for emp_data in payroll_request.employees:
            # Get employee details
            employee = await db.employees.find_one({"employee_id": emp_data.employee_id})
            if not employee or employee.get('status') != 'active':
                continue
            
            salary = employee.get('salary_structure', {})
            
            # Calculate earnings (same logic as payslip generation)
            earnings = {
                "basic_salary": salary.get('basic_salary', 0),
                "house_rent_allowance": salary.get('house_rent_allowance', 0) or salary.get('hra', 0),
                "medical_allowance": salary.get('medical_allowance', 0),
                "leave_travel_allowance": salary.get('leave_travel_allowance', 0) or salary.get('travel_allowance', 0),
                "conveyance_allowance": salary.get('conveyance_allowance', 0) or salary.get('food_allowance', 0),
                "performance_incentive": salary.get('performance_incentive', 0) or salary.get('internet_allowance', 0),
                "other_benefits": salary.get('other_benefits', 0) or salary.get('special_allowance', 0)
            }
            
            # Calculate deductions
            deductions = {
                "pf_employee": salary.get('pf_employee', 0),
                "esi_employee": salary.get('esi_employee', 0),
                "professional_tax": salary.get('professional_tax', 0),
                "tds": emp_data.tds,  # Use dynamic TDS from payroll form
                "loan_deductions": emp_data.loan_deductions,  # Use dynamic loan deductions from payroll form
                "others": salary.get('others', 0)
            }
            
            gross = sum(earnings.values())
            deductions_total = sum(deductions.values())
            
            # Apply adjustments
            net = gross - deductions_total + emp_data.bonus + emp_data.adjustments
            
            total_gross += gross
            total_deductions += deductions_total
            total_net += net
            
            processed_employees.append({
                "employee_id": emp_data.employee_id,
                "days_worked": emp_data.days_worked,
                "days_in_month": emp_data.days_in_month,
                "overtime_hours": emp_data.overtime_hours,
                "bonus": emp_data.bonus,
                "adjustments": emp_data.adjustments,
                "loan_deductions": emp_data.loan_deductions,
                "tds": emp_data.tds,
                "gross_salary": gross,
                "total_deductions": deductions_total,
                "net_salary": net,
                "earnings": earnings,
                "deductions": deductions
            })
        
        # Create payroll run record
        payroll_run = {
            "id": str(uuid.uuid4()),
            "month": payroll_request.month,
            "year": payroll_request.year,
            "employees": processed_employees,
            "total_employees": len(processed_employees),
            "total_gross": total_gross,
            "total_deductions": total_deductions,
            "total_net": total_net,
            "status": "completed",
            "processed_date": datetime.now(timezone.utc),
            "processed_by": current_user.username
        }
        
        prepared_data = prepare_for_mongo(payroll_run)
        await db.payroll_runs.insert_one(prepared_data)
        
        return {
            "message": "Payroll run completed successfully",
            "payroll_run_id": payroll_run["id"],
            "total_employees": len(processed_employees),
            "total_net": total_net
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process payroll: {str(e)}"
        )

@api_router.get("/payroll-runs")
async def get_payroll_runs(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get payroll runs, optionally filtered by month and year"""
    try:
        query = {}
        if month is not None:
            query["month"] = month
        if year is not None:
            query["year"] = year
        
        payroll_runs = await db.payroll_runs.find(query, {"_id": 0}).sort("processed_date", -1).to_list(length=None)
        
        # Enrich with employee details
        for run in payroll_runs:
            for emp_data in run.get("employees", []):
                employee = await db.employees.find_one(
                    {"employee_id": emp_data["employee_id"]},
                    {"_id": 0}
                )
                if employee:
                    emp_data["employee_name"] = employee.get("name", "")
                    bank_info = employee.get("bank_info", {})
                    emp_data["account_number"] = bank_info.get("account_number", "")
                    emp_data["ifsc_code"] = bank_info.get("ifsc_code", "")
                    emp_data["bank_name"] = bank_info.get("bank_name", "")
                    emp_data["branch"] = bank_info.get("branch", "")
        
        return payroll_runs
    except Exception as e:
        logging.error(f"Error fetching payroll runs: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch payroll runs: {str(e)}"
        )

@api_router.get("/payroll/history")
async def get_payroll_history(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get payroll run history"""
    try:
        # Get all payroll runs, sorted by date (newest first)
        payroll_runs = await db.payroll_runs.find({}, {"_id": 0}).sort("processed_date", -1).to_list(length=None)
        
        return payroll_runs
    except Exception as e:
        logging.error(f"Error fetching payroll history: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch payroll history: {str(e)}"
        )

@api_router.get("/payroll/history/{payroll_run_id}")
async def get_payroll_run_details(
    payroll_run_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get detailed information for a specific payroll run"""
    try:
        payroll_run = await db.payroll_runs.find_one({"id": payroll_run_id}, {"_id": 0})
        if not payroll_run:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Payroll run not found"
            )
        
        return payroll_run
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching payroll run details: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch payroll run details: {str(e)}"
        )

@api_router.post("/payroll/{payroll_run_id}/generate-payslips")
async def generate_payslips_from_payroll(
    payroll_run_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Generate payslips from a completed payroll run with pro-ration"""
    try:
        # Get payroll run
        payroll_run = await db.payroll_runs.find_one({"id": payroll_run_id})
        if not payroll_run:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Payroll run not found"
            )
        
        # Calculate actual days in the month
        from calendar import monthrange
        actual_days_in_month = monthrange(payroll_run["year"], payroll_run["month"])[1]
        
        generated_count = 0
        updated_count = 0
        
        # Generate payslips for each employee in the payroll run
        for emp_data in payroll_run["employees"]:
            payslip_id = f"{emp_data['employee_id']}-{payroll_run['year']}-{payroll_run['month']:02d}"
            
            # Check if payslip already exists
            existing_payslip = await db.payslips.find_one({"id": payslip_id})
            
            # Get employee details for payslip (for name, designation, etc.)
            employee = await db.employees.find_one({"employee_id": emp_data["employee_id"]})
            
            # Get days worked from payroll data
            days_worked = emp_data.get("days_worked", actual_days_in_month)
            
            # Get original earnings, bonus and adjustments
            original_earnings = emp_data.get("earnings", {})
            bonus = emp_data.get("bonus", 0)
            adjustments = emp_data.get("adjustments", 0)
            
            # Indian Payroll Logic: Only prorate if worked less than full month
            # If worked full month (days_worked >= actual_days_in_month), use full salary
            if days_worked >= actual_days_in_month:
                # Full month worked - NO pro-ration
                prorated_earnings = {
                    "basic_salary": original_earnings.get("basic_salary", 0),
                    "house_rent_allowance": original_earnings.get("house_rent_allowance", 0),
                    "medical_allowance": original_earnings.get("medical_allowance", 0),
                    "leave_travel_allowance": original_earnings.get("leave_travel_allowance", 0),
                    "conveyance_allowance": original_earnings.get("conveyance_allowance", 0),
                    "performance_incentive": original_earnings.get("performance_incentive", 0),
                    "other_benefits": original_earnings.get("other_benefits", 0) + bonus  # Add bonus
                }
            else:
                # Partial month - Prorate based on actual days in month
                prorated_earnings = {
                    "basic_salary": round((original_earnings.get("basic_salary", 0) / actual_days_in_month) * days_worked, 2),
                    "house_rent_allowance": round((original_earnings.get("house_rent_allowance", 0) / actual_days_in_month) * days_worked, 2),
                    "medical_allowance": round((original_earnings.get("medical_allowance", 0) / actual_days_in_month) * days_worked, 2),
                    "leave_travel_allowance": round((original_earnings.get("leave_travel_allowance", 0) / actual_days_in_month) * days_worked, 2),
                    "conveyance_allowance": round((original_earnings.get("conveyance_allowance", 0) / actual_days_in_month) * days_worked, 2),
                    "performance_incentive": round((original_earnings.get("performance_incentive", 0) / actual_days_in_month) * days_worked, 2),
                    "other_benefits": round((original_earnings.get("other_benefits", 0) / actual_days_in_month) * days_worked + bonus, 2)
                }
            
            # Get deductions and add adjustments to others
            original_deductions = emp_data.get("deductions", {})
            deductions = {
                "pf_employee": original_deductions.get("pf_employee", 0),
                "esi_employee": original_deductions.get("esi_employee", 0),
                "professional_tax": original_deductions.get("professional_tax", 0),
                "tds": original_deductions.get("tds", 0),
                "loan_deductions": original_deductions.get("loan_deductions", 0),
                "others": original_deductions.get("others", 0) + abs(adjustments)  # Include adjustments here (always positive)
            }
            
            # Calculate totals
            gross_salary = sum(prorated_earnings.values())  # Already includes bonus in other_benefits
            total_deductions = sum(deductions.values())  # Already includes adjustments in others
            net_salary = gross_salary - total_deductions
            
            # Create comprehensive payslip data
            payslip_data = {
                "id": payslip_id,
                "employee_id": emp_data["employee_id"],
                "month": payroll_run["month"],
                "year": payroll_run["year"],
                "days_worked": days_worked,
                "days_in_month": actual_days_in_month,
                "earnings": prorated_earnings,  # Bonus already included in other_benefits
                "deductions": deductions,  # Adjustments already included in others
                "bonus": bonus,  # Store for reference
                "adjustments": adjustments,  # Store for reference
                "gross_salary": round(gross_salary, 2),
                "total_deductions": round(total_deductions, 2),
                "net_salary": round(net_salary, 2),
                "status": "generated",
                "employee_details": {
                    "name": employee.get("name") if employee else "Unknown",
                    "employee_id": emp_data["employee_id"],
                    "designation": employee.get("designation") if employee else "N/A",
                    "department": employee.get("department") if employee else "N/A",
                    "date_of_joining": employee.get("date_of_joining") if employee else None,
                    "pan": employee.get("bank_info", {}).get("pan") or employee.get("pan_number") if employee else "N/A",
                    "bank_account": employee.get("bank_info", {}).get("account_number") if employee else "N/A",
                    "ifsc": employee.get("bank_info", {}).get("ifsc_code") if employee else "N/A"
                }
            }
            
            if existing_payslip:
                # Update existing payslip
                payslip_data["updated_date"] = datetime.now(timezone.utc)
                prepared_data = prepare_for_mongo(payslip_data)
                await db.payslips.update_one(
                    {"id": payslip_id},
                    {"$set": prepared_data}
                )
                updated_count += 1
            else:
                # Create new payslip
                payslip_data["generated_date"] = datetime.now(timezone.utc)
                prepared_data = prepare_for_mongo(payslip_data)
                await db.payslips.insert_one(prepared_data)
                generated_count += 1
        
        return {
            "message": f"Payslips processed: {generated_count} generated, {updated_count} updated",
            "generated_count": generated_count,
            "updated_count": updated_count,
            "total_employees": len(payroll_run["employees"])
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate payslips from payroll: {str(e)}"
        )

# Migration endpoint for salary structure
@api_router.post("/migrate-salary-structure")
async def migrate_salary_structure(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Migrate existing salary structure to new field names"""
    try:
        employees = await db.employees.find({"salary_structure": {"$exists": True}}).to_list(length=None)
        migrated_count = 0
        
        for employee in employees:
            salary = employee.get('salary_structure', {})
            
            # Check if migration is needed
            if 'hra' in salary and 'house_rent_allowance' not in salary:
                # Migrate old field names to new ones
                update_data = {
                    "salary_structure.house_rent_allowance": salary.get('hra', 0),
                    "salary_structure.leave_travel_allowance": salary.get('travel_allowance', 0),
                    "salary_structure.conveyance_allowance": salary.get('food_allowance', 0),
                    "salary_structure.performance_incentive": salary.get('internet_allowance', 0),
                    "salary_structure.other_benefits": salary.get('special_allowance', 0),
                    "salary_structure.loan_deductions": 0,
                    "salary_structure.others": 0,
                    "updated_at": datetime.now(timezone.utc)
                }
                
                update_data = prepare_for_mongo(update_data)
                
                await db.employees.update_one(
                    {"employee_id": employee['employee_id']},
                    {"$set": update_data}
                )
                migrated_count += 1
        
        return {"message": f"Successfully migrated {migrated_count} employee salary structures"}
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Migration failed: {str(e)}"
        )

# Manual user initialization endpoint for testing
@api_router.post("/init-users")
async def init_users():
    try:
        # Initialize default admin user
        admin_exists = await db.users.find_one({"username": "admin", "role": "admin"})
        if not admin_exists:
            admin_user = User(
                username="admin",
                email="admin@company.com",
                role=UserRole.ADMIN,
                hashed_password=get_password_hash("Admin$2022"),
                is_active=True
            )
            await db.users.insert_one(prepare_for_mongo(admin_user.dict()))
            print("Default admin user created: admin/password")
        
        # Create employee users for existing employees
        employees = await db.employees.find({}).to_list(length=None)
        created_employees = []
        for emp in employees:
            employee_user_exists = await db.users.find_one({
                "username": emp["employee_id"], 
                "role": "employee"
            })
            if not employee_user_exists:
                # Generate random PIN for employee
                default_pin = generate_random_pin()
                employee_user = User(
                    username=emp["employee_id"],
                    email=emp.get("email"),
                    role=UserRole.EMPLOYEE,
                    employee_id=emp["employee_id"],
                    pin=default_pin,
                    is_active=True
                )
                await db.users.insert_one(prepare_for_mongo(employee_user.dict()))
                created_employees.append(emp["employee_id"])
                print(f"Employee user created: {emp['employee_id']}/PIN:1234")
        
        return {
            "message": "Users initialized successfully",
            "admin_created": not admin_exists,
            "employees_created": created_employees
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to initialize users: {str(e)}"
        )

@api_router.post("/update-employee-pins")
async def update_employee_pins(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Update all employee PINs to random 4-digit PINs"""
    try:
        updated_count = 0
        
        # Get all employee users
        employees = await db.users.find({"role": "employee"}).to_list(length=None)
        
        for emp in employees:
            employee_id = emp.get("employee_id", "")
            new_pin = generate_random_pin()
                
            # Update PIN
            result = await db.users.update_one(
                {"employee_id": employee_id},
                {"$set": {"pin": new_pin}}
            )
            
            if result.modified_count > 0:
                updated_count += 1
                
        return {
            "message": f"Updated PINs for {updated_count} employees",
            "updated_count": updated_count
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update employee PINs: {str(e)}"
        )

# Admin PIN management endpoints


@api_router.get("/payroll/download-template")
async def download_payroll_template(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Download Excel template for payroll import"""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Data"
        
        # Define headers
        headers = [
            "Payroll Date (Last Day of Month)",
            "Employee ID",
            "Full Name",
            "Basic",
            "House Rent Allowance",
            "Medical Allowance",
            "Leave Travel Allowance",
            "Bonus",
            "Performance Incentive",
            "Other Benefits",
            "Provident Fund",
            "Profession Tax",
            "ESI",
            "Advances/Loans",
            "Income Tax",
            "Other Deductions"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Set column widths
            ws.column_dimensions[cell.column_letter].width = 20
        
        # Add sample data row with Excel date
        from datetime import datetime as dt
        # Last day of October 2025 as Excel date
        sample_date = dt(2025, 10, 31)
        
        sample_data = [
            sample_date,  # Month/Year (Excel date - last day of month)
            "ET-MUM-00001",  # Employee ID
            "John Doe",  # Full Name
            50000,  # Basic
            20000,  # HRA
            1250,  # Medical
            2000,  # LTA
            1600,  # Bonus
            5000,  # Performance Incentive
            2000,  # Other Benefits
            6000,  # PF
            200,  # PT
            0,  # ESI
            0,  # Advances/Loans
            5000,  # Income Tax
            0   # Other Deductions
        ]
        
        for col_num, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.border = border
            if col_num == 1:  # Date column
                cell.number_format = 'DD/MM/YYYY'
            elif col_num > 3:  # Numeric columns
                cell.number_format = '#,##0.00'
        
        # Add instructions sheet
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            ["Payroll Import Template - Instructions", ""],
            ["", ""],
            ["Column", "Description"],
            ["Month/Year", "Excel date format (use last date of payroll month, e.g., 31/10/2025 for October 2025). Can also use MM/YYYY text format."],
            ["Employee ID", "Must match existing employee ID in the system"],
            ["Full Name", "Employee's full name (for reference only)"],
            ["Basic", "Basic salary amount"],
            ["House Rent Allowance", "HRA amount"],
            ["Medical Allowance", "Medical allowance amount"],
            ["Leave Travel Allowance", "LTA amount"],
            ["Bonus", "Bonus amount"],
            ["Performance Incentive", "Performance bonus/incentive amount"],
            ["Other Benefits", "Any other earnings"],
            ["Provident Fund", "Employee PF contribution (deduction)"],
            ["Profession Tax", "Professional tax (deduction)"],
            ["ESI", "Employee ESI contribution (deduction)"],
            ["Advances/Loans", "Loan EMI deduction for the month"],
            ["Income Tax", "TDS/Income tax (deduction)"],
            ["Other Deductions", "Any other deductions"],
            ["", ""],
            ["Important Notes:", ""],
            ["1.", "Date Format: Use Excel date (last day of month) OR MM/YYYY text format"],
            ["2.", "Example: For October 2025 payroll, use 31/10/2025 as an Excel date"],
            ["3.", "Employee ID must exist in the system"],
            ["4.", "All amounts should be numeric values only"],
            ["5.", "Gross Salary = Sum of all earnings columns"],
            ["6.", "Net Salary = Gross Salary - Sum of all deductions"],
            ["7.", "Importing will create/update payslips for the specified month"],
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws2.cell(row=row_num, column=1).value = instruction[0]
            ws2.cell(row=row_num, column=2).value = instruction[1]
            if row_num == 1:
                ws2.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            elif row_num == 3:
                ws2.cell(row=row_num, column=1).font = Font(bold=True)
                ws2.cell(row=row_num, column=2).font = Font(bold=True)
        
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 60
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as downloadable file
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Payroll_Import_Template.xlsx"}
        )
        
    except Exception as e:
        logging.error(f"Error generating payroll template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate template: {str(e)}"
        )


@api_router.post("/payroll/import-excel")
async def import_payroll_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Import payroll data from Excel file"""
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid file type. Please upload an Excel file (.xlsx or .xls)"
            )
        
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Parse data
        imported_count = 0
        error_count = 0
        errors = []
        
        # Get ALL employees (including inactive) for validation
        employees = await db.employees.find({}).to_list(length=None)
        employee_dict = {emp["employee_id"]: emp for emp in employees}
        
        # Process rows (skip header)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0] or not row[1]:  # Skip empty rows
                    continue
                
                # Parse month/year from Excel date
                month_year_value = row[0]
                try:
                    # Check if it's an Excel date (datetime object or number)
                    if isinstance(month_year_value, datetime):
                        # Already a datetime object
                        month = month_year_value.month
                        year = month_year_value.year
                    elif isinstance(month_year_value, (int, float)):
                        # Excel serial date number
                        # Excel's epoch is 1899-12-30, but Python's datetime.fromordinal starts from 0001-01-01
                        # Excel serial 1 = 1900-01-01, but Excel incorrectly treats 1900 as a leap year
                        # So we adjust: Excel date = Python ordinal date - 693594 + 1
                        from datetime import datetime as dt
                        excel_epoch = dt(1899, 12, 30)
                        date_obj = excel_epoch + timedelta(days=month_year_value)
                        month = date_obj.month
                        year = date_obj.year
                    elif isinstance(month_year_value, str):
                        # String format MM/YYYY or similar
                        month_year_str = month_year_value.strip()
                        if '/' in month_year_str:
                            parts = month_year_str.split('/')
                            month = int(parts[0])
                            year = int(parts[1])
                        elif '-' in month_year_str:
                            parts = month_year_str.split('-')
                            month = int(parts[0])
                            year = int(parts[1])
                        else:
                            errors.append(f"Row {row_num}: Invalid date format. Use MM/YYYY or Excel date")
                            error_count += 1
                            continue
                    else:
                        errors.append(f"Row {row_num}: Unrecognized date format")
                        error_count += 1
                        continue
                    
                    # Validate month and year
                    if not (1 <= month <= 12):
                        errors.append(f"Row {row_num}: Invalid month {month}. Must be between 1 and 12")
                        error_count += 1
                        continue
                    if year < 2000 or year > 2100:
                        errors.append(f"Row {row_num}: Invalid year {year}")
                        error_count += 1
                        continue
                        
                except Exception as e:
                    errors.append(f"Row {row_num}: Error parsing date - {str(e)}")
                    error_count += 1
                    continue
                
                # Validate minimum columns exist (at least Employee ID)
                if len(row) < 2:
                    errors.append(f"Row {row_num}: Insufficient columns (need at least 2)")
                    error_count += 1
                    continue
                
                employee_id = str(row[1]).strip() if row[1] else ""
                
                if not employee_id:
                    errors.append(f"Row {row_num}: Employee ID is empty")
                    error_count += 1
                    continue
                
                # Validate employee exists
                if employee_id not in employee_dict:
                    errors.append(f"Row {row_num}: Employee ID '{employee_id}' not found in system")
                    error_count += 1
                    continue
                
                employee = employee_dict[employee_id]
                
                # Helper function to safely get column value (defaults to 0 for empty/missing)
                def get_col_value(row, index, default=0):
                    try:
                        if len(row) > index and row[index] is not None:
                            value = row[index]
                            # Handle empty strings or whitespace
                            if isinstance(value, str):
                                value = value.strip()
                                if not value:
                                    return default
                            return float(value)
                        return default
                    except (ValueError, TypeError):
                        return default
                
                # Parse earnings (columns 3-9)
                basic = get_col_value(row, 3)
                hra = get_col_value(row, 4)
                medical = get_col_value(row, 5)
                lta = get_col_value(row, 6)
                conveyance = get_col_value(row, 7)
                performance_incentive = get_col_value(row, 8)
                other_benefits = get_col_value(row, 9)
                
                # Parse deductions (columns 10-15)
                pf = get_col_value(row, 10)
                pt = get_col_value(row, 11)
                esi = get_col_value(row, 12)
                loans = get_col_value(row, 13)
                income_tax = get_col_value(row, 14)
                other_deductions = get_col_value(row, 15)
                
                # Calculate totals
                earnings = {
                    "basic_salary": basic,
                    "house_rent_allowance": hra,
                    "medical_allowance": medical,
                    "leave_travel_allowance": lta,
                    "conveyance_allowance": conveyance,
                    "performance_incentive": performance_incentive,
                    "other_benefits": other_benefits
                }
                
                deductions = {
                    "pf_employee": pf,
                    "professional_tax": pt,
                    "esi_employee": esi,
                    "loan_deductions": loans,
                    "tds": income_tax,
                    "others": other_deductions
                }
                
                gross_salary = sum(earnings.values())
                total_deductions = sum(deductions.values())
                net_salary = gross_salary - total_deductions
                
                # Create or update payslip
                payslip_data = {
                    "id": str(uuid.uuid4()),
                    "employee_id": employee_id,
                    "employee_name": employee.get("name"),
                    "month": month,
                    "year": year,
                    "earnings": earnings,
                    "deductions": deductions,
                    "gross_salary": gross_salary,
                    "total_deductions": total_deductions,
                    "net_salary": net_salary,
                    "generated_date": datetime.now(timezone.utc),
                    "status": "generated"
                }
                
                # Check if payslip already exists
                existing = await db.payslips.find_one({
                    "employee_id": employee_id,
                    "month": month,
                    "year": year
                })
                
                if existing:
                    # Update existing
                    await db.payslips.update_one(
                        {"employee_id": employee_id, "month": month, "year": year},
                        {"$set": payslip_data}
                    )
                else:
                    # Insert new
                    await db.payslips.insert_one(payslip_data)
                
                # Create notification for employee
                notification = {
                    "id": str(uuid.uuid4()),
                    "title": "Payslip Generated",
                    "message": f"Your payslip for {month}/{year} has been generated",
                    "notification_type": "success",
                    "category": "payslip",
                    "recipient_id": employee_id,
                    "recipient_role": "employee",
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                await db.notifications.insert_one(notification)
                
                # Send real-time notification
                await send_realtime_notification(notification)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
        
        return {
            "success": True,
            "imported_count": imported_count,
            "error_count": error_count,
            "errors": errors[:20]  # Return first 20 errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error importing payroll from Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import payroll: {str(e)}"
        )

@api_router.get("/admin/employee-pins")
async def get_employee_pins(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Get employee PINs for active employees only"""
    try:
        # Get all active employees first
        active_employees = await db.employees.find({"status": "active"}).to_list(length=None)
        active_employee_ids = [emp["employee_id"] for emp in active_employees]
        
        # Get users for active employees only
        users = await db.users.find({
            "role": "employee",
            "employee_id": {"$in": active_employee_ids}
        }).to_list(length=None)
        
        pin_data = []
        for user in users:
            employee_id = user.get("employee_id")
            # Get employee details
            employee_details = next((emp for emp in active_employees if emp["employee_id"] == employee_id), None)
            
            if employee_details:  # Only add if employee is truly active
                # Get the most recent login history entry for location data
                recent_login = await db.login_history.find_one(
                    {"employee_id": employee_id},
                    sort=[("login_time", -1)]
                )
                
                location_data = None
                if recent_login and recent_login.get("location"):
                    location_data = recent_login.get("location")
                
                pin_data.append({
                    "employee_id": employee_id,
                    "name": employee_details.get("name", "Unknown"),
                    "email": employee_details.get("email", ""),
                    "pin": user.get("pin"),
                    "last_login": user.get("last_login"),
                    "last_login_ip": user.get("last_login_ip"),
                    "last_login_device": user.get("last_login_device"),
                    "last_login_location": location_data,
                    "status": employee_details.get("status", "active")
                })
        
        return {"employee_pins": pin_data}
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch employee PINs: {str(e)}"
        )



@api_router.get("/employees/{employee_id}/login-history")
async def get_employee_login_history(
    employee_id: str,
    days: int = 365,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get login history for an employee for the specified number of days"""
    try:
        # Calculate the date threshold
        threshold_date = datetime.now(timezone.utc) - timedelta(days=days)
        
        # Fetch login history from the database
        login_history = await db.login_history.find({
            "employee_id": employee_id,
            "login_time": {"$gte": threshold_date}
        }).sort("login_time", -1).to_list(length=None)
        
        # Format the response
        formatted_history = []
        for entry in login_history:
            formatted_entry = {
                "id": entry.get("id"),
                "login_time": entry.get("login_time"),
                "ip_address": entry.get("ip_address"),
                "device_name": entry.get("device_name"),
                "pc_name": entry.get("pc_name", "Desktop"),
                "location": entry.get("location")
            }
            formatted_history.append(formatted_entry)
        
        return {
            "employee_id": employee_id,
            "days": days,
            "total_logins": len(formatted_history),
            "history": formatted_history
        }
        
    except Exception as e:
        logging.error(f"Failed to fetch login history for {employee_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch login history: {str(e)}"
        )

class PinUpdateRequest(BaseModel):
    employee_id: str
    new_pin: Optional[str] = None  # If None, generate random PIN

@api_router.put("/admin/employee-pins")
async def update_employee_pin(
    request: PinUpdateRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Update a specific employee's PIN"""
    try:
        # Generate random PIN if not provided
        new_pin = request.new_pin if request.new_pin else generate_random_pin()
        
        # Validate PIN format (4 digits)
        if not new_pin.isdigit() or len(new_pin) != 4:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="PIN must be exactly 4 digits"
            )
        
        # Update employee PIN
        result = await db.users.update_one(
            {"employee_id": request.employee_id, "role": "employee"},
            {"$set": {"pin": new_pin}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee not found"
            )
        
        return {
            "message": f"PIN updated successfully for employee {request.employee_id}",
            "new_pin": new_pin
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update employee PIN: {str(e)}"
        )

@api_router.post("/admin/cleanup-users")
async def cleanup_user_accounts(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Clean up invalid user accounts and ensure data consistency"""
    try:
        # Get all users and employees
        users = await db.users.find({"role": "employee"}).to_list(length=None)
        employees = await db.employees.find({}).to_list(length=None)
        
        # Create lookup for valid employee IDs with their status
        valid_employees = {}
        active_employees = set()
        
        for emp in employees:
            valid_employees[emp['employee_id']] = emp.get('status', 'active')
            if emp.get('status', 'active') == 'active':
                active_employees.add(emp['employee_id'])
        
        # Identify accounts to delete
        accounts_to_delete = []
        duplicates = {}
        
        for user in users:
            employee_id = user.get('employee_id')
            
            # Track duplicates
            if employee_id in duplicates:
                duplicates[employee_id].append(user['_id'])
            else:
                duplicates[employee_id] = [user['_id']]
            
            # Mark for deletion if:
            # 1. Test account patterns
            if any(pattern in employee_id for pattern in ['TEST', 'MIN082', 'FULL082', 'BLANK082', 'AUTH082', 'BULK090', 'STATUS090', 'EDIT090', 'PIN135']):
                accounts_to_delete.append((user['_id'], f"Test account: {employee_id}"))
            
            # 2. Not in employees collection
            elif employee_id not in valid_employees:
                accounts_to_delete.append((user['_id'], f"No matching employee: {employee_id}"))
            
            # 3. Employee is not active
            elif valid_employees.get(employee_id) != 'active':
                status_val = valid_employees.get(employee_id, 'unknown')
                accounts_to_delete.append((user['_id'], f"Non-active employee ({status_val}): {employee_id}"))
        
        # Handle duplicates - keep only the most recent one per employee
        for employee_id, user_ids in duplicates.items():
            if len(user_ids) > 1:
                # Keep the first one, delete the rest
                for user_id in user_ids[1:]:
                    accounts_to_delete.append((user_id, f"Duplicate account: {employee_id}"))
        
        # Delete invalid accounts
        deleted_count = 0
        if accounts_to_delete:
            user_ids_to_delete = [user_id for user_id, _ in accounts_to_delete]
            result = await db.users.delete_many({"_id": {"$in": user_ids_to_delete}})
            deleted_count = result.deleted_count
        
        # Get final count
        remaining_users = await db.users.find({"role": "employee"}).to_list(length=None)
        active_user_count = sum(1 for user in remaining_users if user.get('employee_id') in active_employees)
        
        return {
            "message": "User account cleanup completed",
            "deleted_count": deleted_count,
            "remaining_users": len(remaining_users),
            "active_users": active_user_count,
            "active_employees": len(active_employees),
            "cleanup_details": [reason for _, reason in accounts_to_delete[:20]]  # First 20 reasons
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to cleanup user accounts: {str(e)}"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update employee PIN: {str(e)}"
        )

@api_router.put("/employee/change-pin")
async def change_employee_pin(
    request: EmployeePinChangeRequest,
    current_user: User = Depends(get_current_user)
):
    """Allow employee to change their own PIN"""
    try:
        if current_user.role != UserRole.EMPLOYEE:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only employees can change their PIN"
            )
        
        # Validate new PIN format (4 digits)
        if not request.new_pin.isdigit() or len(request.new_pin) != 4:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="New PIN must be exactly 4 digits"
            )
        
        # Verify current PIN
        if current_user.pin != request.current_pin:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Current PIN is incorrect"
            )
        
        # Update PIN
        result = await db.users.update_one(
            {"username": current_user.username, "role": "employee"},
            {"$set": {"pin": request.new_pin}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee account not found"
            )
        
        return {
            "message": "PIN changed successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to change PIN: {str(e)}"
        )

# Email Payslip Models
class EmailPayslipRequest(BaseModel):
    month: int
    year: int
    employee_ids: Optional[List[str]] = None  # If None, send to all

class EmailStatus(BaseModel):
    employee_id: str
    employee_name: str
    email: str
    status: str  # 'sent', 'failed', 'pending'
    error: Optional[str] = None

# PDF Generation using reportlab
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

def generate_payslip_pdf(payslip_data: dict, company_settings: dict) -> bytes:
    """Generate PDF payslip matching frontend format using reportlab"""
    buffer = io.BytesIO()
    
    # Create PDF with same margins as frontend
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.7*inch, bottomMargin=0.7*inch, leftMargin=0.7*inch, rightMargin=0.7*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Month name
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
    
    # Company Header Section
    header_data = [
        [Paragraph(f"<b><font size=18 color='#10b981'>{company_settings.get('company_name', 'Company Name')}</font></b>", styles['Normal']),
         Paragraph("<b><font size=14>SALARY SLIP</font></b>", ParagraphStyle('Right', parent=styles['Normal'], alignment=TA_RIGHT))],
        [Paragraph(f"<font size=9 color='#6b7280'>{company_settings.get('address', 'N/A')}</font>", styles['Normal']),
         Paragraph(f"<font size=9 color='#6b7280'>For the month of {month_names[payslip_data['month']]} {payslip_data['year']}</font>", ParagraphStyle('Right', parent=styles['Normal'], alignment=TA_RIGHT))],
    ]
    
    header_table = Table(header_data, colWidths=[4*inch, 3.2*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('LINEBELOW', (0, 1), (-1, 1), 2, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Employee Details - Two Column Layout
    emp_details_left = [
        ['<b>Employee Details</b>', ''],
        ['Name:', payslip_data['employee_name']],
        ['Employee ID:', payslip_data['employee_id']],
        ['Designation:', payslip_data.get('designation', 'N/A')],
        ['Department:', payslip_data.get('department', 'N/A')],
    ]
    
    emp_details_right = [
        ['<b>Bank Details</b>', ''],
        ['PAN:', payslip_data.get('pan', 'N/A')],
        ['Bank Account:', '***' + str(payslip_data.get('bank_account', 'N/A'))[-4:] if payslip_data.get('bank_account') else 'N/A'],
        ['IFSC Code:', payslip_data.get('ifsc', 'N/A')],
        ['Days Worked:', f"{payslip_data.get('days_worked', 30)} of {payslip_data.get('days_in_month', 30)}"],
    ]
    
    left_table = Table(emp_details_left, colWidths=[1.3*inch, 2*inch])
    left_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#6b7280')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    right_table = Table(emp_details_right, colWidths=[1.3*inch, 2*inch])
    right_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#6b7280')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    combined_emp_table = Table([[left_table, right_table]], colWidths=[3.6*inch, 3.6*inch])
    combined_emp_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(combined_emp_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Earnings Section
    earnings_data = [
        ['💰 Earnings', ''],
        ['Basic Salary:', f"₹{payslip_data.get('basic_salary', 0):,.2f}"],
        ['House Rent Allowance:', f"₹{payslip_data.get('hra', 0):,.2f}"],
        ['Medical Allowance:', f"₹{payslip_data.get('medical_allowance', 0):,.2f}"],
        ['Leave Travel Allowance:', f"₹{payslip_data.get('lta', 0):,.2f}"],
        ['Bonus:', f"₹{payslip_data.get('conveyance', 0):,.2f}"],
        ['Performance Incentive:', f"₹{payslip_data.get('performance_incentive', 0):,.2f}"],
        ['Other Benefits:', f"₹{payslip_data.get('other_benefits', 0):,.2f}"],
        ['<b>Gross Earnings:</b>', f"<b>₹{payslip_data.get('gross_salary', 0):,.2f}</b>"],
    ]
    
    earnings_table = Table(earnings_data, colWidths=[2.3*inch, 1.3*inch])
    earnings_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ecfdf5')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#065f46')),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#10b981')),
        ('LEFTPADDING', (0, 0), (0, -1), 4),
    ]))
    
    # Deductions Section
    deductions_data = [
        ['💸 Deductions', ''],
        ['PF (Employee):', f"₹{payslip_data.get('pf_employee', 0):,.2f}"],
        ['ESI (Employee):', f"₹{payslip_data.get('esi_employee', 0):,.2f}"],
        ['Professional Tax:', f"₹{payslip_data.get('professional_tax', 0):,.2f}"],
        ['TDS:', f"₹{payslip_data.get('tds', 0):,.2f}"],
        ['Loan Deductions:', f"₹{payslip_data.get('loan_deductions', 0):,.2f}"],
        ['Others:', f"₹{payslip_data.get('others', 0):,.2f}"],
        ['', ''],  # Spacer
        ['<b>Total Deductions:</b>', f"<b>₹{payslip_data.get('total_deductions', 0):,.2f}</b>"],
    ]
    
    deductions_table = Table(deductions_data, colWidths=[2.3*inch, 1.3*inch])
    deductions_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fef2f2')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#991b1b')),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#dc2626')),
        ('LEFTPADDING', (0, 0), (0, -1), 4),
    ]))
    
    # Combine earnings and deductions side by side
    combined_salary_table = Table([[earnings_table, deductions_table]], colWidths=[3.6*inch, 3.6*inch])
    combined_salary_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(combined_salary_table)
    elements.append(Spacer(1, 0.25*inch))
    
    # Net Salary Section
    net_salary_data = [[
        Paragraph('<font size=12><b>Net Salary:</b></font>', styles['Normal']),
        Paragraph(f'<font size=16 color="#10b981"><b>₹{payslip_data.get("net_salary", 0):,.2f}</b></font>', ParagraphStyle('Right', parent=styles['Normal'], alignment=TA_RIGHT))
    ]]
    
    net_table = Table(net_salary_data, colWidths=[4*inch, 3.2*inch])
    net_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
    ]))
    elements.append(net_table)
    elements.append(Spacer(1, 0.1*inch))
    
    # Generated date
    from datetime import datetime
    generated_text = Paragraph(
        f'<font size=8 color="#6b7280">Generated on: {datetime.now().strftime("%d %B %Y")}</font>',
        styles['Normal']
    )
    elements.append(generated_text)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()

# Email sending utility function
async def send_payslip_email(employee_email: str, employee_name: str, employee_id: str, payslip_data: dict, pdf_bytes: bytes, company_logo_url: str = None):
    """Send payslip email with PDF attachment"""
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"{employee_name} - Payslip - {payslip_data['month_name']} - {payslip_data['year']}"
        msg['From'] = SMTP_FROM
        msg['To'] = employee_email
        msg['Bcc'] = SMTP_BCC
        
        # Create HTML email body with salary breakdown
        logo_html = f'<img src="{company_logo_url}" alt="Company Logo" style="max-width: 150px; margin-bottom: 20px;">' if company_logo_url else ''
        
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ 
                    font-family: 'Aptos Display', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                    line-height: 1.6; 
                    color: #006B5B; 
                    margin: 0; 
                    padding: 20px; 
                }}
                .container {{ 
                    max-width: 600px; 
                    margin: 0 auto; 
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <p>Dear {employee_name},</p>
                <p>Your salary slip for {payslip_data['month_name']} {payslip_data['year']} has been processed successfully.</p>
                <p>Best regards,<br>
                {payslip_data.get('company_name', 'Company Name')} Accounts Team</p>
            </div>
        </body>
        </html>
        """
        
        html_part = MIMEText(html_body, 'html')
        msg.attach(html_part)
        
        # Attach PDF with proper filename
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(pdf_bytes)
        encoders.encode_base64(part)
        
        # Format: EmployeeName_Month_Year.pdf (e.g., JohnDoe_October_2025.pdf)
        safe_name = employee_name.replace(' ', '_')
        pdf_filename = f"{safe_name}_{payslip_data['month_name']}_{payslip_data['year']}.pdf"
        
        part.add_header(
            'Content-Disposition',
            f'attachment; filename="{pdf_filename}"'
        )
        msg.attach(part)
        
        # Send email
        await aiosmtplib.send(
            msg,
            hostname=SMTP_HOST,
            port=SMTP_PORT,
            username=SMTP_USER,
            password=SMTP_PASSWORD,
            start_tls=True
        )
        
        return True, None
        
    except Exception as e:
        logging.error(f"Error sending email to {employee_email}: {str(e)}")
        return False, str(e)

@api_router.post("/payslips/email-with-pdfs")
async def email_payslips_with_pdfs(
    pdfs: List[UploadFile] = File(...),
    employee_ids: List[str] = Form(...),
    employee_names: List[str] = Form(...),
    employee_emails: List[str] = Form(...),
    month: int = Form(...),
    year: int = Form(...),
    request_id: str = Form(None),
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Send payslips via email with pre-generated PDFs from frontend"""
    try:
        # Log request for debugging
        if request_id:
            logger.info(f"Processing email request with ID: {request_id}")
        else:
            logger.info(f"Processing email request without ID for {len(employee_ids)} employees")
        if len(pdfs) != len(employee_ids):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Mismatch between number of PDFs and employees"
            )
        
        # Get company settings from the settings collection
        settings = await db.settings.find_one({})
        logger.info(f"Retrieved settings from DB: {settings}")
        if settings and 'company_settings' in settings:
            company_settings = settings['company_settings']
            logger.info(f"Using company_settings: {company_settings}")
        else:
            company_settings = {"company_name": "Company Name", "address": "N/A"}
            logger.info("Using fallback company settings")
        
        company_logo_url = company_settings.get('logo_url')
        
        email_results = []
        successful = 0
        failed = 0
        
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        
        # Send emails with provided PDFs
        for i, pdf_file in enumerate(pdfs):
            employee_id = employee_ids[i]
            employee_name = employee_names[i]
            employee_email = employee_emails[i]
            
            if not employee_email or employee_email == 'N/A':
                email_results.append({
                    "employee_id": employee_id,
                    "employee_name": employee_name,
                    "email": "N/A",
                    "status": "failed",
                    "error": "No email address"
                })
                failed += 1
                continue
            
            # Read PDF bytes
            pdf_bytes = await pdf_file.read()
            
            # Get payslip data for email body
            payslip = await db.payslips.find_one({
                "employee_id": employee_id,
                "month": month,
                "year": year
            })
            
            if not payslip:
                email_results.append({
                    "employee_id": employee_id,
                    "employee_name": employee_name,
                    "email": employee_email,
                    "status": "failed",
                    "error": "Payslip not found"
                })
                failed += 1
                continue
            
            # Prepare payslip data for email body
            payslip_data = {
                "month": month,
                "year": year,
                "month_name": month_names[month],
                "company_name": company_settings.get('company_name', 'Company Name'),
                "basic_salary": payslip.get('basic_salary', 0),
                "hra": payslip.get('hra', 0),
                "medical_allowance": payslip.get('medical_allowance', 0),
                "lta": payslip.get('lta', 0),
                "conveyance": payslip.get('conveyance', 0),
                "performance_incentive": payslip.get('performance_incentive', 0),
                "other_benefits": payslip.get('other_benefits', 0),
                "gross_salary": payslip.get('gross_salary', 0),
                "pf_employee": payslip.get('pf_employee', 0),
                "esi_employee": payslip.get('esi_employee', 0),
                "professional_tax": payslip.get('professional_tax', 0),
                "tds": payslip.get('tds', 0),
                "loan_deductions": payslip.get('loan_deductions', 0),
                "others": payslip.get('others', 0),
                "total_deductions": payslip.get('total_deductions', 0),
                "net_salary": payslip.get('net_salary', 0),
                "days_worked": payslip.get('days_worked', 30),
                "days_in_month": payslip.get('days_in_month', 30)
            }
            
            # Send email
            success, error = await send_payslip_email(
                employee_email,
                employee_name,
                employee_id,
                payslip_data,
                pdf_bytes,
                company_logo_url
            )
            
            email_results.append({
                "employee_id": employee_id,
                "employee_name": employee_name,
                "email": employee_email,
                "status": "sent" if success else "failed",
                "error": error
            })
            
            if success:
                successful += 1
                # Update payslip with email sent status
                await db.payslips.update_one(
                    {"employee_id": employee_id, "month": month, "year": year},
                    {"$set": {"email_sent": True, "email_sent_at": datetime.now(timezone.utc).isoformat()}}
                )
            else:
                failed += 1
            
            # Small delay to avoid spam limits (reduced for better performance)
            await asyncio.sleep(0.1)
        
        return {
            "message": f"Email sending completed: {successful} sent, {failed} failed",
            "total": len(pdfs),
            "successful": successful,
            "failed": failed,
            "results": email_results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error sending payslip emails: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to send emails: {str(e)}"
        )

@api_router.post("/payslips/email")
async def email_payslips(
    request: EmailPayslipRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Send payslips via email to employees (OLD - kept for compatibility)"""
    try:
        # Get payslips for the specified month/year
        query = {"month": request.month, "year": request.year}
        if request.employee_ids:
            query["employee_id"] = {"$in": request.employee_ids}
        
        payslips = await db.payslips.find(query).to_list(length=None)
        
        if not payslips:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No payslips found for {request.month}/{request.year}"
            )
        
        # Get company settings for PDF generation
        settings = await db.settings.find_one({})
        if settings and 'company_settings' in settings:
            company_settings = settings['company_settings']
        else:
            company_settings = {
                "company_name": "Company Name",
                "address": "Company Address"
            }
        
        email_results = []
        successful = 0
        failed = 0
        
        # Send emails with delay (individual sending to avoid spam limits)
        for payslip in payslips:
            employee_id = payslip.get('employee_id')
            
            # Get employee details
            employee = await db.employees.find_one({"employee_id": employee_id})
            if not employee or not employee.get('email'):
                email_results.append({
                    "employee_id": employee_id,
                    "employee_name": employee.get('name', 'Unknown') if employee else 'Unknown',
                    "email": "N/A",
                    "status": "failed",
                    "error": "No email address"
                })
                failed += 1
                continue
            
            # Prepare payslip data with ALL fields
            month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                          'July', 'August', 'September', 'October', 'November', 'December']
            
            payslip_data = {
                "month": request.month,
                "year": request.year,
                "month_name": month_names[request.month],
                "employee_id": employee_id,
                "employee_name": employee.get('name', 'Employee'),
                "department": employee.get('department', 'N/A'),
                "designation": employee.get('designation', 'N/A'),
                "pan": employee.get('pan_number', 'N/A'),
                "bank_account": employee.get('bank_info', {}).get('account_number', 'N/A'),
                "ifsc": employee.get('bank_info', {}).get('ifsc_code', 'N/A'),
                "company_name": company_settings.get('company_name', 'Company Name'),
                # Earnings
                "basic_salary": payslip.get('basic_salary', 0),
                "hra": payslip.get('hra', 0),
                "medical_allowance": payslip.get('medical_allowance', 0),
                "lta": payslip.get('lta', 0),
                "conveyance": payslip.get('conveyance', 0),
                "performance_incentive": payslip.get('performance_incentive', 0),
                "other_benefits": payslip.get('other_benefits', 0),
                "gross_salary": payslip.get('gross_salary', 0),
                # Deductions
                "pf_employee": payslip.get('pf_employee', 0),
                "esi_employee": payslip.get('esi_employee', 0),
                "professional_tax": payslip.get('professional_tax', 0),
                "tds": payslip.get('tds', 0),
                "loan_deductions": payslip.get('loan_deductions', 0),
                "others": payslip.get('others', 0),
                "total_deductions": payslip.get('total_deductions', 0),
                # Net
                "net_salary": payslip.get('net_salary', 0),
                "days_worked": payslip.get('days_worked', 30),
                "days_in_month": payslip.get('days_in_month', 30)
            }
            
            # Generate PDF
            try:
                pdf_bytes = generate_payslip_pdf(payslip_data, company_settings)
            except Exception as pdf_error:
                logging.error(f"Error generating PDF for {employee_id}: {str(pdf_error)}")
                email_results.append({
                    "employee_id": employee_id,
                    "employee_name": employee.get('name'),
                    "email": employee.get('email'),
                    "status": "failed",
                    "error": f"PDF generation failed: {str(pdf_error)}"
                })
                failed += 1
                continue
            
            # Get company logo URL if available
            company_logo_url = company_settings.get('logo_url')
            
            # Send email
            success, error = await send_payslip_email(
                employee.get('email'),
                employee.get('name', 'Employee'),
                employee_id,
                payslip_data,
                pdf_bytes,
                company_logo_url
            )
            
            email_results.append({
                "employee_id": employee_id,
                "employee_name": employee.get('name'),
                "email": employee.get('email'),
                "status": "sent" if success else "failed",
                "error": error
            })
            
            if success:
                successful += 1
                # Update payslip with email sent status
                await db.payslips.update_one(
                    {"_id": payslip["_id"]},
                    {"$set": {"email_sent": True, "email_sent_at": datetime.now(timezone.utc).isoformat()}}
                )
            else:
                failed += 1
            
            # Small delay to avoid spam limits (0.5 seconds)
            await asyncio.sleep(0.5)
        
        return {
            "message": f"Email sending completed: {successful} sent, {failed} failed",
            "total": len(payslips),
            "successful": successful,
            "failed": failed,
            "results": email_results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error sending payslip emails: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to send emails: {str(e)}"
        )

# Backup and Restore Endpoints
@api_router.get("/backup/download")
async def download_backup(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Download complete database backup as JSON"""
    try:
        backup_data = {
            "backup_date": datetime.now(timezone.utc).isoformat(),
            "version": "1.0",
            "collections": {}
        }
        
        # List of all collections to backup
        collection_names = [
            "employees",
            "users", 
            "payslips",
            "leave_requests",
            "leave_balances",
            "loans",
            "payroll_runs",
            "company_settings"
        ]
        
        for collection_name in collection_names:
            try:
                collection = db[collection_name]
                documents = await collection.find({}).to_list(length=None)
                
                # Convert ObjectId and datetime to strings for JSON serialization
                serialized_docs = []
                for doc in documents:
                    # Remove _id (MongoDB ObjectId) as it's not JSON serializable
                    if '_id' in doc:
                        del doc['_id']
                    
                    # Convert any remaining datetime objects to ISO strings
                    for key, value in doc.items():
                        if isinstance(value, datetime):
                            doc[key] = value.isoformat()
                        elif isinstance(value, date):
                            doc[key] = value.isoformat()
                    
                    serialized_docs.append(doc)
                
                backup_data["collections"][collection_name] = serialized_docs
                logger.info(f"Backed up {len(serialized_docs)} documents from {collection_name}")
                
            except Exception as e:
                logging.error(f"Error backing up collection {collection_name}: {str(e)}")
                backup_data["collections"][collection_name] = []
        
        # Return as JSON response with download headers
        from fastapi.responses import JSONResponse
        
        return JSONResponse(
            content=backup_data,
            headers={
                "Content-Disposition": f"attachment; filename=payroll_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            }
        )
        
    except Exception as e:
        logging.error(f"Error creating backup: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create backup: {str(e)}"
        )

@api_router.post("/backup/restore")
async def restore_backup(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Restore database from backup JSON file"""
    try:
        # Read and parse the uploaded file
        contents = await file.read()
        backup_data = json.loads(contents)
        
        if "collections" not in backup_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid backup file format"
            )
        
        restore_stats = {
            "restored_collections": [],
            "total_documents": 0,
            "errors": []
        }
        
        # Restore each collection
        for collection_name, documents in backup_data["collections"].items():
            try:
                if not documents:
                    continue
                
                collection = db[collection_name]
                
                # Clear existing data (optional - can be made configurable)
                # await collection.delete_many({})
                
                # Insert documents
                if documents:
                    # Convert ISO date strings back to datetime objects where needed
                    processed_docs = []
                    for doc in documents:
                        # Convert date strings back to date/datetime objects
                        for key, value in doc.items():
                            if isinstance(value, str):
                                # Try to parse as datetime
                                try:
                                    if 'T' in value and ('+' in value or 'Z' in value):
                                        doc[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                    elif '-' in value and len(value) == 10:  # YYYY-MM-DD
                                        doc[key] = datetime.fromisoformat(value).date()
                                except (ValueError, AttributeError):
                                    pass
                        processed_docs.append(doc)
                    
                    # Use insert_many with ordered=False to continue on duplicates
                    result = await collection.insert_many(processed_docs, ordered=False)
                    
                    restore_stats["restored_collections"].append(collection_name)
                    restore_stats["total_documents"] += len(result.inserted_ids)
                    logger.info(f"Restored {len(result.inserted_ids)} documents to {collection_name}")
                    
            except Exception as e:
                error_msg = f"Error restoring {collection_name}: {str(e)}"
                logging.error(error_msg)
                restore_stats["errors"].append(error_msg)
        
        return {
            "message": "Backup restored successfully",
            "stats": restore_stats
        }
        
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid JSON file"
        )
    except Exception as e:
        logging.error(f"Error restoring backup: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to restore backup: {str(e)}"
        )

@api_router.get("/backup/info")
async def get_backup_info(current_user: User = Depends(require_role(UserRole.ADMIN))):
    """Get backup information and statistics"""
    try:
        stats = {}
        collection_names = ["employees", "users", "payslips", "leave_requests", "leave_balances", "loans", "payroll_runs", "company_settings"]
        
        total_documents = 0
        for collection_name in collection_names:
            count = await db[collection_name].count_documents({})
            stats[collection_name] = count
            total_documents += count
        
        return {
            "total_documents": total_documents,
            "collections": stats,
            "last_backup": None,  # Can be implemented with a backup log collection
            "backup_available": True
        }
        
    except Exception as e:
        logging.error(f"Error getting backup info: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get backup info: {str(e)}"
        )

# ================ EVENT MANAGEMENT ENDPOINTS ================

@api_router.get("/events")
async def get_events():
    """Get all company events (accessible to all employees)"""
    try:
        events = await db.events.find({}, {"_id": 0}).to_list(length=None)
        return events
    except Exception as e:
        logging.error(f"Error fetching events: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch events"
        )

@api_router.post("/events")
async def create_event(
    event: EventCreate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Create a new company event (admin only)"""
    try:
        # Validate date format
        from datetime import datetime
        try:
            datetime.fromisoformat(event.date)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use ISO format (YYYY-MM-DD)"
            )
        
        new_event = Event(
            **event.dict(),
            created_by=current_user.employee_id or current_user.username
        )
        
        event_dict = new_event.dict()
        event_dict["created_at"] = event_dict["created_at"].isoformat()
        event_dict["updated_at"] = event_dict["updated_at"].isoformat()
        
        await db.events.insert_one(event_dict)
        
        return {
            "message": "Event created successfully",
            "event": new_event
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating event: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create event"
        )

@api_router.get("/events/{event_id}")
async def get_event(event_id: str):
    """Get a specific event by ID"""
    try:
        event = await db.events.find_one({"id": event_id}, {"_id": 0})
        if not event:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Event not found"
            )
        return event
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching event: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch event"
        )

@api_router.put("/events/{event_id}")
async def update_event(
    event_id: str,
    event_update: EventUpdate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Update an existing event (admin only)"""
    try:
        # Build update dict with only provided fields
        update_data = {k: v for k, v in event_update.dict().items() if v is not None}
        
        if not update_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No update data provided"
            )
        
        # Validate date format if provided
        if "date" in update_data:
            from datetime import datetime
            try:
                datetime.fromisoformat(update_data["date"])
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Invalid date format. Use ISO format (YYYY-MM-DD)"
                )
        
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        result = await db.events.update_one(
            {"id": event_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Event not found"
            )
        
        # Fetch and return updated event
        updated_event = await db.events.find_one({"id": event_id}, {"_id": 0})
        
        return {
            "message": "Event updated successfully",
            "event": updated_event
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating event: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update event"
        )

@api_router.delete("/events/{event_id}")
async def delete_event(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete an event (admin only)"""
    try:
        result = await db.events.delete_one({"id": event_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Event not found"
            )
        
        return {"message": "Event deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting event: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete event"
        )


# ==================== Company Bank Accounts API ====================

@api_router.get("/company-bank-accounts")
async def get_company_bank_accounts(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get all company bank accounts"""
    try:
        accounts = await db.company_bank_accounts.find({}, {"_id": 0}).to_list(length=None)
        return accounts
    except Exception as e:
        logging.error(f"Error fetching company bank accounts: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch company bank accounts"
        )

@api_router.post("/company-bank-accounts")
async def create_company_bank_account(
    account: CompanyBankAccountCreate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Create a new company bank account"""
    try:
        # Check for duplicate account number
        existing = await db.company_bank_accounts.find_one({
            "account_number": account.account_number
        })
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Account number already exists"
            )
        
        new_account = CompanyBankAccount(**account.dict())
        await db.company_bank_accounts.insert_one(new_account.dict())
        
        return {
            "message": "Company bank account created successfully",
            "account": new_account.dict()
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating company bank account: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create company bank account"
        )

@api_router.put("/company-bank-accounts/{account_id}")
async def update_company_bank_account(
    account_id: str,
    account_update: CompanyBankAccountUpdate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Update a company bank account"""
    try:
        # Check if account exists
        existing_account = await db.company_bank_accounts.find_one({"id": account_id})
        if not existing_account:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company bank account not found"
            )
        
        # Prepare update data
        update_data = {k: v for k, v in account_update.dict(exclude_unset=True).items() if v is not None}
        
        if update_data:
            update_data["updated_at"] = datetime.now(timezone.utc)
            
            # Check for duplicate account number if being updated
            if "account_number" in update_data:
                duplicate = await db.company_bank_accounts.find_one({
                    "account_number": update_data["account_number"],
                    "id": {"$ne": account_id}
                })
                if duplicate:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Account number already exists"
                    )
            
            await db.company_bank_accounts.update_one(
                {"id": account_id},
                {"$set": update_data}
            )
        
        updated_account = await db.company_bank_accounts.find_one({"id": account_id})
        return {
            "message": "Company bank account updated successfully",
            "account": updated_account
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating company bank account: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update company bank account"
        )

@api_router.delete("/company-bank-accounts/{account_id}")
async def delete_company_bank_account(
    account_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete a company bank account"""
    try:
        # Check if any employees are mapped to this account
        mappings_count = await db.employee_source_mapping.count_documents({
            "company_account_id": account_id
        })
        
        if mappings_count > 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Cannot delete account. {mappings_count} employees are mapped to this account."
            )
        
        result = await db.company_bank_accounts.delete_one({"id": account_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company bank account not found"
            )
        
        return {"message": "Company bank account deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting company bank account: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete company bank account"
        )

# ==================== Employee Source Mapping API ====================

@api_router.get("/employee-source-mapping")
async def get_employee_source_mappings(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get all employee source mappings with employee and account details"""
    try:
        mappings = await db.employee_source_mapping.find({}, {"_id": 0}).to_list(length=None)
        
        # Enrich mappings with employee and account details
        enriched_mappings = []
        for mapping in mappings:
            # Get employee details
            employee = await db.employees.find_one({"employee_id": mapping["employee_id"]}, {"_id": 0})
            # Get company account details
            account = await db.company_bank_accounts.find_one({"id": mapping["company_account_id"]}, {"_id": 0})
            
            enriched_mapping = {
                **mapping,
                "employee_name": employee.get("name", "Unknown") if employee else "Unknown",
                "employee_department": employee.get("department", "Unknown") if employee else "Unknown",
                "company_account_name": account.get("account_name", "Unknown") if account else "Unknown",
                "company_bank_name": account.get("bank_name", "Unknown") if account else "Unknown"
            }
            enriched_mappings.append(enriched_mapping)
        
        return enriched_mappings
    except Exception as e:
        logging.error(f"Error fetching employee source mappings: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch employee source mappings"
        )

@api_router.post("/employee-source-mapping")
async def create_employee_source_mapping(
    mapping: EmployeeSourceMappingCreate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Create or update employee source mapping"""
    try:
        # Validate employee exists
        employee = await db.employees.find_one({"employee_id": mapping.employee_id})
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Employee {mapping.employee_id} not found"
            )
        
        # Validate company account exists
        account = await db.company_bank_accounts.find_one({"id": mapping.company_account_id})
        if not account:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company bank account not found"
            )
        
        # Check if mapping already exists
        existing_mapping = await db.employee_source_mapping.find_one({
            "employee_id": mapping.employee_id
        })
        
        if existing_mapping:
            # Update existing mapping
            await db.employee_source_mapping.update_one(
                {"employee_id": mapping.employee_id},
                {
                    "$set": {
                        "company_account_id": mapping.company_account_id,
                        "updated_at": datetime.now(timezone.utc)
                    }
                }
            )
            return {
                "message": "Employee source mapping updated successfully",
                "mapping_id": existing_mapping["id"]
            }
        else:
            # Create new mapping
            new_mapping = EmployeeSourceMapping(**mapping.dict())
            await db.employee_source_mapping.insert_one(new_mapping.dict())
            return {
                "message": "Employee source mapping created successfully",
                "mapping_id": new_mapping.id
            }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating employee source mapping: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create employee source mapping"
        )

@api_router.post("/employee-source-mapping/bulk")
async def create_bulk_employee_source_mapping(
    bulk_mapping: BulkEmployeeSourceMappingCreate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Create or update multiple employee source mappings"""
    try:
        # Validate company account exists
        account = await db.company_bank_accounts.find_one({"id": bulk_mapping.company_account_id})
        if not account:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Company bank account not found"
            )
        
        success_count = 0
        failed_count = 0
        errors = []
        
        for employee_id in bulk_mapping.employee_ids:
            try:
                # Validate employee exists
                employee = await db.employees.find_one({"employee_id": employee_id})
                if not employee:
                    failed_count += 1
                    errors.append(f"Employee {employee_id} not found")
                    continue
                
                # Check if mapping already exists
                existing_mapping = await db.employee_source_mapping.find_one({
                    "employee_id": employee_id
                })
                
                if existing_mapping:
                    # Update existing mapping
                    await db.employee_source_mapping.update_one(
                        {"employee_id": employee_id},
                        {
                            "$set": {
                                "company_account_id": bulk_mapping.company_account_id,
                                "updated_at": datetime.now(timezone.utc)
                            }
                        }
                    )
                else:
                    # Create new mapping
                    new_mapping = EmployeeSourceMapping(
                        employee_id=employee_id,
                        company_account_id=bulk_mapping.company_account_id
                    )
                    await db.employee_source_mapping.insert_one(new_mapping.dict())
                
                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append(f"Error with employee {employee_id}: {str(e)}")
        
        return {
            "message": "Bulk employee source mapping completed",
            "success_count": success_count,
            "failed_count": failed_count,
            "errors": errors if errors else None
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating bulk employee source mapping: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create bulk employee source mapping"
        )

@api_router.delete("/employee-source-mapping/{mapping_id}")
async def delete_employee_source_mapping(
    mapping_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete an employee source mapping"""
    try:
        result = await db.employee_source_mapping.delete_one({"id": mapping_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee source mapping not found"
            )
        
        return {"message": "Employee source mapping deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting employee source mapping: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete employee source mapping"
        )

@api_router.delete("/employee-source-mapping/by-employee/{employee_id}")
async def delete_employee_source_mapping_by_employee(
    employee_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete employee source mapping by employee_id"""
    try:
        result = await db.employee_source_mapping.delete_one({"employee_id": employee_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Employee source mapping not found"
            )
        
        return {"message": "Employee source mapping deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting employee source mapping: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete employee source mapping"
        )

# ==================== Bank Template API ====================

@api_router.get("/bank-templates")
async def get_bank_templates(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get all bank templates"""
    try:
        templates = await db.bank_templates.find({}, {"_id": 0}).to_list(length=None)
        # Don't send full template_data in list view to reduce response size
        for template in templates:
            template["template_size"] = len(template.get("template_data", ""))
            template.pop("template_data", None)
        return templates
    except Exception as e:
        logging.error(f"Error fetching bank templates: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch bank templates"
        )

@api_router.post("/bank-templates")
async def upload_bank_template(
    template: BankTemplateUpload,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Upload a bank template"""
    try:
        # Validate base64 data
        try:
            base64.b64decode(template.template_data)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid base64 template data"
            )
        
        # Check if template for this bank already exists
        existing_template = await db.bank_templates.find_one({
            "bank_name": template.bank_name
        })
        
        if existing_template:
            # Update existing template
            await db.bank_templates.update_one(
                {"bank_name": template.bank_name},
                {
                    "$set": {
                        "template_data": template.template_data,
                        "file_name": template.file_name,
                        "uploaded_at": datetime.now(timezone.utc),
                        "uploaded_by": current_user.username
                    }
                }
            )
            return {
                "message": f"Bank template for {template.bank_name} updated successfully",
                "template_id": existing_template["id"]
            }
        else:
            # Create new template
            new_template = BankTemplate(
                **template.dict(),
                uploaded_by=current_user.username
            )
            await db.bank_templates.insert_one(new_template.dict())
            return {
                "message": f"Bank template for {template.bank_name} uploaded successfully",
                "template_id": new_template.id
            }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error uploading bank template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to upload bank template"
        )

@api_router.get("/bank-templates/{template_id}")
async def get_bank_template(
    template_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get a specific bank template with full data"""
    try:
        template = await db.bank_templates.find_one({"id": template_id}, {"_id": 0})
        if not template:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Bank template not found"
            )
        return template
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching bank template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch bank template"
        )

@api_router.delete("/bank-templates/{template_id}")
async def delete_bank_template(
    template_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Delete a bank template"""
    try:
        result = await db.bank_templates.delete_one({"id": template_id})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Bank template not found"
            )
        
        return {"message": "Bank template deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting bank template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete bank template"
        )

@api_router.get("/bank-templates/by-bank/{bank_name}")
async def get_bank_template_by_name(
    bank_name: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get bank template by bank name"""
    try:
        template = await db.bank_templates.find_one({"bank_name": bank_name}, {"_id": 0})
        if not template:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Bank template for {bank_name} not found"
            )
        return template
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching bank template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch bank template"
        )

# ==================== Bank Advice Generation API ====================

@api_router.post("/bank-advice/generate")
async def generate_bank_advice(
    request: BankAdviceGenerate,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Generate bank advice for salary disbursement"""
    try:
        # Get all active employees
        employees = await db.employees.find({"status": "active"}, {"_id": 0}).to_list(length=None)
        
        if not employees:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No active employees found"
            )
        
        # Get employee source mappings
        mappings = await db.employee_source_mapping.find({}, {"_id": 0}).to_list(length=None)
        mapping_dict = {m["employee_id"]: m["company_account_id"] for m in mappings}
        
        # Get company accounts
        company_accounts = await db.company_bank_accounts.find({"is_active": True}, {"_id": 0}).to_list(length=None)
        account_dict = {acc["id"]: acc for acc in company_accounts}
        
        # Get payroll data for the month
        payroll_data = await db.payroll_runs.find_one(
            {"month": request.month, "year": request.year},
            {"_id": 0}
        )
        
        # Group employees by company account
        accounts_data = {}
        unmapped_employees = []
        
        for employee in employees:
            emp_id = employee["employee_id"]
            company_account_id = mapping_dict.get(emp_id)
            
            if not company_account_id:
                unmapped_employees.append(emp_id)
                continue
            
            if company_account_id not in accounts_data:
                accounts_data[company_account_id] = {
                    "account": account_dict.get(company_account_id),
                    "employees": []
                }
            
            # Calculate salary
            salary_structure = employee.get("salary_structure", {})
            gross_salary = (
                salary_structure.get("basic_salary", 0) +
                salary_structure.get("hra", 0) +
                salary_structure.get("medical_allowance", 0) +
                salary_structure.get("travel_allowance", 0) +
                salary_structure.get("food_allowance", 0) +
                salary_structure.get("internet_allowance", 0) +
                salary_structure.get("special_allowance", 0)
            )
            
            total_deductions = (
                salary_structure.get("pf_employee", 0) +
                salary_structure.get("esi_employee", 0) +
                salary_structure.get("professional_tax", 0) +
                salary_structure.get("tds", 0)
            )
            
            net_salary = gross_salary - total_deductions
            
            accounts_data[company_account_id]["employees"].append({
                "employee_id": emp_id,
                "employee_name": employee.get("name"),
                "account_number": employee.get("bank_info", {}).get("account_number"),
                "ifsc_code": employee.get("bank_info", {}).get("ifsc_code"),
                "bank_name": employee.get("bank_info", {}).get("bank_name"),
                "net_salary": net_salary
            })
        
        # Filter by specific account if requested
        if request.company_account_id:
            if request.company_account_id not in accounts_data:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="No employees mapped to the selected company account"
                )
            accounts_data = {request.company_account_id: accounts_data[request.company_account_id]}
        
        # Create bank advice records
        bank_advices = []
        for account_id, data in accounts_data.items():
            if not data["account"]:
                continue
            
            account = data["account"]
            employees_list = data["employees"]
            
            total_amount = sum(emp["net_salary"] for emp in employees_list)
            employee_count = len(employees_list)
            
            logging.info(f"Bank Advice - Account: {account['account_name']}, Employees: {employee_count}, Total Amount: {total_amount}")
            
            # Get unique banks from employees
            unique_banks = len(set(emp["bank_name"] for emp in employees_list if emp["bank_name"]))
            
            # Generate reference number
            month_abbr = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", 
                         "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"][request.month - 1]
            
            existing_count = await db.bank_advices.count_documents({
                "month": request.month,
                "year": request.year
            })
            
            reference_number = f"BA/{month_abbr}{request.year}/{str(existing_count + 1).zfill(3)}"
            
            # Create bank advice record
            bank_advice = BankAdvice(
                month=request.month,
                year=request.year,
                reference_number=reference_number,
                company_account_id=account_id,
                template_id=request.template_id,
                total_amount=total_amount,
                employee_count=employee_count,
                status="generated"
            )
            
            # Save to database
            await db.bank_advices.insert_one(bank_advice.dict())
            
            bank_advices.append({
                **bank_advice.dict(),
                "company_account_name": account["account_name"],
                "bank_name": account["bank_name"],
                "bank_count": unique_banks
            })
        
        return {
            "message": f"Generated {len(bank_advices)} bank advice(s) successfully",
            "bank_advices": bank_advices,
            "unmapped_employees": unmapped_employees if unmapped_employees else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating bank advice: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate bank advice: {str(e)}"
        )

@api_router.get("/bank-advice")
async def get_bank_advices(
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Get all bank advices"""
    try:
        advices = await db.bank_advices.find({}, {"_id": 0}).sort("generated_date", -1).to_list(length=None)
        
        # Enrich with company account details
        for advice in advices:
            account = await db.company_bank_accounts.find_one(
                {"id": advice["company_account_id"]},
                {"_id": 0}
            )
            if account:
                advice["company_account_name"] = account.get("account_name")
                advice["company_bank_name"] = account.get("bank_name")
        
        return advices
    except Exception as e:
        logging.error(f"Error fetching bank advices: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to fetch bank advices"
        )

@api_router.get("/bank-advice/{advice_id}/download")
async def download_bank_advice(
    advice_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN))
):
    """Download bank advice as Excel file"""
    try:
        # Get bank advice
        advice = await db.bank_advices.find_one({"id": advice_id}, {"_id": 0})
        if not advice:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Bank advice not found"
            )
        
        # Get company account
        account = await db.company_bank_accounts.find_one(
            {"id": advice["company_account_id"]},
            {"_id": 0}
        )
        
        # Get employee source mappings for this account
        mappings = await db.employee_source_mapping.find(
            {"company_account_id": advice["company_account_id"]},
            {"_id": 0}
        ).to_list(length=None)
        
        employee_ids = [m["employee_id"] for m in mappings]
        
        # Get employees
        employees = await db.employees.find(
            {"employee_id": {"$in": employee_ids}, "status": "active"},
            {"_id": 0}
        ).to_list(length=None)
        
        # Check if template was used
        template = None
        if advice.get("template_id"):
            template = await db.bank_templates.find_one(
                {"id": advice["template_id"]},
                {"_id": 0}
            )
        
        # If template exists, use it; otherwise use standard format
        if template:
            # Decode base64 template
            template_data = base64.b64decode(template["template_data"])
            template_file = io.BytesIO(template_data)
            wb = load_workbook(template_file)
            ws = wb.active
            
            # Find the header row by searching for specific keywords
            header_row = None
            header_mapping = {}
            
            for row in range(1, 30):  # Check first 30 rows
                row_values = []
                for col in range(1, 20):  # Check first 20 columns
                    cell_value = str(ws.cell(row=row, column=col).value or "").strip()
                    row_values.append(cell_value.lower())
                
                # Check if this row contains header keywords
                if any("debit account" in val for val in row_values) or \
                   any("beneficiary name" in val for val in row_values) or \
                   any("transaction amount" in val for val in row_values):
                    header_row = row
                    
                    # Map headers to column numbers
                    for col in range(1, 20):
                        cell_value = str(ws.cell(row=row, column=col).value or "").strip().lower()
                        
                        if "debit account" in cell_value and "number" in cell_value:
                            header_mapping["debit_account"] = col
                        elif "transaction amount" in cell_value:
                            header_mapping["transaction_amount"] = col
                        elif "transaction currency" in cell_value or ("currency" in cell_value and "transaction" in cell_value):
                            header_mapping["currency"] = col
                        elif "beneficiary name" in cell_value and "nickname" not in cell_value:
                            header_mapping["beneficiary_name"] = col
                        elif "beneficiary account" in cell_value and "number" in cell_value:
                            header_mapping["beneficiary_account"] = col
                        elif ("beneficiary ifsc" in cell_value or "ifsc code" in cell_value) and "beneficiary" in cell_value:
                            header_mapping["ifsc_code"] = col
                        elif "transaction date" in cell_value:
                            header_mapping["transaction_date"] = col
                        elif "payment mode" in cell_value:
                            header_mapping["payment_mode"] = col
                        elif "reference number" in cell_value and "customer" in cell_value:
                            header_mapping["reference_number"] = col
                        elif ("nickname" in cell_value or "beneficiary nickname" in cell_value) and "beneficiary" in cell_value:
                            header_mapping["beneficiary_code"] = col
                            logging.info(f"Found Beneficiary Nickname/Code at column {col}: {cell_value}")
                    
                    break
            
            if not header_row:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Could not identify template headers. Please ensure the template has proper headers."
                )
            
            logging.info(f"Header row found at: {header_row}")
            logging.info(f"Header mapping: {header_mapping}")
            
            # Data starts from the row after header
            data_start_row = header_row + 1
            
            # Get company account number for debit account
            company_account_number = account.get("account_number", "") if account else ""
            
            # Populate employee data
            current_row = data_start_row
            for idx, employee in enumerate(employees, 1):
                salary_structure = employee.get("salary_structure", {})
                
                # Calculate net salary
                gross_salary = (
                    salary_structure.get("basic_salary", 0) +
                    salary_structure.get("hra", 0) +
                    salary_structure.get("medical_allowance", 0) +
                    salary_structure.get("travel_allowance", 0) +
                    salary_structure.get("food_allowance", 0) +
                    salary_structure.get("internet_allowance", 0) +
                    salary_structure.get("special_allowance", 0)
                )
                
                total_deductions = (
                    salary_structure.get("pf_employee", 0) +
                    salary_structure.get("esi_employee", 0) +
                    salary_structure.get("professional_tax", 0) +
                    salary_structure.get("tds", 0)
                )
                
                net_salary = gross_salary - total_deductions
                bank_info = employee.get("bank_info", {})
                
                # Populate data based on header mapping
                if "debit_account" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["debit_account"], value=company_account_number)
                
                if "transaction_amount" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["transaction_amount"], value=net_salary)
                
                if "currency" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["currency"], value="INR")
                
                if "beneficiary_name" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["beneficiary_name"], value=employee.get("name", ""))
                
                if "beneficiary_account" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["beneficiary_account"], value=bank_info.get("account_number", ""))
                
                if "ifsc_code" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["ifsc_code"], value=bank_info.get("ifsc_code", ""))
                
                if "transaction_date" in header_mapping:
                    # Use current date or last day of the month
                    from datetime import date
                    transaction_date = date(advice["year"], advice["month"], 1)
                    # Get last day of month
                    if advice["month"] == 12:
                        next_month = date(advice["year"] + 1, 1, 1)
                    else:
                        next_month = date(advice["year"], advice["month"] + 1, 1)
                    last_day = (next_month - timedelta(days=1)).strftime("%d/%m/%Y")
                    ws.cell(row=current_row, column=header_mapping["transaction_date"], value=last_day)
                
                if "payment_mode" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["payment_mode"], value="NEFT")
                
                if "reference_number" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["reference_number"], value=employee.get("employee_id", ""))
                
                if "beneficiary_code" in header_mapping:
                    ws.cell(row=current_row, column=header_mapping["beneficiary_code"], value=employee.get("name", ""))
                
                current_row += 1
        else:
            # Use standard format (existing code)
            wb = Workbook()
            ws = wb.active
            ws.title = "Bank Advice"
            
            # Style definitions
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add header information
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Bank Advice - {advice['reference_number']}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws['A2'] = f"Company Account: {account['account_name'] if account else 'N/A'}"
            ws['A3'] = f"Bank: {account['bank_name'] if account else 'N/A'}"
            ws['A4'] = f"Period: {advice['month']}/{advice['year']}"
            ws['A5'] = f"Total Amount: ₹{advice['total_amount']:,.2f}"
            ws['A6'] = f"Total Employees: {advice['employee_count']}"
            
            # Add empty row
            current_row = 8
            
            # Add table headers
            headers = ["S.No", "Employee ID", "Employee Name", "Bank Name", "Account Number", "IFSC Code", "Amount"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Add employee data
            current_row += 1
            for idx, employee in enumerate(employees, 1):
                salary_structure = employee.get("salary_structure", {})
                
                # Calculate net salary
                gross_salary = (
                    salary_structure.get("basic_salary", 0) +
                    salary_structure.get("hra", 0) +
                    salary_structure.get("medical_allowance", 0) +
                    salary_structure.get("travel_allowance", 0) +
                    salary_structure.get("food_allowance", 0) +
                    salary_structure.get("internet_allowance", 0) +
                    salary_structure.get("special_allowance", 0)
                )
                
                total_deductions = (
                    salary_structure.get("pf_employee", 0) +
                    salary_structure.get("esi_employee", 0) +
                    salary_structure.get("professional_tax", 0) +
                    salary_structure.get("tds", 0)
                )
                
                net_salary = gross_salary - total_deductions
                
                bank_info = employee.get("bank_info", {})
                
                row_data = [
                    idx,
                    employee.get("employee_id", ""),
                    employee.get("name", ""),
                    bank_info.get("bank_name", ""),
                    bank_info.get("account_number", ""),
                    bank_info.get("ifsc_code", ""),
                    net_salary
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    if col == 7:  # Amount column
                        cell.number_format = '₹#,##0.00'
                
                current_row += 1
            
            # Add total row
            ws.cell(row=current_row, column=6, value="Total:").font = Font(bold=True)
            total_cell = ws.cell(row=current_row, column=7, value=advice['total_amount'])
            total_cell.font = Font(bold=True)
            total_cell.number_format = '₹#,##0.00'
            total_cell.border = border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create filename
        template_name = template["bank_name"] if template else "Standard"
        filename = f"Bank_Advice_{advice['reference_number'].replace('/', '_')}_{template_name}.xlsx"
        
        # Return as streaming response
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error downloading bank advice: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to download bank advice: {str(e)}"
        )


# ============================================================================
# SALARY COMPONENTS MANAGEMENT ENDPOINTS
# ============================================================================

@api_router.get("/salary-components")
async def get_salary_components(
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    """Get all salary components with optional category filter"""
    query = {**company_filter}
    
    if category:
        query["category"] = category
    
    components = await db.salary_components.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    
    return {
        "components": components,
        "total": len(components)
    }


@api_router.get("/salary-components/{component_id}")
async def get_salary_component(
    component_id: str,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    """Get a specific salary component"""
    component = await db.salary_components.find_one(
        {"component_id": component_id, **company_filter},
        {"_id": 0}
    )
    
    if not component:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Salary component not found"
        )
    
    return component


@api_router.post("/salary-components")
async def create_salary_component(
    component_data: SalaryComponentCreate,
    current_user: User = Depends(require_admin_or_super_admin),
    company_filter: dict = Depends(get_company_filter)
):
    """Create a new salary component"""
    # Get company_id from current user
    company_id = current_user.company_id
    if not company_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Company ID is required"
        )
    
    # Check if component with same name already exists for this company
    existing = await db.salary_components.find_one({
        "company_id": company_id,
        "category": component_data.category,
        "component_name": component_data.component_name
    })
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"A {component_data.category} component with this name already exists"
        )
    
    # Create new component
    new_component = SalaryComponent(
        **component_data.dict(),
        company_id=company_id,
        created_by=current_user.username
    )
    
    await db.salary_components.insert_one(prepare_for_mongo(new_component.dict()))
    
    return {
        "message": "Salary component created successfully",
        "component_id": new_component.component_id
    }


@api_router.put("/salary-components/{component_id}")
async def update_salary_component(
    component_id: str,
    component_data: SalaryComponentUpdate,
    current_user: User = Depends(require_admin_or_super_admin),
    company_filter: dict = Depends(get_company_filter)
):
    """Update a salary component"""
    component = await db.salary_components.find_one(
        {"component_id": component_id, **company_filter}
    )
    
    if not component:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Salary component not found"
        )
    
    # Prepare update data
    update_data = {k: v for k, v in component_data.dict(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data:
        await db.salary_components.update_one(
            {"component_id": component_id, **company_filter},
            {"$set": update_data}
        )
    
    return {"message": "Salary component updated successfully"}


@api_router.delete("/salary-components/{component_id}")
async def delete_salary_component(
    component_id: str,
    current_user: User = Depends(require_admin_or_super_admin),
    company_filter: dict = Depends(get_company_filter)
):
    """Delete a salary component"""
    component = await db.salary_components.find_one(
        {"component_id": component_id, **company_filter}
    )
    
    if not component:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Salary component not found"
        )
    
    # TODO: Add validation to check if component is used in any employee's salary structure
    # For now, allow deletion
    
    await db.salary_components.delete_one(
        {"component_id": component_id, **company_filter}
    )
    
    return {"message": "Salary component deleted successfully"}


@api_router.get("/salary-components/types/{category}")
async def get_component_types(
    category: str,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    """Get all unique component types for a category"""
    components = await db.salary_components.find(
        {"category": category, **company_filter},
        {"_id": 0, "component_type": 1}
    ).to_list(length=None)
    
    # Extract unique types
    types = list(set([c["component_type"] for c in components if c.get("component_type")]))
    
    # Add predefined types based on category
    predefined_types = {
        "earnings": ["Basic", "House Rent Allowance", "Dearness Allowance", "Conveyance Allowance", "Medical Allowance", "Bonus", "Commission", "Overtime", "Incentive"],
        "deductions": ["Provident Fund", "ESI", "Professional Tax", "TDS", "Loan Deduction", "Advance Deduction", "Late Arrival Fine"],
        "benefits": ["Health Insurance", "Life Insurance", "Food Coupons", "Transport Allowance", "Phone Reimbursement"],
        "reimbursements": ["Travel Reimbursement", "Medical Reimbursement", "Fuel Reimbursement", "Mobile Reimbursement", "Internet Reimbursement"]
    }
    
    # Combine predefined and custom types
    all_types = list(set(predefined_types.get(category, []) + types))
    all_types.sort()
    
    return {"types": all_types}


# ============================================================================
# TAX CONFIGURATION ENDPOINTS
# ============================================================================

@api_router.get("/tax-configuration")
async def get_tax_configurations(
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    """Get all tax configurations for the company"""
    configs = await db.tax_configurations.find(company_filter, {"_id": 0}).to_list(length=None)
    return {"configurations": configs}


@api_router.get("/tax-configuration/{component_type}")
async def get_tax_configuration(
    component_type: str,
    current_user: User = Depends(get_current_user),
    company_filter: dict = Depends(get_company_filter)
):
    """Get a specific tax configuration"""
    config = await db.tax_configurations.find_one(
        {"component_type": component_type, **company_filter},
        {"_id": 0}
    )
    
    if not config:
        # Return default configuration if not found
        return {
            "component_type": component_type,
            "is_enabled": False,
            "company_id": current_user.company_id
        }
    
    return config


@api_router.post("/tax-configuration")
async def create_tax_configuration(
    config_data: TaxConfigurationCreate,
    current_user: User = Depends(require_admin_or_super_admin),
    company_filter: dict = Depends(get_company_filter)
):
    """Create or update tax configuration"""
    company_id = current_user.company_id
    if not company_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Company ID is required"
        )
    
    # Check if configuration already exists
    existing = await db.tax_configurations.find_one({
        "company_id": company_id,
        "component_type": config_data.component_type
    })
    
    if existing:
        # Update existing configuration
        update_data = config_data.dict(exclude_unset=True)
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.tax_configurations.update_one(
            {"company_id": company_id, "component_type": config_data.component_type},
            {"$set": update_data}
        )
        
        return {"message": "Tax configuration updated successfully"}
    else:
        # Create new configuration
        new_config = TaxConfiguration(
            **config_data.dict(),
            company_id=company_id,
            created_by=current_user.username
        )
        
        await db.tax_configurations.insert_one(prepare_for_mongo(new_config.dict()))
        
        return {
            "message": "Tax configuration created successfully",
            "config_id": new_config.config_id
        }


@api_router.put("/tax-configuration/{component_type}")
async def update_tax_configuration(
    component_type: str,
    config_data: TaxConfigurationUpdate,
    current_user: User = Depends(require_admin_or_super_admin),
    company_filter: dict = Depends(get_company_filter)
):
    """Update tax configuration"""
    config = await db.tax_configurations.find_one(
        {"component_type": component_type, **company_filter}
    )
    
    if not config:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tax configuration not found"
        )
    
    # Prepare update data
    update_data = {k: v for k, v in config_data.dict(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data:
        await db.tax_configurations.update_one(
            {"component_type": component_type, **company_filter},
            {"$set": update_data}
        )
    
    return {"message": "Tax configuration updated successfully"}


@api_router.delete("/tax-configuration/{component_type}")
async def delete_tax_configuration(
    component_type: str,
    current_user: User = Depends(require_admin_or_super_admin),
    company_filter: dict = Depends(get_company_filter)
):
    """Delete/disable tax configuration"""
    config = await db.tax_configurations.find_one(
        {"component_type": component_type, **company_filter}
    )
    
    if not config:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tax configuration not found"
        )
    
    # Instead of deleting, just disable it
    await db.tax_configurations.update_one(
        {"component_type": component_type, **company_filter},
        {"$set": {"is_enabled": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Tax configuration disabled successfully"}


# Include the router in the main app
app.include_router(api_router)

# WebSocket endpoint for real-time notifications
@app.websocket("/ws/notifications/{user_id}")
async def websocket_endpoint(websocket: WebSocket, user_id: str):
    await manager.connect(websocket, user_id)
    try:
        while True:
            # Keep connection alive and listen for client messages
            data = await websocket.receive_text()
            # Echo back for heartbeat/ping-pong
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect(user_id)
    except Exception as e:
        logging.error(f"WebSocket error for {user_id}: {e}")
        manager.disconnect(user_id)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_db():
    print("Application startup - initializing users...")
    try:
        # Initialize default admin user
        admin_exists = await db.users.find_one({"username": "admin", "role": "admin"})
        if not admin_exists:
            admin_user = User(
                username="admin",
                email="admin@company.com",
                role=UserRole.ADMIN,
                hashed_password=get_password_hash("Admin$2022"),
                is_active=True
            )
            await db.users.insert_one(prepare_for_mongo(admin_user.dict()))
            print("Default admin user created: admin/password")
        
        # Create employee users for existing employees
        employees = await db.employees.find({}).to_list(length=None)
        for emp in employees:
            employee_user_exists = await db.users.find_one({
                "username": emp["employee_id"], 
                "role": "employee"
            })
            if not employee_user_exists:
                # Generate random PIN for employee
                default_pin = generate_random_pin()
                employee_user = User(
                    username=emp["employee_id"],
                    email=emp.get("email"),
                    role=UserRole.EMPLOYEE,
                    employee_id=emp["employee_id"],
                    pin=default_pin,
                    is_active=True
                )
                await db.users.insert_one(prepare_for_mongo(employee_user.dict()))
                print(f"Employee user created: {emp['employee_id']}/PIN:1234")
        print("User initialization completed")
    except Exception as e:
        print(f"Error during user initialization: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
